"""Analyser tab — v3.2 per-asset layout with v3.1 capital-loss netting.

Restructured in v3.0 around two design goals:

1. **See the answer immediately.** Fund summary at the top shows BOTH scenarios
   side-by-side (`If no reset (default)` / `If elected to reset` / `Difference`)
   with per-member tax breakdown. No tab navigation required.
2. **Single source of truth.** Per-member tax formulas read `band1`/`band2`
   straight from Inputs col D/E — what the user sees on Inputs IS what the
   calc uses.

v3.1 (capital-loss netting) — three changes on this sheet:

- **Div 296 fund earnings** (row 8) — formula is `MAX(0, SUM(...))` not
  `SUMIF(...,">0")`. Capital losses now net intra-year against gains;
  the net is floored at zero.
- **Per-asset Ord CGT (col G, v3.2)** — "Per-asset Ord CGT (info only)",
  greyed, no totals-row sum. It is now a STANDALONE DIAGNOSTIC VIEW, not
  the real tax. Loss rows show "—". The authoritative fund Ord CGT is in
  the Reconciliation panel and uses the s102-5 method statement.
- **Reconciliation panel** — "Fund Ordinary CGT (after intra-year netting)"
  replaces the old `SUM(F:F)` cell. Carry-forward losses cell is now
  fund-level net unused gross loss (`MAX(0, gross_losses - gross_gains)`).

Hidden recon helper cells (cols O, P, Q at row 70) drive the fund Ord CGT:

    O70  disc_gains    — sum of positive gains where held>12m ("Yes")
    P70  nond_gains    — sum of positive gains where held<>"Yes"
    Q70  gross_losses  — sum of |negative gains| (positive number)

Layout (every formula references either an Inputs cell or a named range —
never a magic number):

    Row 1     Title
    Row 2     Parameters in effect (read-only, sourced from named ranges)
    Row 3     Sample-data warning (conditional)
    Row 4     Estimate disclaimer banner (always visible)

    Row 6     Section band: "Fund summary"
    Row 7     Header row: blank | If no reset (default) | If elected | Diff
    Row 8     Fund Div 296 earnings    (3 scenario columns + diff) — v3.1: MAX(0, SUM)
    Rows 9-12 Member 1-4 Div 296 tax   (3 scenario columns + diff)
    Row 13    Total Div 296 tax (headline)  (3 scenario columns + diff)

    Row 15    Section band: "Per-asset analysis (elected-reset scenario)"
    Row 16    Column headers (A..L visible, M..N hidden helpers)
    Rows 17-66 50 data rows. Always shows the elected-reset cost base.
                  A #row-num   B Asset                  C Proceeds
                  D Original CB              E Ord gross gain (info only)
                  F 1/3 CGT discount eligible? (info)    G Per-asset Ord CGT (info)
                  H Div 296 CB              I Div 296 gross gain (info)
                  J Div 296 adj gain (post-discount)     K Div 296 tax
                  L Reset impact   M helper (with reset)    N helper (without reset)
    Row 67    Totals (greyed info-only cols omitted)
    Row 68    Footnote: the greyed info-only columns are diagnostic, not tax

    Row 70    Section band: "Reconciliation"
              (hidden recon helpers also live on this row at cols O/P/Q)
    Row 71    Fund Ordinary CGT (after intra-year netting)  ← v3.1 new formula
    Row 72    Div 296 tax payable (= elected-reset headline)
    Row 73    Capital losses carried forward  ← v3.1 fund net
    Row 74    Plain-English caption  ← v3.1 reworded

Per-asset Div 296 tax (col K) is the pro-rata of the elected-reset headline
($D$13):
    K{r} = IF(SUMIF(J17:J66,">0")=0, 0,
              MAX(0, J{r}) / SUMIF(J17:J66,">0") * $D$13)

(SUMIF(>0) is correct as the attribution DENOMINATOR — loss assets bear
$0 tax. Only the fund-EARNINGS formula at row 8 changed in v3.1.)

Hidden helper columns M, N compute both reset scenarios unconditionally
(they do not reference any toggle), so the Reset Impact column L = M - N
remains scenario-agnostic.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from div296._formulas import div296_adj_gain_formula, per_member_div296_tax_formula
from div296.assumptions import ASSUMPTIONS
from div296.styles import (
    BODY_FONT, CENTER, COLOUR_DIV296_HEADER, FMT_CURRENCY, ROWNUM_FILL, ROWNUM_FONT,
    SECTION_BAND_FILL, SECTION_BAND_FONT, STATE_STRIP_FILL, TITLE_FONT, TRAP_FILL,
    PROC_HEADER_FILL, PROC_DATA_FILL,
    ORD_HEADER_FILL, ORD_DATA_FILL,
    DIV_HEADER_FILL, DIV_DATA_FILL,
    RESET_HEADER_FILL, RESET_DATA_FILL,
    RESET_HEADER_FONT, GROUP_HEADER_FONT,
    TOTALS_FILL, TOTALS_FONT,
)
from div296.tabs.inputs import (
    MEMBERS_FIRST_DATA_ROW,
    REGISTER_FIRST_DATA_ROW,
    sample_detect_expr,
)


SHEET = "Analyser"
INPUTS_SHEET = "'Inputs'"

# --- Layout constants (v3.0 — control panel removed; side-by-side fund summary) ---
TITLE_ROW = 1
STATE_STRIP_ROW = 2                            # "Parameters in effect" line
SAMPLE_WARN_ROW = 3                            # conditional sample-data warning
ESTIMATE_BANNER_ROW = 4                        # was 9; moved up since lever rows gone

FUND_BAND_ROW = 6                              # was 11
FUND_HEADER_ROW = 7                            # NEW — scenario column headers
FUND_EARNINGS_ROW = 8                          # was 12
MEMBER_TAX_FIRST_ROW = 9                       # was 13
MEMBER_TAX_LAST_ROW = MEMBER_TAX_FIRST_ROW + ASSUMPTIONS.member_count - 1   # row 12
HEADLINE_ROW = MEMBER_TAX_LAST_ROW + 1         # row 13 (was 17)

PERASSET_BAND_ROW = 15                         # was 19
PERASSET_HEADER_ROW = 16                       # was 20
PERASSET_FIRST_ROW = 17                        # was 21
PERASSET_LAST_ROW = PERASSET_FIRST_ROW + ASSUMPTIONS.asset_register_rows - 1   # row 66
TOTALS_ROW = PERASSET_LAST_ROW + 1             # row 67

# v3.1: col F footnote (Per-asset Ord CGT info only) — squeezed in before
# the trap legend; everything downstream shifts +1.
F_INFO_FOOTNOTE_ROW = TOTALS_ROW + 1           # row 68 (NEW in v3.1)
TRAP_LEGEND_ROW = TOTALS_ROW + 2               # row 69 (was 68)
RECON_BAND_ROW = TOTALS_ROW + 3                # row 70 (was 69)
RECON_ORD_CGT_ROW = RECON_BAND_ROW + 1         # row 71 (was 70)
RECON_DIV296_ROW = RECON_BAND_ROW + 2          # row 72 (was 71)
RECON_LOSSES_ROW = RECON_BAND_ROW + 3          # row 73 (was 72)
RECON_LOSSES_CAPTION_ROW = RECON_LOSSES_ROW + 1   # row 74 (was 73)

# v3.2: hidden helper cells driving the Fund Ord CGT formula.
# Placed at row 70 (= RECON_BAND_ROW) in hidden cols O, P, Q (was M, N, O
# in v3.1 — shifted +2 to make room for the new visible Disc-flag and
# Div 296 gross cols). Adjacent to the consumer at B71 but invisible.
HELPER_DISC_GAINS_COL = 15                     # col O — recon helper: discountable gains
HELPER_NOND_GAINS_COL = 16                     # col P — recon helper: non-discountable gains
HELPER_GROSS_LOSSES_COL = 17                   # col Q — recon helper: gross losses (as +ve)

# Inputs↔Analyser offset (Inputs row = Analyser row - OFFSET)
ROW_OFFSET = PERASSET_FIRST_ROW - REGISTER_FIRST_DATA_ROW   # 17 - 16 = 1

# --- Fund summary column layout (v3.0 — side-by-side scenarios) ---
# A & B merged for label; C/D/E for scenario data.
FUND_LABEL_COL_RANGE = ("A", "B")              # merged for labels
FUND_NORESET_COL = 3                           # col C
FUND_ELECTED_COL = 4                           # col D
FUND_DIFF_COL = 5                              # col E

# --- Per-asset column constants (v3.2 Option B layout) ---
# Visible cols A..L (12 cols); hidden helpers M..Q (5 cols).
ROWNUM_COL = 1                                 # col A
ASSET_COL = 2                                  # col B
PROCEEDS_COL = 3                               # col C
ORIG_CB_COL = 4                                # col D
ORD_GAIN_COL = 5                               # col E — Ord GROSS gain (v3.2: semantics changed from "post-disc" to "gross")
DISC_FLAG_COL = 6                              # col F — NEW "1/3 CGT discount eligible? (Yes/No)" flag
ORD_CGT_COL = 7                                # col G — Per-asset Ord CGT (info-only, derived from E + F)
DIV_CB_COL = 8                                 # col H — Div 296 cost base
DIV_GROSS_COL = 9                              # col I — NEW Div 296 GROSS gain (proceeds − MV)
DIV_GAIN_COL = 10                              # col J — Per-asset Div 296 post-disc gain (info-only; semantics unchanged from v3.1, position shifted from H)
DIV_TAX_COL = 11                               # col K — Div 296 tax (pro-rata of headline)
RESET_IMPACT_COL = 12                          # col L — Reset impact = with-reset − no-reset
HELPER_WITH_RESET_COL = 13                     # col M — Div 296 gain WITH reset (hidden)
HELPER_NORESET_COL = 14                        # col N — Div 296 gain WITHOUT reset (hidden)
LAST_VISIBLE_COL_LETTER = "L"

# Hidden helper columns (per-asset scenario helpers M/N + recon-row helpers
# O/P/Q), derived from the constants above so a column shift can't strand the
# tuple (audit 2026-06-10: was a hardcoded ("M","N","O","P","Q") literal).
HELPER_HIDDEN_COLS = tuple(
    get_column_letter(c)
    for c in (HELPER_WITH_RESET_COL, HELPER_NORESET_COL,
              HELPER_DISC_GAINS_COL, HELPER_NOND_GAINS_COL, HELPER_GROSS_LOSSES_COL)
)


# --- Cross-tab cell-address constants (v3.2 slice 1) ---
# Surfaced so other tabs (Comparison) reference these instead of hardcoding
# literal cell addresses. Derived from the row/col constants above so they
# auto-track future layout shifts.
FUND_ORD_CGT_CELL = f"B{RECON_ORD_CGT_ROW}"                                    # "B71"
HEADLINE_NORESET_CELL = f"{get_column_letter(FUND_NORESET_COL)}{HEADLINE_ROW}"  # "C13"
HEADLINE_ELECTED_CELL = f"{get_column_letter(FUND_ELECTED_COL)}{HEADLINE_ROW}"  # "D13"


def _band(ws: Worksheet, row: int, text: str, last_col_letter: str = "L") -> None:
    ws.cell(row=row, column=1, value=text).font = SECTION_BAND_FONT
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    for col_idx in range(1, ord(last_col_letter) - ord("A") + 2):
        ws.cell(row=row, column=col_idx).fill = SECTION_BAND_FILL


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)

    # --- Title ---
    ws.cell(row=TITLE_ROW, column=1,
            value="Division 296 Cost Base Reset Model — Analyser").font = TITLE_FONT
    ws.merge_cells(f"A{TITLE_ROW}:L{TITLE_ROW}")

    # --- Row 2: Parameters in effect (v3.0 — replaces v2.x toggle-narration state strip) ---
    state_formula = (
        '="Parameters in effect: Rates "&TEXT(rate_tier1,"0%")&" tier 1 / "'
        '&TEXT(rate_tier2,"0%")&" tier 2 · Thresholds "'
        '&TEXT(threshold_1,"$#,##0")&" / "&TEXT(threshold_2,"$#,##0")'
        '&" · CGT discount "&TEXT(discount_rate,"0.00%")'
        '&" · Fund CGT "&TEXT(fund_cgt_rate,"0%")'
    )
    state_cell = ws.cell(row=STATE_STRIP_ROW, column=1, value=state_formula)
    state_cell.font = Font(name="Arial", size=9, italic=True, color="555555")
    state_cell.fill = STATE_STRIP_FILL
    state_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A{STATE_STRIP_ROW}:L{STATE_STRIP_ROW}")
    ws.row_dimensions[STATE_STRIP_ROW].height = 20

    # --- Row 3: Sample-data warning (conditional) ---
    sample_detect = sample_detect_expr(f"{INPUTS_SHEET}!")
    badge = ws.cell(
        row=SAMPLE_WARN_ROW, column=1,
        value=(
            f'=IF({sample_detect},'
            '"⚠  Sample data detected — figures below are illustrative only '
            'until the asset register on Inputs is replaced with the actual '
            'fund\'s holdings.","")'
        ),
    )
    badge.font = Font(name="Arial", size=10, bold=True, italic=True, color="8A6D00")
    badge.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A{SAMPLE_WARN_ROW}:L{SAMPLE_WARN_ROW}")
    ws.row_dimensions[SAMPLE_WARN_ROW].height = 20
    sample_amber_rule = FormulaRule(
        formula=[sample_detect],
        fill=PatternFill("solid", fgColor="FFF4CE"),
    )
    ws.conditional_formatting.add(f"A{SAMPLE_WARN_ROW}:L{SAMPLE_WARN_ROW}", sample_amber_rule)

    # --- Row 4: Estimate disclaimer banner (always visible) ---
    est = ws.cell(
        row=ESTIMATE_BANNER_ROW, column=1,
        value=("Important assumption: the figures shown in this analysis are "
               "estimates only and are based on the information entered on the "
               "Inputs tab. Actual Division 296 outcomes may differ depending "
               "on final taxable income, asset values, member balances and "
               "other relevant adjustments."),
    )
    est.font = Font(name="Arial", size=10, italic=True, color="8A6D00")
    est.fill = PatternFill("solid", fgColor="FFF8E1")  # very pale amber
    est.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    est.border = Border(
        top=Side(style="thin", color="E0C77A"),
        bottom=Side(style="thin", color="E0C77A"),
    )
    ws.merge_cells(f"A{ESTIMATE_BANNER_ROW}:L{ESTIMATE_BANNER_ROW}")
    ws.row_dimensions[ESTIMATE_BANNER_ROW].height = 42

    # --- Fund summary block (v3.0 — side-by-side scenarios + signed Difference) ---
    _band(ws, FUND_BAND_ROW, "Fund summary")

    # Helper col ranges for the per-asset gain columns we SUMIF over.
    # v3.2: variable names track semantics, not column letters. The post-disc
    # Div 296 col moved from H to J; the no-reset hidden helper moved from
    # L to N (the +2 shift comes from the new visible Disc-flag and Div 296
    # gross cols inserted upstream).
    postdisc_col = get_column_letter(DIV_GAIN_COL)          # "J"
    noreset_col = get_column_letter(HELPER_NORESET_COL)            # "N"
    postdisc_first = f"{postdisc_col}{PERASSET_FIRST_ROW}"
    postdisc_last = f"{postdisc_col}{PERASSET_LAST_ROW}"
    noreset_first = f"{noreset_col}{PERASSET_FIRST_ROW}"
    noreset_last = f"{noreset_col}{PERASSET_LAST_ROW}"

    # --- Row 7: Header row ---
    fund_headers = [
        (1, "", None),                                  # cols A:B label zone
        (FUND_NORESET_COL, "If no reset (default)", COLOUR_DIV296_HEADER),  # green-ish
        (FUND_ELECTED_COL, "If elected to reset", "C7A752"),      # gold (reset theme)
        (FUND_DIFF_COL, "Difference", "1D3B34"),                  # dark teal
    ]
    # Merge label cols A:B on header row too (keeps the band visually clean).
    ws.merge_cells(
        f"{FUND_LABEL_COL_RANGE[0]}{FUND_HEADER_ROW}:"
        f"{FUND_LABEL_COL_RANGE[1]}{FUND_HEADER_ROW}"
    )
    for col_idx, label, fg in fund_headers[1:]:
        c = ws.cell(row=FUND_HEADER_ROW, column=col_idx, value=label)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[FUND_HEADER_ROW].height = 30

    # --- Row 8: Fund Div 296 earnings ---
    earnings_label_range = (
        f"{FUND_LABEL_COL_RANGE[0]}{FUND_EARNINGS_ROW}:"
        f"{FUND_LABEL_COL_RANGE[1]}{FUND_EARNINGS_ROW}"
    )
    ws.merge_cells(earnings_label_range)
    elabel = ws.cell(row=FUND_EARNINGS_ROW, column=1, value="Fund Div 296 earnings")
    elabel.font = BODY_FONT
    elabel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # v3.1: No-reset earnings = MAX(0, SUM(no-reset helper col)) — intra-year
    # netting of gains and losses, floored at zero. Was SUMIF(>0) in v3.0
    # (per-asset floor, no netting). v3.2: helper moved from col L to col N.
    e_noreset = ws.cell(
        row=FUND_EARNINGS_ROW, column=FUND_NORESET_COL,
        value=f'=MAX(0, SUM({noreset_first}:{noreset_last}))',
    )
    e_noreset.number_format = FMT_CURRENCY
    e_noreset.comment = Comment(
        ("Sum of Div 296 adjusted gains and losses across all 50 register "
         "rows (intra-year netting), floored at zero — Div 296 earnings "
         "cannot be negative. Uses the ORIGINAL cost base (no reset election). "
         "Mirrors the hidden no-reset helper column of the per-asset table."),
        "v3.1",
    )
    # v3.1: Elected earnings = MAX(0, SUM(post-disc Div 296 gain col)) — same
    # netting logic on the elected-reset scenario. v3.2: col moved from H to J.
    e_elected = ws.cell(
        row=FUND_EARNINGS_ROW, column=FUND_ELECTED_COL,
        value=f'=MAX(0, SUM({postdisc_first}:{postdisc_last}))',
    )
    e_elected.number_format = FMT_CURRENCY
    e_elected.comment = Comment(
        ("Sum of Div 296 adjusted gains and losses across all 50 register "
         "rows (intra-year netting), floored at zero — Div 296 earnings "
         "cannot be negative. Uses the MARKET VALUE at 30 Jun 2026 as cost "
         "base (election made). Mirrors the per-asset Div 296 post-discount "
         "column (info-only diagnostic) of the per-asset table."),
        "v3.1",
    )
    # Difference (signed) = elected - no_reset.
    e_diff = ws.cell(
        row=FUND_EARNINGS_ROW, column=FUND_DIFF_COL,
        value=f"=D{FUND_EARNINGS_ROW}-C{FUND_EARNINGS_ROW}",
    )
    e_diff.number_format = '$#,##0;($#,##0);"-"'

    # --- Rows 9-12: Per-member Div 296 tax (3 scenario columns + signed diff) ---
    # Cell references for the earnings driver of each scenario.
    earnings_cell_noreset = f"$C${FUND_EARNINGS_ROW}"
    earnings_cell_elected = f"$D${FUND_EARNINGS_ROW}"
    for i in range(ASSUMPTIONS.member_count):
        row = MEMBER_TAX_FIRST_ROW + i
        inputs_row = MEMBERS_FIRST_DATA_ROW + i
        # Label (merged A:B)
        ws.merge_cells(f"A{row}:B{row}")
        lbl = ws.cell(row=row, column=1, value=f"Member {i+1} Div 296 tax")
        lbl.font = BODY_FONT
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # No-reset scenario
        t_no = ws.cell(
            row=row, column=FUND_NORESET_COL,
            value=per_member_div296_tax_formula(inputs_row, earnings_cell_noreset),
        )
        t_no.number_format = FMT_CURRENCY
        # Elected scenario
        t_el = ws.cell(
            row=row, column=FUND_ELECTED_COL,
            value=per_member_div296_tax_formula(inputs_row, earnings_cell_elected),
        )
        t_el.number_format = FMT_CURRENCY
        # Difference (signed)
        t_df = ws.cell(
            row=row, column=FUND_DIFF_COL,
            value=f"=D{row}-C{row}",
        )
        t_df.number_format = '$#,##0;($#,##0);"-"'

    # --- Row 13: Headline total ---
    ws.merge_cells(f"A{HEADLINE_ROW}:B{HEADLINE_ROW}")
    hlbl = ws.cell(row=HEADLINE_ROW, column=1, value="Total Div 296 tax (headline)")
    hlbl.font = Font(name="Arial", size=11, bold=True, color="1D3B34")
    hlbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hlbl.fill = PatternFill("solid", fgColor="EFF5F3")
    for col_idx in (FUND_NORESET_COL, FUND_ELECTED_COL, FUND_DIFF_COL):
        col_letter = get_column_letter(col_idx)
        h_total = ws.cell(
            row=HEADLINE_ROW, column=col_idx,
            value=f"=SUM({col_letter}{MEMBER_TAX_FIRST_ROW}:{col_letter}{MEMBER_TAX_LAST_ROW})",
        )
        h_total.font = Font(name="Arial", size=11, bold=True, color="1D3B34")
        h_total.fill = PatternFill("solid", fgColor="EFF5F3")
        if col_idx == FUND_DIFF_COL:
            h_total.number_format = '$#,##0;($#,##0);"-"'
        else:
            h_total.number_format = FMT_CURRENCY
        h_total.border = Border(top=Side(style="thin", color="1D3B34"))

    # Sign-coloured Difference column (negative = saving from electing reset → green;
    # positive = cost increase → red). Matches the Reset-impact convention used on
    # the per-asset table below. Muted RGB to sit visually with the rest of the workbook.
    # Two ranges so CF font weight/size matches the base style of each band.
    diff_letter = get_column_letter(FUND_DIFF_COL)
    body_range = (
        f"{diff_letter}{FUND_EARNINGS_ROW}:{diff_letter}{MEMBER_TAX_LAST_ROW}"
    )
    head_range = f"{diff_letter}{HEADLINE_ROW}:{diff_letter}{HEADLINE_ROW}"
    for rng, size, bold in ((body_range, 10, False), (head_range, 11, True)):
        top = int(rng.split(":")[0][1:])
        fav = FormulaRule(
            formula=[f"AND(ISNUMBER({diff_letter}{top}),{diff_letter}{top}<0)"],
            font=Font(name="Arial", size=size, bold=bold, color="0B6E4F"),
        )
        unfav = FormulaRule(
            formula=[f"AND(ISNUMBER({diff_letter}{top}),{diff_letter}{top}>0)"],
            font=Font(name="Arial", size=size, bold=bold, color="A61B1B"),
        )
        ws.conditional_formatting.add(rng, fav)
        ws.conditional_formatting.add(rng, unfav)

    # --- Per-asset analysis (elected-reset scenario only) ---
    _band(ws, PERASSET_BAND_ROW, "Per-asset analysis (elected-reset scenario)")
    # v3.2: Option B symmetric layout — gross + flag + derived CGT on both
    # ord and Div 296 sides. 12 visible cols (A..L), 5 hidden (M..Q).
    headers = [
        "#",
        "Asset", "Projected sale proceeds", "Original cost base",
        "Ord gross gain (info only)",
        "1/3 CGT discount eligible? (Yes/No)",
        "Per-asset Ord CGT (info only)",
        "Div 296 cost base",
        "Div 296 gross gain (info only)",
        "Per-asset Div 296 gain (post-discount where eligible, info only)",
        "Div 296 tax",
        "Reset impact",
        "Helper: Div 296 gain with reset",
        "Helper: Div 296 gain without reset",
    ]
    # v3.1: greyed styling signals "info only — not the real tax".
    F_INFO_HEADER_FILL = PatternFill("solid", fgColor="F4F4F4")
    F_INFO_HEADER_FONT = Font(name="Arial", size=10, bold=True, italic=True, color="888888")
    HEADER_STYLE = {
        ROWNUM_COL:       (SECTION_BAND_FILL,  SECTION_BAND_FONT),
        ASSET_COL:        (SECTION_BAND_FILL,  SECTION_BAND_FONT),
        PROCEEDS_COL:     (PROC_HEADER_FILL,   GROUP_HEADER_FONT),
        ORIG_CB_COL:      (ORD_HEADER_FILL,    GROUP_HEADER_FONT),
        ORD_GAIN_COL:     (F_INFO_HEADER_FILL, F_INFO_HEADER_FONT),   # v3.2: greyed (info only — gross)
        DISC_FLAG_COL:    (F_INFO_HEADER_FILL, F_INFO_HEADER_FONT),   # v3.2: greyed (info only — flag)
        ORD_CGT_COL:      (F_INFO_HEADER_FILL, F_INFO_HEADER_FONT),   # v3.1: greyed (info only — derived CGT)
        DIV_CB_COL:       (DIV_HEADER_FILL,    GROUP_HEADER_FONT),
        DIV_GROSS_COL:    (F_INFO_HEADER_FILL, F_INFO_HEADER_FONT),   # v3.2: greyed (info only — gross)
        DIV_GAIN_COL:     (F_INFO_HEADER_FILL, F_INFO_HEADER_FONT),   # v3.1.1: greyed (info only — post-disc)
        DIV_TAX_COL:      (DIV_HEADER_FILL,    GROUP_HEADER_FONT),
        RESET_IMPACT_COL: (RESET_HEADER_FILL,  RESET_HEADER_FONT),
        HELPER_WITH_RESET_COL:     (SECTION_BAND_FILL,  SECTION_BAND_FONT),
        HELPER_NORESET_COL:     (SECTION_BAND_FILL,  SECTION_BAND_FONT),
    }
    # v3.1: data-cell styling for greyed "info only" cells.
    F_INFO_DATA_FILL = PatternFill("solid", fgColor="FAFAFA")
    F_INFO_DATA_FONT = Font(name="Arial", size=10, italic=True, color="888888")
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=PERASSET_HEADER_ROW, column=col_idx, value=header)
        fill, font = HEADER_STYLE[col_idx]
        c.font = font
        c.fill = fill
        c.alignment = CENTER

    # v3.2: hide helper cols M, N (Div 296 with/without reset). Cols K, L
    # are now VISIBLE (Div 296 tax, Reset impact respectively).
    for col_letter in ("M", "N"):
        ws.column_dimensions[col_letter].hidden = True

    # v3.2: DISC_FLAG_COL is text (Yes/No), not currency. Excluded from
    # currency_cols.
    currency_cols = [PROCEEDS_COL, ORIG_CB_COL, ORD_GAIN_COL, ORD_CGT_COL,
                     DIV_CB_COL, DIV_GROSS_COL, DIV_GAIN_COL, DIV_TAX_COL,
                     RESET_IMPACT_COL, HELPER_WITH_RESET_COL, HELPER_NORESET_COL]

    DATA_FILL = {
        PROCEEDS_COL:     PROC_DATA_FILL,
        ORIG_CB_COL:      ORD_DATA_FILL,
        ORD_GAIN_COL:     F_INFO_DATA_FILL,   # v3.2: greyed (info only — gross)
        DISC_FLAG_COL:    F_INFO_DATA_FILL,   # v3.2: greyed (info only — flag)
        ORD_CGT_COL:      F_INFO_DATA_FILL,   # v3.1: greyed (info only — derived CGT)
        DIV_CB_COL:       DIV_DATA_FILL,
        DIV_GROSS_COL:    F_INFO_DATA_FILL,   # v3.2: greyed (info only — gross)
        DIV_GAIN_COL:     F_INFO_DATA_FILL,   # v3.1.1: greyed (info only — post-disc)
        DIV_TAX_COL:      DIV_DATA_FILL,
        RESET_IMPACT_COL: RESET_DATA_FILL,
    }

    for offset in range(ASSUMPTIONS.asset_register_rows):
        a_row = PERASSET_FIRST_ROW + offset
        i_row = a_row - ROW_OFFSET

        # Inputs column layout (v2.3): A code / B name / C orig CB / D MV today /
        # E MV 30 Jun / F val source / G proceeds / H projected G/L / I held>12m
        # (raw, user-visible) / J held>12m (hidden, paste-normalised — v3.1.2).
        code = f"{INPUTS_SHEET}!A{i_row}"
        name = f"{INPUTS_SHEET}!B{i_row}"
        orig = f"{INPUTS_SHEET}!C{i_row}"
        mv = f"{INPUTS_SHEET}!E{i_row}"
        proceeds = f"{INPUTS_SHEET}!G{i_row}"
        held = f"{INPUTS_SHEET}!J{i_row}"

        # v3.2: column letters derived from constants for cross-cell refs
        # within this row's formulas.
        e_letter = get_column_letter(ORD_GAIN_COL)            # "E"
        f_letter = get_column_letter(DISC_FLAG_COL)           # "F"
        helper_with_reset_letter = get_column_letter(HELPER_WITH_RESET_COL)     # "M"
        helper_without_reset_letter = get_column_letter(HELPER_NORESET_COL)  # "N"
        postdisc_letter = get_column_letter(DIV_GAIN_COL)     # "J"

        # Col A — Row number
        rn = ws.cell(row=a_row, column=ROWNUM_COL, value=offset + 1)
        rn.fill = ROWNUM_FILL
        rn.font = ROWNUM_FONT
        rn.alignment = Alignment(horizontal="center", vertical="center")
        # Col B — Asset display
        ws.cell(row=a_row, column=ASSET_COL,
                value=f'=IF({code}="","",{code}&" - "&{name})')
        # Col C — Proceeds
        ws.cell(row=a_row, column=PROCEEDS_COL, value=f'=IF({proceeds}="","",{proceeds})')
        # Col D — Original cost base
        ws.cell(row=a_row, column=ORIG_CB_COL, value=f'=IF({orig}="","",{orig})')
        # Col E — Ord GROSS gain (v3.2: semantics changed from "post-disc").
        # Plain proceeds − orig CB; losses pass through as negatives; no
        # discount applied here. The discount factor lives in the col G
        # derivation downstream.
        e_cell = ws.cell(
            row=a_row, column=ORD_GAIN_COL,
            value=f'=IF(OR({proceeds}="",{orig}=""),"",{proceeds}-{orig})',
        )
        e_cell.font = F_INFO_DATA_FONT
        e_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Col F — NEW "1/3 CGT discount eligible? (Yes/No)" flag. Mirrors
        # Inputs!J (the paste-normalised held>12m flag — see inputs.py for
        # rationale). The col G derivation reads this to decide whether to
        # apply the 1/3 discount factor.
        f_cell = ws.cell(
            row=a_row, column=DISC_FLAG_COL,
            value=f'=IF({proceeds}="","",{held})',
        )
        f_cell.font = F_INFO_DATA_FONT
        f_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Col G — Per-asset Ord CGT (info only, derived from gross + flag).
        # Shows "—" for loss rows (gross E<=0). For positive gains, applies
        # the 1/3 discount when F="Yes", then the fund CGT rate. Result is
        # numerically identical to the v3.1 derivation (which lived at col F
        # and worked from a pre-discounted col E).
        g_cell = ws.cell(
            row=a_row, column=ORD_CGT_COL,
            value=(
                f'=IF({e_letter}{a_row}="","",'
                f'IF({e_letter}{a_row}<=0,"—",'
                f'{e_letter}{a_row}*(1-IF({f_letter}{a_row}="Yes",discount_rate,0))'
                f'*fund_cgt_rate))'
            ),
        )
        g_cell.font = F_INFO_DATA_FONT
        g_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Col H — Div 296 cost base. v3.0: always the elected reset cost base.
        ws.cell(
            row=a_row, column=DIV_CB_COL,
            value=f'=IF(OR({proceeds}="",{mv}=""),"",{mv})',
        )
        # Col I — NEW Div 296 GROSS gain. Symmetric to col E: plain
        # proceeds − Div 296 CB (MV). No discount applied; for diagnostic
        # parity with the ord side.
        i_cell = ws.cell(
            row=a_row, column=DIV_GROSS_COL,
            value=f'=IF(OR({proceeds}="",{mv}=""),"",{proceeds}-{mv})',
        )
        i_cell.font = F_INFO_DATA_FONT
        i_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Col J — Per-asset Div 296 gain (post-discount where eligible, info
        # only). Semantics unchanged from v3.1's col H; position shifted right.
        # The fund Div 296 earnings figure (row 8) sums this col with
        # intra-year netting; per-asset Div 296 tax (col K) attributes the
        # fund headline over this col's positive entries.
        j_cell = ws.cell(
            row=a_row, column=DIV_GAIN_COL,
            value=div296_adj_gain_formula(proceeds, mv, held),
        )
        j_cell.font = F_INFO_DATA_FONT
        j_cell.alignment = Alignment(horizontal="right", vertical="center")
        # Col K — Div 296 tax (pro-rata of the elected-reset headline at
        # $D${HEADLINE_ROW}). Denominator SUMIFs the post-disc col J — loss
        # assets bear $0 attribution, even though they reduce the headline
        # at fund level (s102-5 / Div 296 earnings netting at row 8).
        ws.cell(
            row=a_row, column=DIV_TAX_COL,
            value=(
                f'=IF({postdisc_letter}{a_row}="","",'
                f'IF(SUMIF({postdisc_first}:{postdisc_last},">0")=0,0,'
                f'MAX(0,{postdisc_letter}{a_row})/'
                f'SUMIF({postdisc_first}:{postdisc_last},">0")*$D${HEADLINE_ROW}))'
            ),
        )
        # Helper M — Div 296 gain WITH reset (cost base = MV). Unconditional;
        # mirrors col J under the elected-reset scenario.
        ws.cell(row=a_row, column=HELPER_WITH_RESET_COL,
                value=div296_adj_gain_formula(proceeds, mv, held))
        # Helper N — Div 296 gain WITHOUT reset (cost base = orig). Unconditional.
        ws.cell(row=a_row, column=HELPER_NORESET_COL,
                value=div296_adj_gain_formula(proceeds, orig, held))
        # Col L — Reset impact = M − N. Highlights the per-asset Δ of the
        # reset election.
        ws.cell(
            row=a_row, column=RESET_IMPACT_COL,
            value=(
                f'=IF(OR({helper_with_reset_letter}{a_row}="",{helper_without_reset_letter}{a_row}=""),"",'
                f'{helper_with_reset_letter}{a_row}-{helper_without_reset_letter}{a_row})'
            ),
        )

        for col_idx in currency_cols:
            cell = ws.cell(row=a_row, column=col_idx)
            cell.number_format = FMT_CURRENCY
            if col_idx in DATA_FILL:
                cell.fill = DATA_FILL[col_idx]

        # Thin borders on every visible per-asset cell; medium-weight left border
        # on Reset Impact (col L in v3.2; was col J in v3.1) to highlight it.
        thin = Side(style="thin", color="D5DEDA")
        heavy = Side(style="medium", color="C29A3B")
        for col_idx in range(1, RESET_IMPACT_COL + 1):
            cell = ws.cell(row=a_row, column=col_idx)
            cell.border = Border(
                left=heavy if col_idx == RESET_IMPACT_COL else thin,
                right=thin, top=thin, bottom=thin,
            )
        ws.cell(row=a_row, column=RESET_IMPACT_COL).font = Font(
            name="Arial", size=10, bold=True, color="1D3B34",
        )

    # --- Totals row ---
    DOUBLE_TOP = Border(top=Side(style="double", color="1D3B34"))

    sigma_cell = ws.cell(row=TOTALS_ROW, column=ROWNUM_COL, value="Total")
    sigma_cell.fill = TOTALS_FILL
    sigma_cell.font = TOTALS_FONT
    sigma_cell.alignment = Alignment(horizontal="center", vertical="center")
    sigma_cell.border = DOUBLE_TOP

    totals_label = ws.cell(row=TOTALS_ROW, column=ASSET_COL, value="TOTAL")
    totals_label.font = TOTALS_FONT
    totals_label.fill = TOTALS_FILL
    totals_label.border = DOUBLE_TOP

    # v3.2: only proceeds (C) and Div 296 tax (K — shifted from I in v3.1)
    # carry SUM totals. The four info-only gain/CGT cols (E ord gross,
    # G ord CGT, I div296 gross, J div296 post-disc) are NOT summed —
    # per-asset post-discount values don't aggregate meaningfully once
    # fund-level loss netting is in play (s102-5).
    sum_total_cols = (
        (get_column_letter(PROCEEDS_COL), PROCEEDS_COL),    # ("C", 3)
        (get_column_letter(DIV_TAX_COL), DIV_TAX_COL),      # ("K", 11) — shifted from I
    )
    for col_letter, col_const in sum_total_cols:
        rng = f"{col_letter}{PERASSET_FIRST_ROW}:{col_letter}{PERASSET_LAST_ROW}"
        cell = ws.cell(row=TOTALS_ROW, column=col_const, value=f"=SUM({rng})")
        cell.number_format = FMT_CURRENCY
        cell.font = TOTALS_FONT
        cell.fill = TOTALS_FILL
        cell.border = DOUBLE_TOP

    # v3.2: four info-only cols — ord gross (E), per-asset Ord CGT (G),
    # Div 296 gross (I), per-asset Div 296 post-disc (J) — all show
    # "(see fund total)" in grey italic.
    info_only_cols = (ORD_GAIN_COL, ORD_CGT_COL, DIV_GROSS_COL, DIV_GAIN_COL)
    for info_only_col in info_only_cols:
        info_cell = ws.cell(
            row=TOTALS_ROW, column=info_only_col,
            value="(see fund total)",
        )
        info_cell.font = F_INFO_DATA_FONT
        info_cell.fill = TOTALS_FILL
        info_cell.border = DOUBLE_TOP
        info_cell.alignment = Alignment(horizontal="right", vertical="center")

    # v3.2: blank cells with totals-row styling for cols that hold no
    # meaningful aggregate (D orig CB, F flag, H Div 296 CB, L reset impact).
    for col_const in (ORIG_CB_COL, DISC_FLAG_COL, DIV_CB_COL, RESET_IMPACT_COL):
        c = ws.cell(row=TOTALS_ROW, column=col_const)
        c.fill = TOTALS_FILL
        c.border = DOUBLE_TOP

    # --- v3.2: footnote covering all four "info only" gain/CGT columns ---
    f_footnote = ws.cell(
        row=F_INFO_FOOTNOTE_ROW, column=1,
        value=('The greyed "(info only)" columns (Ord gross gain, Per-asset Ord '
               'CGT, Div 296 gross gain, Per-asset Div 296 gain) show each asset '
               'on a standalone basis: the 1/3 CGT discount is applied per-asset '
               'where eligible (s115-100 ITAA 1997 — held > 12 months); losses '
               'and short-held gains pass through gross. These columns do NOT '
               'sum to the fund total because capital losses offset capital gains '
               'within the income year at the fund level (s102-5 ITAA 1997). For '
               'the authoritative fund figures see the Reconciliation panel below '
               '("Fund Ordinary CGT (after intra-year netting)" and "Div 296 '
               'earnings").'),
    )
    f_footnote.font = Font(name="Arial", size=9, italic=True, color="666666")
    f_footnote.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
    f_footnote.fill = PatternFill("solid", fgColor="F4F4F4")
    ws.merge_cells(f"A{F_INFO_FOOTNOTE_ROW}:L{F_INFO_FOOTNOTE_ROW}")
    ws.row_dimensions[F_INFO_FOOTNOTE_ROW].height = 32

    # --- Trap-row + reset-impact legend (plain-English) ---
    legend = ws.cell(
        row=TRAP_LEGEND_ROW, column=1,
        value=("Reset impact (gold column) shows the estimated change in "
               "Division 296 gain if the reset election is applied. A "
               "negative amount (shown in green) generally indicates a "
               "favourable impact, while a positive amount (shown in red) "
               "may indicate additional Division 296 exposure. Particular "
               "care should be taken with red-shaded rows, where a reset "
               "creates taxable Division 296 gain on an asset currently in "
               "an unrealised loss position (the 'reset trap')."),
    )
    legend.font = Font(name="Arial", size=9, italic=True, color="666666")
    legend.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True, indent=1)
    swatch = ws.cell(row=TRAP_LEGEND_ROW, column=1)
    swatch.fill = PatternFill("solid", fgColor="FBE9E9")
    ws.merge_cells(f"A{TRAP_LEGEND_ROW}:L{TRAP_LEGEND_ROW}")
    ws.row_dimensions[TRAP_LEGEND_ROW].height = 20

    # --- Reconciliation panel (stays at bottom per v3.0 decision 8) ---
    _band(ws, RECON_BAND_ROW, "Reconciliation")

    # v3.1: hidden helper cells driving the new Fund Ord CGT formula.
    # Ranges: Inputs!H16:H65 = projected gain/loss (= proceeds − orig cost base).
    #         Inputs!J16:J65 = NORMALISED held>12m flag ("Yes"/"No"/blank).
    # SUMIFS ignores text in the sum_range automatically; rows where Inputs!H
    # is "" (the empty-row IF fallback) are skipped.
    inputs_h_range = (
        f"{INPUTS_SHEET}!H{REGISTER_FIRST_DATA_ROW}:"
        f"H{REGISTER_FIRST_DATA_ROW + ASSUMPTIONS.asset_register_rows - 1}"
    )
    # v3.1.2: read from Inputs!J (paste-normalised), not Inputs!I (raw user
    # input). Inputs!I has a Yes/No dropdown but its DataValidation only
    # catches direct typing — pastes from another sheet/CSV bypass it.
    # Inputs!J runs TRIM+UPPER on I and returns a clean "Yes"/"No"/"".
    # See inputs.py for the rationale.
    inputs_j_range = (
        f"{INPUTS_SHEET}!J{REGISTER_FIRST_DATA_ROW}:"
        f"J{REGISTER_FIRST_DATA_ROW + ASSUMPTIONS.asset_register_rows - 1}"
    )
    helper_disc_gains = (
        f'=SUMIFS({inputs_h_range}, {inputs_h_range}, ">0", '
        f'{inputs_j_range}, "Yes")'
    )
    helper_nond_gains = (
        f'=SUMIFS({inputs_h_range}, {inputs_h_range}, ">0", '
        f'{inputs_j_range}, "<>Yes")'
    )
    helper_gross_losses = f'=-SUMIF({inputs_h_range}, "<0")'
    helper_disc_cell = ws.cell(
        row=RECON_BAND_ROW, column=HELPER_DISC_GAINS_COL, value=helper_disc_gains,
    )
    helper_disc_cell.number_format = FMT_CURRENCY
    helper_nond_cell = ws.cell(
        row=RECON_BAND_ROW, column=HELPER_NOND_GAINS_COL, value=helper_nond_gains,
    )
    helper_nond_cell.number_format = FMT_CURRENCY
    helper_loss_cell = ws.cell(
        row=RECON_BAND_ROW, column=HELPER_GROSS_LOSSES_COL, value=helper_gross_losses,
    )
    helper_loss_cell.number_format = FMT_CURRENCY
    # Hide the helper cols.
    for col in HELPER_HIDDEN_COLS:
        ws.column_dimensions[col].hidden = True

    # Fund Ordinary CGT — s102-5 method statement, losses to non-discount
    # gains first (taxpayer-favourable; standard SMSF practice).
    #   nd_after  = MAX(0, N - L)
    #   d_after   = MAX(0, D - MAX(0, L - N))
    #   net_taxable = nd_after + d_after * (1 - discount_rate)
    #   CGT = net_taxable * fund_cgt_rate
    # v3.2: helper cells moved from M70/N70/O70 to O70/P70/Q70 (shift +2)
    # to make room for the new Disc-flag and Div 296 gross visible cols.
    d_ref = f"{get_column_letter(HELPER_DISC_GAINS_COL)}{RECON_BAND_ROW}"  # disc gains  (O70 in v3.2)
    n_ref = f"{get_column_letter(HELPER_NOND_GAINS_COL)}{RECON_BAND_ROW}"  # nondisc gains (P70)
    l_ref = f"{get_column_letter(HELPER_GROSS_LOSSES_COL)}{RECON_BAND_ROW}"  # gross losses  (Q70)
    fund_ord_cgt_formula = (
        f"=(MAX(0,{n_ref}-{l_ref})"
        f"+MAX(0,{d_ref}-MAX(0,{l_ref}-{n_ref}))*(1-discount_rate))"
        f"*fund_cgt_rate"
    )

    ord_cgt_label = ws.cell(
        row=RECON_ORD_CGT_ROW, column=1,
        value="Fund Ordinary CGT (after intra-year netting)",
    )
    ord_cgt_label.font = BODY_FONT
    fund_ord_cgt_cell = ws.cell(
        row=RECON_ORD_CGT_ROW, column=2, value=fund_ord_cgt_formula,
    )
    fund_ord_cgt_cell.number_format = FMT_CURRENCY
    fund_ord_cgt_cell.comment = Comment(
        ("Per s102-5 ITAA 1997 method statement: gross capital gains and "
         "losses are netted within the income year, then the 1/3 CGT discount "
         "is applied to the net positive long-held portion. Losses are applied "
         "to non-discount gains first (preserving the discount on long-held "
         "gains where possible) — common SMSF practice; the taxpayer may elect "
         "a different loss-application order. Consult adviser if a non-default "
         "allocation is preferred."),
        "v3.1",
    )

    ws.cell(row=RECON_DIV296_ROW, column=1,
            value="Div 296 tax payable (elected-reset headline)").font = BODY_FONT
    # v3.0: headline lives at col D row 13 (elected-reset scenario).
    ws.cell(row=RECON_DIV296_ROW, column=2,
            value=f"=D{HEADLINE_ROW}").number_format = FMT_CURRENCY

    # v3.1: Carry-forward losses = net unused gross loss at fund level
    # (MAX(0, gross_losses - gross_gains)). Was per-asset gross sum in v3.0.
    cf_formula = (
        f"=MAX(0, {l_ref} - ({d_ref} + {n_ref}))"
    )
    ws.cell(row=RECON_LOSSES_ROW, column=1,
            value="Capital losses carried forward").font = BODY_FONT
    ws.cell(
        row=RECON_LOSSES_ROW, column=2, value=cf_formula,
    ).number_format = FMT_CURRENCY

    caption = ws.cell(
        row=RECON_LOSSES_CAPTION_ROW, column=1,
        value=("Capital gains and losses are netted within the income year for "
               "both ordinary CGT (per s102-5 ITAA 1997) and Div 296. Any net "
               "unused gross loss is the carry-forward figure above, available "
               "against future realised capital gains. Neither Div 296 earnings "
               "nor net taxable capital gain can be negative."),
    )
    caption.font = Font(name="Arial", size=9, italic=True, color="666666")
    caption.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)
    ws.merge_cells(f"A{RECON_LOSSES_CAPTION_ROW}:L{RECON_LOSSES_CAPTION_ROW}")
    ws.row_dimensions[RECON_LOSSES_CAPTION_ROW].height = 32

    # --- Trap shading on per-asset rows: red when ord raw gain < 0 AND
    # Div 296 adj gain > 0. v3.2: the post-disc Div 296 gain trigger moved
    # from col H to col J.
    postdisc_letter = get_column_letter(DIV_GAIN_COL)   # "J"
    rng = (
        f"A{PERASSET_FIRST_ROW}:{LAST_VISIBLE_COL_LETTER}{PERASSET_LAST_ROW}"
    )
    trap_rule = FormulaRule(
        formula=[
            f"AND($B{PERASSET_FIRST_ROW}<>\"\","
            f"($C{PERASSET_FIRST_ROW}-$D{PERASSET_FIRST_ROW})<0,"
            f"${postdisc_letter}{PERASSET_FIRST_ROW}>0)"
        ],
        fill=TRAP_FILL,
    )
    ws.conditional_formatting.add(rng, trap_rule)

    # --- Reset Impact — green for favourable (<0), red for unfavourable (>0).
    # v3.2: Reset Impact moved from col J to col L.
    reset_letter = get_column_letter(RESET_IMPACT_COL)   # "L"
    j_range = f"{reset_letter}{PERASSET_FIRST_ROW}:{reset_letter}{PERASSET_LAST_ROW}"
    j_fav_rule = FormulaRule(
        formula=[f"AND(ISNUMBER({reset_letter}{PERASSET_FIRST_ROW}),"
                 f"{reset_letter}{PERASSET_FIRST_ROW}<0)"],
        font=Font(name="Arial", size=10, bold=True, color="0B6E4F"),
    )
    j_unfav_rule = FormulaRule(
        formula=[f"AND(ISNUMBER({reset_letter}{PERASSET_FIRST_ROW}),"
                 f"{reset_letter}{PERASSET_FIRST_ROW}>0)"],
        font=Font(name="Arial", size=10, bold=True, color="A61B1B"),
    )
    ws.conditional_formatting.add(j_range, j_fav_rule)
    ws.conditional_formatting.add(j_range, j_unfav_rule)

    # --- Column widths ---
    # v3.2: 12 visible cols (A..L). Width per col, in order:
    #   A # | B Asset | C Proceeds | D Orig CB | E Ord gross | F Disc? |
    #   G Per-asset Ord CGT | H Div 296 CB | I Div 296 gross |
    #   J Div 296 post-disc | K Div 296 tax | L Reset impact
    widths = [5, 32, 14, 16, 16, 14, 22, 14, 16, 24, 14, 14]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    # Free scroll (no freeze panes). Header row repeats on every printed page.
    ws.print_title_rows = f"{PERASSET_HEADER_ROW}:{PERASSET_HEADER_ROW}"

    # --- Print header watermark ---
    ws.oddHeader.center.text = "ILLUSTRATIVE — NOT ADVICE"
    ws.oddHeader.center.size = 28
    ws.oddHeader.center.color = "CCCCCC"
    # v3.4 audit: 12 visible cols (~201 chars) spill ~3 portrait pages — fit
    # to one page wide in landscape, let the 50-row register spill vertically.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- Sheet protection ---
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws
