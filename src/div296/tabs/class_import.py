"""CLASS Import tab (v3.3) — staging area to map a CLASS Super
'Investment Summary Report' export into the Inputs asset register.

Workflow (copy-paste staging, decisions in
docs/superpowers/specs/2026-06-01-class-import-mapping-design.md):

    1. In CLASS, generate the Investment Summary Report on a TAX COST BASE
       basis (not Accounting) and export to CSV.
    2. Clear the green PASTE ZONE first (select A7:R56, press Delete) — the
       tab ships with demo data and stale rows will transfer if not removed.
       Then paste the CLASS data rows into the green zone.
    3. The MAPPED BLOCK (right) filters + maps each row into register shape.
    4. Copy mapped-block cols A:G only (physical range T7:Z56) -> Inputs!A16
       -> Paste-Special > Values.
       (NOT col H — that is the register's own Projected gain/loss formula,
       and it is a locked cell so an accidental A:I paste is blocked.)
    5. Fill MV @ 30 Jun (E), Valuation source/date (F), Projected proceeds (G),
       Held>12m (I) by hand, and resolve any negative-cost-base flag.

Only 5 CLASS columns are read: Security Code (B), Holding Account Name (C),
G/L Class (H), Total Cost (L), Market Value (M). Foreign holdings are already
AUD-converted by CLASS.

Filter (blacklist): a paste row is DROPPED when its G/L Class contains "cash",
or its Security Code is "REASEDCGT" (the realised-CGT line), or it is blank.
Dropped rows emit blanks (no row compaction). A negative tax cost base is passed
through verbatim and flagged red for manual review (CGT-event-E4 territory).

This tab holds no cross-sheet references — it is a staging template, not a live
link into the register (decision 2).
"""

from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from div296.assumptions import ASSUMPTIONS
from div296.styles import (
    BODY_FONT, CENTER, FMT_CURRENCY, FMT_TEXT, INPUT_FILL, INPUT_FONT,
    SECTION_BAND_FILL, SECTION_BAND_FONT, THIN_BOX, TITLE_FONT,
)
from div296.tabs.inputs import REGISTER_FIRST_DATA_ROW

SHEET = "CLASS Import"

# --- Row layout ---
TITLE_ROW = 1
BASIS_BANNER_ROW = 2
HOWTO_BANNER_ROW = 3
CAPACITY_WARN_ROW = 4
ALIGN_WARN_ROW = CAPACITY_WARN_ROW + 1
HEADER_ROW = 6
FIRST_DATA_ROW = 7
PASTE_ROWS = ASSUMPTIONS.asset_register_rows   # 50 — must equal the register depth
LAST_DATA_ROW = FIRST_DATA_ROW + PASTE_ROWS - 1   # row 56

# --- Paste zone: CLASS native columns A..R (18). Columns read by the map: ---
PASTE_COL_CODE = "B"      # Security Code     -> register Asset code
PASTE_COL_NAME = "C"      # Holding Acct Name -> register Asset name
PASTE_COL_GL = "H"        # G/L Class         -> filter
PASTE_COL_COST = "L"      # Total Cost (tax)  -> register Original cost base
PASTE_COL_MV = "M"        # Market Value      -> register Current market value
PASTE_LAST_COL_IDX = 18   # R — full CLASS export width

CLASS_HEADERS = [
    "Fund Code", "Security Code", "Holding Account Name", "Feed Reference",
    "Market Type", "Asset Pool", "Asset Class", "G/L Class", "Units",
    "Average Cost", "Market Price", "Total Cost", "Market Value",
    "Unrealised Gain", "Gain/loss %", "Portfolio weight %",
    "Estimated Income", "Estimated Yield %",
]
PASTE_USED_COL_IDX = {2, 3, 8, 12, 13}  # B, C, H, L, M

# --- Mapped block: register-shaped A..I at columns T..AB, flag at AC. ---
MAP_COL_START = 20  # T
MAP_HEADERS = [
    "A  Asset code", "B  Asset name", "C  Original cost base (tax)",
    "D  Current market value (as at today)", "E  Market value at 30 Jun 2026",
    "F  Valuation source / date", "G  Projected sale proceeds",
    "H  Projected gain/loss (formula — DON'T copy)", "I  Held > 12 months?",
]
MAP_FLAG_COL_IDX = MAP_COL_START + len(MAP_HEADERS)  # AC
# 0-based offsets into the mapped block.
_POPULATED = {0: PASTE_COL_CODE, 1: PASTE_COL_NAME, 2: PASTE_COL_COST, 3: PASTE_COL_MV}
_BLANK_USER = {4, 5, 6, 8}   # E, F, G, I — blank template cells (amber)
_FORMULA_SLOT = 7            # H — register formula, do not copy (grey)

# Sample: DEMO_SMSF_2020 TAX COST BASE export (2026-06-09).
# (security_code, holding_account_name, gl_class, total_cost, market_value)
SAMPLE_ROWS = [
    ("BACCT", "Macquarie CMA", "Cash At Bank", 883309.21, 883309.21),
    ("BACCT", "USD Account", "Foreign Cash At Bank", 45836.03, 48456.72),
    ("3988", "Bank of China Ltd (HKEX:3988)", "Shares in Listed Companies - Foreign", 18415.46, 18884.81),
    ("7203", "Toyota Motor Corp (TSE:7203)", "Shares in Listed Companies - Foreign", 10103.37, 13717.39),
    ("AAPL", "Apple Inc. (NASDAQ:AAPL)", "Shares in Listed Companies - Foreign", 164332.70, 265773.03),
    ("FB", "Facebook Inc (NASDAQ:FB)", "Shares in Listed Companies - Foreign", 12214.43, 82715.30),
    ("GOOGL", "Google Inc (NASDAQ:GOOGL)", "Shares in Listed Companies - Foreign", -1772.96, 51655.63),
    ("SNAP", "Snap Inc (NYSE:SNAP)", "Shares in Listed Companies - Foreign", 24395.19, 34227.02),
    ("AKE", "Allkem Limited (ASX:AKE)", "Shares in Listed Companies", 9999.00, 9999.00),
    ("AMP", "AMP Limited (ASX:AMP)", "Shares in Listed Companies", 10625.00, 4637.50),
    ("ANZ", "ANZ Banking Group Limited (ASX:ANZ)", "Shares in Listed Companies", 75086.00, 50328.00),
    ("ASX", "ASX Limited (ASX:ASX)", "Shares in Listed Companies", 16535.00, 42690.00),
    ("BHP", "BHP Billiton Limited (ASX:BHP)", "Shares in Listed Companies", 10000.00, 358200.00),
    ("BKW", "Brickworks Limited (ASX:BKW)", "Shares in Listed Companies", 6350.00, 7915.00),
    ("CBA", "Commonwealth Bank Of Australia. (ASX:CBA)", "Shares in Listed Companies", 50000.00, 208260.00),
    ("COH", "Cochlear Limited (ASX:COH)", "Shares in Listed Companies", 12342.00, 37786.00),
    ("COL", "Coles Group Limited. (ASX:COL)", "Shares in Listed Companies", 15175.68, 35936.81),
    ("CSL", "CSL Limited (ASX:CSL)", "Shares in Listed Companies", 117550.30, 1516795.00),
    ("CYB", "Aucyber Limited (ASX:CYB)", "Shares in Listed Companies", 1646.10, 1115.20),
    ("IOO", "Ishares Global 100 ETF (ASX:IOO)", "Units In Listed Unit Trusts", 48836.00, 76650.00),
    ("LTM", "Arcadium Lithium PLC (ASX:LTM)", "Shares in Listed Companies", 9999.00, 9999.00),
    ("NAB", "National Australia Bank Limited (ASX:NAB)", "Shares in Listed Companies", 46944.50, 29917.24),
    ("NAN", "Nanosonics Limited (ASX:NAN)", "Shares in Listed Companies", 3300.00, 6820.00),
    ("NDQ", "Betashares Nasdaq 100 ETF (ASX:NDQ)", "Units In Listed Unit Trusts", 100000.00, 123650.00),
    ("ORI", "Orica Limited (ASX:ORI)", "Shares in Listed Companies", 4130.00, 3328.00),
    ("RMD", "Resmed Inc (ASX:RMD)", "Shares in Listed Companies", 500000.00, 137700.00),
    ("S32", "South32 Limited (ASX:S32)", "Shares in Listed Companies", 2227.27, 2040.00),
    ("TLS", "Telstra Corporation Limited. (ASX:TLS)", "Shares in Listed Companies", 40545.00, 26605.00),
    ("VAS", "Vanguard Australian Shares Index ETF (ASX:VAS)", "Units In Listed Unit Trusts", 119170.90, 826210.00),
    ("WBC", "Westpac Banking Corporation (ASX:WBC)", "Shares in Listed Companies", 47652.00, 29617.50),
    ("WES", "Wesfarmers Limited (ASX:WES)", "Shares in Listed Companies", 30677.82, 48999.19),
    ("FSF0046AU", "Colonial First State Personal Pension Plan - Balanced", "Managed Investments", 100000.00, 163509.19),
    ("SBC0816AU", "UBS Property Securities Fund", "Managed Investments", 279331.04, 268653.28),
    ("PD", "13 GREENHILL ROAD", "Direct Property", 1090000.00, 1090000.00),
    ("PD", "Demo Property", "Direct Property", 330150.81, 554900.00),
    ("REASEDCGT", "Reaslied CGT", "Wrap/Platform Assets", 0.0, -3081.19),
]

AMBER_FILL = PatternFill("solid", fgColor="FFF4CE")
GREY_FILL = PatternFill("solid", fgColor="EDEDED")
FORMULA_FILL = PatternFill("solid", fgColor="F2F6F4")
SLATE_FILL = PatternFill("solid", fgColor="3F4F4A")


def _drop_expr(r: int) -> str:
    """Excel boolean: TRUE when paste row r is dropped (blank / cash / realised)."""
    code = f"${PASTE_COL_CODE}{r}"
    gl = f"${PASTE_COL_GL}{r}"
    return (
        f'OR(TRIM({code})="",UPPER(TRIM({code}))="REASEDCGT",'
        f'IF({gl}="",FALSE,ISNUMBER(SEARCH("cash",{gl}))))'
    )


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False

    # --- Title ---
    ws.cell(row=TITLE_ROW, column=1, value="Division 296 Cost Base Reset Model — CLASS Import").font = TITLE_FONT

    # --- Basis-acknowledgement banner (cannot be auto-detected from the CSV) ---
    basis = ws.cell(
        row=BASIS_BANNER_ROW, column=1,
        value=("⚠  REQUIRES the Investment Summary Report exported on a TAX COST BASE basis "
               "(not Accounting Cost Base). The CSV looks identical either way — an accounting "
               "export silently overstates cost bases for trusts/ETFs/managed funds."),
    )
    basis.font = Font(name="Arial", size=10, bold=True, color="8A6D00")
    basis.fill = AMBER_FILL
    basis.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=BASIS_BANNER_ROW, start_column=1, end_row=BASIS_BANNER_ROW, end_column=PASTE_LAST_COL_IDX)
    ws.row_dimensions[BASIS_BANNER_ROW].height = 30

    # --- How-to-paste banner ---
    map_a = get_column_letter(MAP_COL_START)          # T — mapped col "A"
    map_g = get_column_letter(MAP_COL_START + 6)      # Z — mapped col "G"
    copy_range = f"{map_a}{FIRST_DATA_ROW}:{map_g}{LAST_DATA_ROW}"
    howto = ws.cell(
        row=HOWTO_BANNER_ROW, column=1,
        value=(f"HOW TO USE:  1) Clear the green zone first — select A{FIRST_DATA_ROW}:R{LAST_DATA_ROW} and press "
               f"Delete (it ships with demo data).  2) Paste the CLASS data rows into the green zone.  "
               f"3) Copy the mapped block's register cols A:G — physical range {copy_range} — then go to "
               f"Inputs!A{REGISTER_FIRST_DATA_ROW} and Paste-Special > Values.  Do NOT copy mapped col H "
               f"(the register's own locked formula).  4) On Inputs, fill E / F / G / I by hand afterwards."),
    )
    howto.font = Font(name="Arial", size=10, color="555555")
    howto.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=HOWTO_BANNER_ROW, start_column=1, end_row=HOWTO_BANNER_ROW, end_column=PASTE_LAST_COL_IDX)
    ws.row_dimensions[HOWTO_BANNER_ROW].height = 40

    # --- Capacity warning (fires if >50 holdings pasted) ---
    cap = ws.cell(
        row=CAPACITY_WARN_ROW, column=1,
        value=(f'=IF(COUNTA({PASTE_COL_CODE}{LAST_DATA_ROW + 1}:{PASTE_COL_CODE}{LAST_DATA_ROW + 200})>0,'
               f'"⚠ More than 50 holdings pasted — rows below {LAST_DATA_ROW} are NOT mapped. '
               f'The register holds 50; split the fund or trim.","")'),
    )
    cap.font = Font(name="Arial", size=10, bold=True, color="A61B1B")
    ws.merge_cells(start_row=CAPACITY_WARN_ROW, start_column=1, end_row=CAPACITY_WARN_ROW, end_column=PASTE_LAST_COL_IDX)

    # --- Paste-alignment warning: codes present but no numeric Total Cost ---
    align = ws.cell(
        row=ALIGN_WARN_ROW, column=1,
        value=(f'=IF(AND(COUNTA({PASTE_COL_CODE}{FIRST_DATA_ROW}:{PASTE_COL_CODE}{LAST_DATA_ROW})>0,'
               f'COUNT({PASTE_COL_COST}{FIRST_DATA_ROW}:{PASTE_COL_COST}{LAST_DATA_ROW})=0),'
               f'"⚠ Total Cost column (L) contains no numbers — check paste alignment / header row included.","")'),
    )
    align.font = Font(name="Arial", size=10, bold=True, color="A61B1B")
    ws.merge_cells(start_row=ALIGN_WARN_ROW, start_column=1, end_row=ALIGN_WARN_ROW, end_column=PASTE_LAST_COL_IDX)

    # --- Copy-range hint pinned next to the mapped block itself ---
    hint = ws.cell(
        row=ALIGN_WARN_ROW, column=MAP_COL_START,
        value=f"Copy {copy_range} → Inputs!A{REGISTER_FIRST_DATA_ROW} → Paste-Special > Values",
    )
    hint.font = Font(name="Arial", size=9, italic=True, color="555555")

    # --- Paste-zone header (CLASS native order) ---
    for idx, header in enumerate(CLASS_HEADERS, start=1):
        c = ws.cell(row=HEADER_ROW, column=idx, value=header)
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL if idx in PASTE_USED_COL_IDX else SLATE_FILL
        c.alignment = CENTER

    # --- Paste-zone data cells (unlocked; sample preloaded) ---
    for offset in range(PASTE_ROWS):
        r = FIRST_DATA_ROW + offset
        sample = SAMPLE_ROWS[offset] if offset < len(SAMPLE_ROWS) else None
        for idx in range(1, PASTE_LAST_COL_IDX + 1):
            cell = ws.cell(row=r, column=idx)
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL
            cell.border = THIN_BOX
            cell.protection = Protection(locked=False)
        if sample is not None:
            code, name, gl, cost, mv = sample
            ws[f"{PASTE_COL_CODE}{r}"].value = code
            ws[f"{PASTE_COL_NAME}{r}"].value = name
            ws[f"{PASTE_COL_GL}{r}"].value = gl
            cost_cell = ws[f"{PASTE_COL_COST}{r}"]
            cost_cell.value = cost
            cost_cell.number_format = FMT_CURRENCY
            mv_cell = ws[f"{PASTE_COL_MV}{r}"]
            mv_cell.value = mv
            mv_cell.number_format = FMT_CURRENCY

    # --- Overflow landing zone (rows 57-256, paste-zone cols only) ---
    # Unlocked so a >50-row CLASS paste lands instead of being rejected by
    # sheet protection; nothing below LAST_DATA_ROW is mapped, and the
    # capacity banner (CAPACITY_WARN_ROW) fires on COUNTA over this band.
    for r in range(LAST_DATA_ROW + 1, LAST_DATA_ROW + 201):
        for idx in range(1, PASTE_LAST_COL_IDX + 1):
            ws.cell(row=r, column=idx).protection = Protection(locked=False)

    # --- Mapped-block header (register-shaped) ---
    for off, header in enumerate(MAP_HEADERS):
        col = MAP_COL_START + off
        c = ws.cell(row=HEADER_ROW, column=col, value=header)
        c.font = SECTION_BAND_FONT
        c.alignment = CENTER
        if off in _BLANK_USER:
            c.fill = PatternFill("solid", fgColor="C7A752")   # gold — you fill later
        elif off == _FORMULA_SLOT:
            c.fill = PatternFill("solid", fgColor="4A6B5E")    # sage — don't copy
        else:
            c.fill = SECTION_BAND_FILL
    flag_hdr = ws.cell(row=HEADER_ROW, column=MAP_FLAG_COL_IDX, value="Review flag")
    flag_hdr.font = SECTION_BAND_FONT
    flag_hdr.fill = SECTION_BAND_FILL
    flag_hdr.alignment = CENTER

    # --- Mapped-block data (formulas; locked) ---
    for offset in range(PASTE_ROWS):
        r = FIRST_DATA_ROW + offset
        drop = _drop_expr(r)
        for off in range(len(MAP_HEADERS)):
            col = MAP_COL_START + off
            cell = ws.cell(row=r, column=col)
            if off in _POPULATED:
                src = f"${_POPULATED[off]}{r}"
                cell.value = f'=IF({drop},"",{src})'
                cell.font = BODY_FONT
                if off in (2, 3):  # cost / MV currency
                    cell.number_format = FMT_CURRENCY
                else:
                    cell.number_format = FMT_TEXT
            elif off in _BLANK_USER:
                cell.fill = AMBER_FILL          # blank template — user fills on Inputs
                cell.number_format = FMT_TEXT
            else:  # _FORMULA_SLOT (H)
                cell.fill = FORMULA_FILL
        # Review flag — negative tax cost base, or a kept row whose Total
        # Cost is blank/non-numeric (header row or column-shifted paste).
        flag = ws.cell(row=r, column=MAP_FLAG_COL_IDX)
        cost_ref = f"${PASTE_COL_COST}{r}"
        flag.value = (
            f'=IF(AND(NOT({drop}),ISNUMBER({cost_ref}),{cost_ref}<0),'
            f'">> NEGATIVE tax cost base - review (CGT event E4?)",'
            f'IF(AND(NOT({drop}),NOT(ISNUMBER({cost_ref}))),'
            f'">> Total Cost blank/non-numeric - check paste alignment",""))'
        )
        flag.font = Font(name="Arial", size=9, bold=True, color="A61B1B")

    # --- Conditional formatting on the mapped block ---
    map_code_col = get_column_letter(MAP_COL_START)
    map_cost_col = get_column_letter(MAP_COL_START + 2)
    map_last_col = get_column_letter(MAP_COL_START + len(MAP_HEADERS) - 1)
    rng = f"{map_code_col}{FIRST_DATA_ROW}:{map_last_col}{LAST_DATA_ROW}"
    # Grey dropped rows: source present (paste code non-blank) but mapped code blank.
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'AND(LEN(TRIM(${PASTE_COL_CODE}{FIRST_DATA_ROW}))>0,'
                     f'${map_code_col}{FIRST_DATA_ROW}="")'],
            fill=GREY_FILL,
        ),
    )
    # Red negative tax cost base on the mapped cost column.
    ws.conditional_formatting.add(
        f"{map_cost_col}{FIRST_DATA_ROW}:{map_cost_col}{LAST_DATA_ROW}",
        FormulaRule(
            formula=[f'AND(ISNUMBER({map_cost_col}{FIRST_DATA_ROW}),{map_cost_col}{FIRST_DATA_ROW}<0)'],
            fill=PatternFill("solid", fgColor="FBE9E9"),
            font=Font(name="Arial", size=10, bold=True, color="A61B1B"),
        ),
    )

    # --- Column widths ---
    paste_widths = [16, 12, 40, 22, 18, 12, 12, 30, 12, 14, 12, 16, 16, 14, 12, 14, 14, 12]
    for idx, w in enumerate(paste_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.column_dimensions[get_column_letter(MAP_COL_START - 1)].width = 3   # spacer (S)
    map_widths = [14, 38, 22, 20, 16, 18, 16, 22, 16]
    for off, w in enumerate(map_widths):
        ws.column_dimensions[get_column_letter(MAP_COL_START + off)].width = w
    ws.column_dimensions[get_column_letter(MAP_FLAG_COL_IDX)].width = 42

    ws.freeze_panes = f"A{FIRST_DATA_ROW}"

    # --- Sheet protection (paste zone unlocked above; everything else locked) ---
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    # --- Print header watermark + page setup (v3.4 audit: was missing here) ---
    ws.oddHeader.center.text = "ILLUSTRATIVE — NOT ADVICE"
    ws.oddHeader.center.size = 28
    ws.oddHeader.center.color = "CCCCCC"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0   # staff-only tab; let the paste zone spill
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws
