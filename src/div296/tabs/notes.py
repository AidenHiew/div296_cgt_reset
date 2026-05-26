"""Notes tab — terminology, caveats, valuation log, provenance.

Per spec §10 plus the locked additional disclosures (loss-offset divergence,
pension-phase exclusion, reset-OFF realised-only, wash sale / Part IVA,
transaction costs, alternative levers).

All text is FACTUAL DISCLOSURE only — no recommendation language.

Hidden provenance block at the bottom carries build_version, build_date,
git short-SHA so any printed/shared file can be traced back to its source.
"""

from __future__ import annotations

import datetime as _dt
import subprocess as _sp

from openpyxl.styles import Alignment, Font, Protection
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from div296 import __version__ as _ver
from div296.assumptions import ASSUMPTIONS
from div296.styles import (
    BODY_FONT, INPUT_FILL, INPUT_FONT, SECTION_BAND_FILL,
    SECTION_BAND_FONT, THIN_BOX, TITLE_FONT,
)
from div296.tabs.inputs import REGISTER_FIRST_DATA_ROW


SHEET = "Notes"
INPUTS_SHEET = "'Inputs'"


# --- Content (all factual disclosure — no recommendation language) ---

TERMINOLOGY = [
    ("Division 296 earnings",
     "Headline earnings figure used to compute Div 296 tax. In this model "
     "(realised-only basis), this equals the sum of positive Div 296 "
     "adjusted taxable capital gains across the asset register."),
    ("Ordinary taxable capital gain",
     "Capital gain on disposal using the asset's ORIGINAL cost base, "
     "discounted by 1/3 if the asset has been held > 12 months and the "
     "CGT discount toggle is ON."),
    ("Div 296 adjusted taxable capital gain",
     "Capital gain on disposal using the asset's Div 296 cost base — which "
     "is the original cost base when reset = OFF, or the market value at "
     "30 June 2026 when reset = ON. Same discount logic as the ordinary gain."),
    ("Ordinary CGT",
     "Capital gains tax on the ordinary taxable gain at the SMSF "
     "accumulation-phase rate of 15%. This model floors each asset at $0 "
     "for the per-asset display — see caveats below."),
    ("Div 296 tax",
     "Additional tax under Division 296 = member's attributed share of "
     "earnings × (proportion of TSB in $3m–$10m band × 15% + proportion "
     "above $10m × 25%). Both bands are always applied per the enacted "
     "law. Member splits are auto-derived from TSB share."),
    ("Original cost base",
     "Cost base used for ordinary CGT — unchanged by the reset election."),
    ("Div 296 cost base",
     "Cost base used ONLY for Div 296 calculations. Equals the original "
     "cost base if reset = OFF, or the market value at 30 June 2026 if "
     "reset = ON."),
    ("Reset election",
     "One-off, all-or-nothing, irrevocable election to reset every CGT "
     "asset's Div 296 cost base to its market value at 30 June 2026. "
     "Ordinary CGT is unaffected."),
    ("Reset impact",
     "(Div 296 adjusted gain WITH reset) minus (Div 296 adjusted gain "
     "WITHOUT reset). Negative = reset reduces the Div 296 gain. "
     "Positive = reset creates a Div 296 gain that did not previously exist "
     "(the 'reset trap', typically on assets held in an unrealised-loss "
     "position at 30 June 2026)."),
]

CAVEATS = [
    ("Illustrative only — not financial, tax or legal advice.",
     "Confirm against the final ATO method, regulations, and a registered "
     "tax agent or licensed financial adviser before relying on any figure."),
    ("Royal Assent / enactment status.",
     "Division 296 is enacted law; some operational detail sits in "
     "regulations. The Royal Assent date held on the Notes tab is "
     "user-editable — verify against current ATO guidance before use."),
    ("The reset is all-or-nothing and irrevocable.",
     "This model lets you toggle it freely FOR COMPARISON ONLY. The actual "
     "election applies to every CGT asset held at 30 June 2026."),
    ("30 June 2026 valuations are the load-bearing input.",
     "Garbage in, garbage out. Every Div 296 figure flows from the MV cells; "
     "use the Valuation Log below to record source and date for each asset."),
    ("Loss-offset divergence from ATO method.",
     "This model computes Ordinary CGT on a per-asset siloed basis "
     "(MAX(0, gain) × 15% per asset, with losses kept as carry-forward). "
     "ATO Subdivision 102-A nets gross gains and losses BEFORE the discount, "
     "which generally produces lower Ordinary CGT and a smaller carry-forward "
     "loss balance in years where gains and losses both occur. Real-world "
     "figures must be reconciled by the firm's tax practitioner."),
    ("Pension phase is NOT modelled.",
     "This model assumes 100% accumulation phase (fund earnings tax = 15%). "
     "Retirement-phase assets supporting a pension are taxed at 0% — for "
     "funds with pension members, this model overstates Ordinary CGT and "
     "the relative attractiveness of the reset."),
    ("Reset OFF scenario is realised-only.",
     "In reality, a fund that does NOT elect the reset is taxed under Div 296 "
     "on the year-on-year movement in TSB (unrealised + realised). The "
     "Comparison tab compares realised vs realised, so it understates the "
     "no-reset Div 296 burden."),
    ("Multi-member split is a TSB-proportion approximation.",
     "A real multi-member fund determines each member's share of Div 296 "
     "earnings via an actuarial certificate based on time-weighted average "
     "balances. This model approximates that split as each member's TSB "
     "divided by total fund TSB (auto-derived on Inputs Section 1)."),
    ("Wash sale / Part IVA risk.",
     "Disposing of an asset pre-30 June 2026 purely for the tax outcome and "
     "reacquiring it sits in anti-avoidance territory (TR 2008/1, Part IVA). "
     "Any pre-reset disposal strategy must be considered against these rules."),
    ("Transaction costs, liquidity, market-timing risk.",
     "Any pre-reset disposal incurs brokerage, stamp duty (for property), "
     "potential illiquidity for unlisted assets, and market-timing risk. "
     "These are not modelled."),
    ("Scope deliberately excludes dividends, interest and other earnings.",
     "This model addresses Div 296 capital gains only."),
    ("Alternative levers not modelled.",
     "TSB recontribution / withdrawal splits, pension commencement, and "
     "estate-timing considerations can materially change the optimal "
     "outcome. None are reflected in this model."),
    ("Sheet protection is tamper-evident, not tamper-proof.",
     "Sheet protection (where applied) is passwordless. It exists to "
     "prevent accidental overwrites, not to enforce immutability."),
]


def _band(ws: Worksheet, row: int, text: str, last_col_letter: str = "D") -> None:
    ws.cell(row=row, column=1, value=text).font = SECTION_BAND_FONT
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    for col_idx in range(1, ord(last_col_letter) - ord("A") + 2):
        ws.cell(row=row, column=col_idx).fill = SECTION_BAND_FILL


def _git_short_sha() -> str:
    try:
        out = _sp.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True, text=True, check=False, timeout=2,
        )
        return out.stdout.strip() or "uncommitted"
    except Exception:
        return "unknown"


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False

    # --- Title ---
    ws["A1"] = "Division 296 Cost Base Reset Model — Notes"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # --- Preparer attribution (visible, mirrors workbook Author metadata) ---
    ws["A2"] = f"Prepared by: Aiden Hiew  ·  Model version v{_ver}  ·  Built {_dt.date.today().isoformat()}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells("A2:D2")

    # --- Enactment status (editable) ---
    _band(ws, 3, "Enactment status (editable)")
    ws.cell(row=4, column=1, value="Royal Assent / enactment date:").font = BODY_FONT
    enact = ws.cell(row=4, column=2)
    enact.value = "Verify against current ATO guidance"
    enact.font = INPUT_FONT
    enact.fill = INPUT_FILL
    enact.border = THIN_BOX
    enact.protection = Protection(locked=False)   # editable under sheet protection
    ws.merge_cells("B4:D4")

    # --- Terminology ---
    row = 6
    _band(ws, row, "Terminology")
    row += 1
    for term, definition in TERMINOLOGY:
        ws.cell(row=row, column=1, value=term).font = Font(name="Arial", size=10, bold=True)
        d = ws.cell(row=row, column=2, value=definition)
        d.font = BODY_FONT
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 36
        row += 1

    # --- Caveats ---
    row += 1
    _band(ws, row, "Caveats (factual disclosure)")
    row += 1
    for headline, body in CAVEATS:
        ws.cell(row=row, column=1, value=headline).font = Font(name="Arial", size=10, bold=True)
        d = ws.cell(row=row, column=2, value=body)
        d.font = BODY_FONT
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 48
        row += 1

    # --- Valuation log (mirrors each asset's valuation source/date) ---
    row += 1
    _band(ws, row, "Valuation log (mirrored from Inputs)")
    row += 1
    log_headers = ["Asset code", "Asset name", "Valuation source / date", "Market value at 30 Jun 2026"]
    for col_idx, header in enumerate(log_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL
    row += 1

    for offset in range(ASSUMPTIONS.asset_register_rows):
        i_row = REGISTER_FIRST_DATA_ROW + offset
        ws.cell(row=row, column=1, value=f"=IF({INPUTS_SHEET}!A{i_row}=\"\",\"\",{INPUTS_SHEET}!A{i_row})")
        ws.cell(row=row, column=2, value=f"=IF({INPUTS_SHEET}!B{i_row}=\"\",\"\",{INPUTS_SHEET}!B{i_row})")
        # v2.3 Inputs col layout: F = Valuation source/date (was G), E = MV 30 Jun (was F)
        ws.cell(row=row, column=3, value=f"=IF({INPUTS_SHEET}!F{i_row}=\"\",\"\",{INPUTS_SHEET}!F{i_row})")
        ws.cell(row=row, column=4, value=f"=IF({INPUTS_SHEET}!E{i_row}=\"\",\"\",{INPUTS_SHEET}!E{i_row})")
        ws.cell(row=row, column=4).number_format = '$#,##0;($#,##0);"-"'
        row += 1

    # --- Provenance (hidden) ---
    build_date = _dt.date.today().isoformat()
    prov_first = row + 2
    ws.cell(row=prov_first, column=1, value="build_version")
    ws.cell(row=prov_first, column=2, value=_ver)
    ws.cell(row=prov_first + 1, column=1, value="build_date")
    ws.cell(row=prov_first + 1, column=2, value=build_date)
    ws.cell(row=prov_first + 2, column=1, value="git_short_sha")
    ws.cell(row=prov_first + 2, column=2, value=_git_short_sha())
    ws.cell(row=prov_first + 3, column=1, value="model_logic")
    ws.cell(row=prov_first + 3, column=2, value=f"per build plan dated {build_date}")
    for r in range(prov_first, prov_first + 4):
        ws.row_dimensions[r].hidden = True

    # --- Column widths ---
    widths = {"A": 32, "B": 50, "C": 28, "D": 22}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # --- Print header watermark (compliance signal on every printed page) ---
    ws.oddHeader.center.text = "ILLUSTRATIVE — NOT ADVICE"
    ws.oddHeader.center.size = 28
    ws.oddHeader.center.color = "CCCCCC"

    # --- Sheet protection ---
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws
