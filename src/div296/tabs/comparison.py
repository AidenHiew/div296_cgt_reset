"""Comparison tab — print-ready landscape A4 tearsheet.

Layout (top to bottom; row numbers are the module's layout constants):
    Rows 1-10    Header block (banner, title, firm / prepared / date, disclaimer)
    Row 12       Section band: "Members & TSB"
    Rows 13-18   Fund-context strip (per-member TSB panel + fund total)
    Row 20       Band: "Headline — total Div 296 tax"
    Rows 21-23   Three metric cards: no-reset | elected | Net effect (+ footnote)
    Row 25       Band: "Per-scenario subtotals"
    Rows 26-30   Subtotals (Earnings / Ord CGT (unchanged) / Div 296 tax / TOTAL BURDEN)
    Row 33       Band: "Per-member breakdown"
    Rows 34-39   Per-member TSB + Div 296 tax (both scenarios) + Total row
    Row 41       Band: "Per-asset detail"
    Rows 42-43   Panel titles + column sub-headers
    Rows 44-53   Top-10 assets by |Δ Div 296 tax| (DISPLAY_ROWS rows)
    Row 54       Overflow note ("Showing top 10 — see Analyser for the full register")
    Rows 55-56   Reminder + sort-order notes

Columns:
    A-E:  Panel A (no reset)        — Asset / Proceeds / Cost base / Adj gain / Tax
    F-J:  Panel B (reset elected)   — same five columns
    K:    Change column (Δ = panel B tax − panel A tax)
    L-M:  Hidden fund helpers (per-member tax, headline totals)
    N-P:  Hidden per-register grid (adj gain A / adj gain B / |Δ| sort metric)
    R:    Hidden matched-register-row lookup for the top-10 panel

Both panels compute independently of any election state (v3.0 removed the
reset toggle), so the comparison always shows the real before/after picture.
The Net effect card is SIGNED (elected − no-reset): positive = the reset costs
more Div 296 tax; negative (green brackets) = a saving. No recommendation
language anywhere on the tab.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from div296._formulas import div296_adj_gain_formula, per_member_div296_tax_formula
from div296.assumptions import ASSUMPTIONS
from div296.styles import (
    BODY_FONT, CENTER, FMT_CURRENCY, FMT_CURRENCY_DELTA, INPUT_FILL, INPUT_FONT, SECTION_BAND_FILL, SECTION_BAND_FONT,
    THIN_BOX, TITLE_FONT, TRAP_FILL,
    PROC_DATA_FILL, ORD_DATA_FILL, DIV_DATA_FILL,
)
from div296.tabs import analyser as analyser_tab
from div296.tabs.inputs import (
    MEMBERS_FIRST_DATA_ROW, REGISTER_FIRST_DATA_ROW,
    REGISTER_LAST_DATA_ROW,
    sample_detect_expr,
)


SHEET = "Comparison"
INPUTS_SHEET = "'Inputs'"
ANALYSER_SHEET = "'Analyser'"

# How many assets to display on the printable comparison page.
# v1.7: 10 rows, sorted by absolute Δ (Scenario B − Scenario A) descending so
# the assets most affected by the reset election appear at the top. Empty
# register rows sort to the bottom and render blank.
DISPLAY_ROWS = 10

# --- Layout rows ---
WATERMARK_ROW = 1
TITLE_ROW = 2
HEADER_FIRM_ROW = 4
HEADER_PREPARED_FOR_ROW = 5
HEADER_PREPARED_BY_ROW = 6
HEADER_DATE_ROW = 7
HEADER_DISCLAIMER_ROW = 8
HEADER_LOGO_ROW = 10

BAND_BODY_ROW = 12
CONTEXT_LABEL_ROW = 13
# v2.5 FB-8: strip now shows per-member TSB (rows 14..14+member_count-1)
# plus a Total row. Other context tiles (proportion, discount, tier) get
# vertically merged across those rows on the right.
CONTEXT_VALUE_ROW = 14   # first member row (kept name for back-compat)
CONTEXT_MEMBER_LAST_ROW = CONTEXT_VALUE_ROW + ASSUMPTIONS.member_count - 1   # row 17
CONTEXT_TOTAL_ROW = CONTEXT_MEMBER_LAST_ROW + 1                              # row 18

# v2.5 step 13: Aiden inserted a blank spacer row before each section band
# during his manual formatting pass; layout below mirrors his .xlsx.
BAND_HEADLINE_ROW = CONTEXT_TOTAL_ROW + 2    # row 20 (spacer at 19)
CARD_LABEL_ROW = BAND_HEADLINE_ROW + 1       # row 21
CARD_VALUE_ROW = CARD_LABEL_ROW + 1          # row 22
CARD_FOOTNOTE_ROW = CARD_VALUE_ROW + 1       # row 23

BAND_SUBTOTALS_ROW = CARD_FOOTNOTE_ROW + 2   # row 25 (spacer at 24)
SUBTOTAL_HEADER_ROW = BAND_SUBTOTALS_ROW + 1 # row 26
SUBTOTAL_EARNINGS_ROW = SUBTOTAL_HEADER_ROW + 1   # row 27
SUBTOTAL_ORD_CGT_ROW = SUBTOTAL_EARNINGS_ROW + 1  # row 28
SUBTOTAL_DIV296_ROW = SUBTOTAL_ORD_CGT_ROW + 1    # row 29
SUBTOTAL_BURDEN_ROW = SUBTOTAL_DIV296_ROW + 1     # row 30

# v2.3: Per-member breakdown block between subtotals and per-asset detail.
# v2.5 step 13: spacer at row 32 between subtotal note (row 31) and band.
PER_MEMBER_BAND_ROW = SUBTOTAL_BURDEN_ROW + 3          # row 33
PER_MEMBER_HEADER_ROW = PER_MEMBER_BAND_ROW + 1        # row 34
PER_MEMBER_FIRST_ROW = PER_MEMBER_HEADER_ROW + 1       # row 35
PER_MEMBER_LAST_ROW = PER_MEMBER_FIRST_ROW + ASSUMPTIONS.member_count - 1   # row 38
# v2.5 step 13: NEW — Total row at end of per-member breakdown.
PER_MEMBER_TOTAL_ROW = PER_MEMBER_LAST_ROW + 1         # row 39

BAND_DETAIL_ROW = PER_MEMBER_TOTAL_ROW + 2             # row 41
PANEL_TITLE_ROW = BAND_DETAIL_ROW + 1                  # row 42
PANEL_HEADER_ROW = BAND_DETAIL_ROW + 2                 # row 43
DATA_FIRST_ROW = BAND_DETAIL_ROW + 3                   # row 44
DATA_LAST_ROW = DATA_FIRST_ROW + DISPLAY_ROWS - 1      # row 53
DATA_OVERFLOW_NOTE_ROW = DATA_LAST_ROW + 1             # row 54

REMINDER_ROW = DATA_OVERFLOW_NOTE_ROW + 1              # row 55
SORT_NOTE_ROW = REMINDER_ROW + 1                       # row 56

# v2.5 FB-3: chart removed. Print area now ends with the sort note (SORT_NOTE_ROW).

# --- Layout columns ---
PANEL_A_COLS = ("A", "B", "C", "D", "E")   # Asset, Proceeds, Cost base, Adj gain, Tax
PANEL_B_COLS = ("F", "G", "H", "I", "J")   # v2.5 FB-7: moved left (was G-K)
DELTA_COL = "K"                             # v2.5 FB-7: Change at the END now (was F)
LAST_VISIBLE_COL = DELTA_COL                # "K"

# Hidden helper columns (right of the visible panels)
HELPER_COL_A = "L"   # Scenario A working column
HELPER_COL_B = "M"   # Scenario B working column

# v1.7: per-register-row helper grid for sort-by-impact lookup.
# Cols N/O/P/Q span register rows (REGISTER_FIRST_DATA_ROW..REGISTER_LAST_DATA_ROW).
PER_REG_GAIN_A_COL = "N"   # Scenario A adj gain per register row
PER_REG_GAIN_B_COL = "O"   # Scenario B adj gain per register row
PER_REG_DELTA_COL = "P"    # |Δ| + row tiebreaker; empty rows = -1
# Col R: matched register row for each visible display row (rows 30..39).
MATCHED_ROW_COL = "R"

# Hidden helper rows in cols L/M
HELPER_FUND_EARNINGS_ROW = 1
HELPER_MEMBER_TAX_FIRST_ROW = 2
HELPER_MEMBER_TAX_LAST_ROW = HELPER_MEMBER_TAX_FIRST_ROW + ASSUMPTIONS.member_count - 1
HELPER_HEADLINE_ROW = HELPER_MEMBER_TAX_LAST_ROW + 1   # row 6

# --- Styling palette specific to v1.5 ---
CARD_FILL_A = PatternFill("solid", fgColor="EFF5F3")    # very light teal
CARD_FILL_B = PatternFill("solid", fgColor="EAF3FB")    # very light blue
CARD_FILL_DELTA = PatternFill("solid", fgColor="FFF5E1")  # very light gold
CARD_BORDER = Border(
    left=Side(style="medium", color="1D3B34"),
    right=Side(style="medium", color="1D3B34"),
    top=Side(style="medium", color="1D3B34"),
    bottom=Side(style="medium", color="1D3B34"),
)
CARD_LABEL_FONT = Font(name="Arial", size=10, italic=True, color="1D3B34")
CARD_VALUE_FONT = Font(name="Arial", size=22, bold=True, color="1D3B34")
CONTEXT_LABEL_FONT = Font(name="Arial", size=9, color="666666")
CONTEXT_VALUE_FONT = Font(name="Arial", size=11, bold=True, color="1D3B34")
TOTAL_BURDEN_FONT = Font(name="Arial", size=11, bold=True, color="1D3B34")
TOTAL_BURDEN_FILL = PatternFill("solid", fgColor="EFF5F3")
COST_BASE_ACCENT_FILL = PatternFill("solid", fgColor="FFF8E6")   # subtle gold tint
DELTA_HEADER_FILL = PatternFill("solid", fgColor="FFF5E1")
DELTA_FONT = Font(name="Arial", size=10, italic=True, color="8A6D00")


# --- Small builders ---

def _band(ws: Worksheet, row: int, text: str, last_col_letter: str = LAST_VISIBLE_COL) -> None:
    ws.cell(row=row, column=1, value=text).font = SECTION_BAND_FONT
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    for col_idx in range(1, ord(last_col_letter) - ord("A") + 2):
        ws.cell(row=row, column=col_idx).fill = SECTION_BAND_FILL


def _input_cell(ws: Worksheet, coord: str, value=None) -> None:
    cell = ws[coord]
    if value is not None:
        cell.value = value
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = THIN_BOX


# v3.0: `_member_tax_formula` moved to `div296._formulas.per_member_div296_tax_formula`
# (shared with analyser.py). Imported at the top of this module.


# --- Big builders for each visual block ---

def _build_header_block(ws: Worksheet) -> None:
    # Row 1: on-screen ILLUSTRATIVE banner
    watermark_font = Font(name="Arial", size=14, bold=True, italic=True, color="999999")
    cell = ws.cell(row=WATERMARK_ROW, column=1, value="ILLUSTRATIVE — NOT ADVICE")
    cell.font = watermark_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{WATERMARK_ROW}:{LAST_VISIBLE_COL}{WATERMARK_ROW}")

    # Row 2: Title
    title_cell = ws.cell(row=TITLE_ROW, column=1,
                         value="Division 296 Cost Base Reset Model — Comparison")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(f"A{TITLE_ROW}:{LAST_VISIBLE_COL}{TITLE_ROW}")

    # Rows 4-7: editable cells for firm / prepared for / by / date
    header_rows = [
        (HEADER_FIRM_ROW, "Firm name:"),
        (HEADER_PREPARED_FOR_ROW, "Prepared for:"),
        (HEADER_PREPARED_BY_ROW, "Prepared by:"),
        (HEADER_DATE_ROW, "Date:"),
    ]
    for row, label in header_rows:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        _input_cell(ws, f"B{row}", value="")
        ws.merge_cells(f"B{row}:E{row}")

    # Row 8: disclaimer line (read-only)
    disc = ws.cell(row=HEADER_DISCLAIMER_ROW, column=1,
                   value="Illustrative only — not financial, tax or legal advice.")
    disc.font = Font(name="Arial", size=10, italic=True, color="666666")
    ws.merge_cells(f"A{HEADER_DISCLAIMER_ROW}:{LAST_VISIBLE_COL}{HEADER_DISCLAIMER_ROW}")

    # Row 10: logo placeholder (top-right area)
    logo = ws.cell(row=HEADER_LOGO_ROW, column=10, value="[ logo ]")
    logo.font = Font(name="Arial", size=10, italic=True, color="999999")
    logo.alignment = Alignment(horizontal="center", vertical="center")
    logo.fill = PatternFill("solid", fgColor="F5F5F5")
    ws.merge_cells(f"J{HEADER_LOGO_ROW}:{LAST_VISIBLE_COL}{HEADER_LOGO_ROW + 1}")

    # v2.2.0: Sample-data warning (row 9, full width left of the logo block).
    # Formula returns the warning string ONLY when the first three asset codes
    # still match the sample register; CF paints amber only in that case.
    sample_detect = sample_detect_expr(f"{INPUTS_SHEET}!")
    sample_row = HEADER_DISCLAIMER_ROW + 1   # row 9
    badge = ws.cell(
        row=sample_row, column=1,
        value=(
            f'=IF({sample_detect},'
            '"⚠  Sample data detected — the figures shown are illustrative '
            'only until the asset register on Inputs is replaced with the '
            "actual fund's holdings.\",\"\")"
        ),
    )
    badge.font = Font(name="Arial", size=10, bold=True, italic=True, color="8A6D00")
    badge.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A{sample_row}:I{sample_row}")
    ws.row_dimensions[sample_row].height = 20
    amber_rule = FormulaRule(
        formula=[sample_detect],
        fill=PatternFill("solid", fgColor="FFF4CE"),
    )
    ws.conditional_formatting.add(f"A{sample_row}:I{sample_row}", amber_rule)


def _build_context_strip(ws: Worksheet) -> None:
    """Top context block. v2.5 polish: only the per-member TSB mini-table
    remains. The v2.5-initial right-side tiles (proportion / discount / tier)
    were dropped per Aiden's feedback — they were either echoes of Inputs
    toggles (discount, tier) or derived metrics already visible per-member
    (proportion). The strip is now a single panel on the left (cols A:C).
    """
    last_input_member_row = MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count - 1

    # --- Row 13: column headers for the mini-table beneath ---
    # v2.5 step 13 (Aiden polish): was a single merged sub-header echoing the
    # band; now functions as two column headings — "Members" | "Total Super
    # Balance" — so the table reads as a proper grid.
    header_a = ws[f"A{CONTEXT_LABEL_ROW}"]
    header_a.value = "Members"
    header_a.font = Font(name="Arial", size=10, bold=True, color="1D3B34")
    header_a.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    header_b = ws[f"B{CONTEXT_LABEL_ROW}"]
    header_b.value = "Total Super Balance"
    header_b.font = Font(name="Arial", size=10, bold=True, color="1D3B34")
    header_b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"B{CONTEXT_LABEL_ROW}:C{CONTEXT_LABEL_ROW}")

    # --- Left tile (cols A:C) — per-member TSB mini-table ---
    # Col A always shows the placeholder label; col B-C merged for TSB value
    # (blank when the corresponding Inputs row has no TSB).
    # v2.5 step 13 (Aiden polish): TSB values left-aligned (was right) so the
    # column reads as a tighter list rather than a financial column.
    for i in range(ASSUMPTIONS.member_count):
        c_row = CONTEXT_VALUE_ROW + i
        inputs_row = MEMBERS_FIRST_DATA_ROW + i
        tsb_ref = f"{INPUTS_SHEET}!B{inputs_row}"

        label_cell = ws.cell(row=c_row, column=1, value=f"Member {i+1}")
        label_cell.font = CONTEXT_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="left", indent=1)

        value_cell = ws.cell(row=c_row, column=2,
                             value=f'=IF({tsb_ref}>0,{tsb_ref},"")')
        value_cell.font = CONTEXT_VALUE_FONT
        value_cell.number_format = FMT_CURRENCY
        value_cell.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(f"B{c_row}:C{c_row}")

        ws.row_dimensions[c_row].height = 16

    # Total row underneath the members.
    total_label = ws.cell(row=CONTEXT_TOTAL_ROW, column=1, value="Total")
    total_label.font = Font(name="Arial", size=10, bold=True, color="1D3B34")
    total_label.alignment = Alignment(horizontal="left", indent=1)

    total_value = ws.cell(
        row=CONTEXT_TOTAL_ROW, column=2,
        value=f"=SUM({INPUTS_SHEET}!B{MEMBERS_FIRST_DATA_ROW}:B{last_input_member_row})",
    )
    total_value.font = Font(name="Arial", size=11, bold=True, color="1D3B34")
    total_value.number_format = FMT_CURRENCY
    total_value.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(f"B{CONTEXT_TOTAL_ROW}:C{CONTEXT_TOTAL_ROW}")
    ws.row_dimensions[CONTEXT_TOTAL_ROW].height = 18

    # v2.5 step 13 (Aiden polish): thin border box across the whole mini-table
    # so it reads as a single grid (was borderless).
    table_border = Border(
        left=Side(style="thin", color="C5CECA"),
        right=Side(style="thin", color="C5CECA"),
        top=Side(style="thin", color="C5CECA"),
        bottom=Side(style="thin", color="C5CECA"),
    )
    for r in range(CONTEXT_LABEL_ROW, CONTEXT_TOTAL_ROW + 1):
        for col in ("A", "B", "C"):
            ws[f"{col}{r}"].border = table_border

    # v2.5 polish: right-side tiles (proportion / discount / tier) removed.
    # The strip occupies cols A:C only now; cols D:K are deliberately blank.


def _build_per_register_helpers(ws: Worksheet) -> tuple[str, str, str]:
    """Per-register-row helper grid (cols N/O/P) used to sort the per-asset
    detail panel by |Δ| descending.

    N = Scenario A Div 296 adj gain (cost base = original)
    O = Scenario B Div 296 adj gain (cost base = MV)
    P = |tax_b − tax_a| + row tiebreaker, where tax_x = MAX(0,gain_x)/SUMIF(>0)
        × headline_x (cells L6/M6). v2.5 FB-7: switched from gain-delta to
        tax-delta so the "top 10 most affected" matches the visible Change col
        (which now displays tax delta, not gain delta).
        Empty register rows return -1 so LARGE pushes them to the bottom.

    Returns the absolute ranges (e.g. '$N$20:$N$69').
    """
    n_first, n_last = REGISTER_FIRST_DATA_ROW, REGISTER_LAST_DATA_ROW
    # Range refs reused inside the per-row tax formula.
    a_range = f"${PER_REG_GAIN_A_COL}${n_first}:${PER_REG_GAIN_A_COL}${n_last}"
    b_range = f"${PER_REG_GAIN_B_COL}${n_first}:${PER_REG_GAIN_B_COL}${n_last}"
    headline_a_abs = f"${HELPER_COL_A}${HELPER_HEADLINE_ROW}"   # $L$6
    headline_b_abs = f"${HELPER_COL_B}${HELPER_HEADLINE_ROW}"   # $M$6
    for n in range(n_first, n_last + 1):
        # v2.3 Inputs column layout: A code / B name / C orig CB / D MV today
        # / E MV 30 Jun / F val source / G proceeds / H projected G/L / I held
        # (raw) / J held (hidden, paste-normalised — v3.1.2)
        proceeds = f"{INPUTS_SHEET}!G{n}"   # was H
        orig = f"{INPUTS_SHEET}!C{n}"        # was D
        mv = f"{INPUTS_SHEET}!E{n}"          # was F
        held = f"{INPUTS_SHEET}!J{n}"        # v3.1.2: J (normalised), was I
        ws[f"{PER_REG_GAIN_A_COL}{n}"] = div296_adj_gain_formula(proceeds, orig, held)
        ws[f"{PER_REG_GAIN_B_COL}{n}"] = div296_adj_gain_formula(proceeds, mv, held)
        # v2.5: per-row tax inlined into the sort metric so the helper grid
        # itself measures tax impact, not raw gain impact.
        tax_a = (
            f'IF(SUMIF({a_range},">0")=0,0,'
            f'MAX(0,{PER_REG_GAIN_A_COL}{n})/SUMIF({a_range},">0")*{headline_a_abs})'
        )
        tax_b = (
            f'IF(SUMIF({b_range},">0")=0,0,'
            f'MAX(0,{PER_REG_GAIN_B_COL}{n})/SUMIF({b_range},">0")*{headline_b_abs})'
        )
        # Tiebreaker: small positive amount decreasing with register row, so
        # earlier-listed assets win ties without ever turning P negative. The
        # +1 keeps the LAST register row strictly positive, so a genuine
        # zero-delta asset there isn't dropped by the LARGE(...)<=0 rank cut.
        tiebreak = f"(({n_last}+1)-ROW())*0.0001"
        # Incomplete rows (either gain cell blank) are excluded from the
        # top-10 ranking like empty rows — they're flagged on Inputs instead.
        ws[f"{PER_REG_DELTA_COL}{n}"] = (
            f'=IF(OR({proceeds}="",{PER_REG_GAIN_A_COL}{n}="",{PER_REG_GAIN_B_COL}{n}=""),'
            f'-1,ABS(({tax_b})-({tax_a}))+{tiebreak})'
        )
    return (
        a_range, b_range,
        f"${PER_REG_DELTA_COL}${n_first}:${PER_REG_DELTA_COL}${n_last}",
    )


def _build_helpers(
    ws: Worksheet, gain_a_range: str, gain_b_range: str,
) -> tuple[str, str]:
    """Hidden L/M block: fund earnings (over the full register), per-member tax,
    and headline totals. Returns (headline_a_abs, headline_b_abs).

    v3.1: fund-earnings cells use MAX(0, SUM(...)) — intra-year netting of
    gains and losses, floored at zero. v3.0 used SUMIF(>0) which floored
    per-asset before summing.
    """
    ws[f"{HELPER_COL_A}{HELPER_FUND_EARNINGS_ROW}"] = f'=MAX(0, SUM({gain_a_range}))'
    ws[f"{HELPER_COL_B}{HELPER_FUND_EARNINGS_ROW}"] = f'=MAX(0, SUM({gain_b_range}))'

    a_earnings = f"${HELPER_COL_A}${HELPER_FUND_EARNINGS_ROW}"
    b_earnings = f"${HELPER_COL_B}${HELPER_FUND_EARNINGS_ROW}"

    for i in range(ASSUMPTIONS.member_count):
        helper_row = HELPER_MEMBER_TAX_FIRST_ROW + i
        inputs_row = MEMBERS_FIRST_DATA_ROW + i
        ws[f"{HELPER_COL_A}{helper_row}"] = per_member_div296_tax_formula(inputs_row, a_earnings)
        ws[f"{HELPER_COL_B}{helper_row}"] = per_member_div296_tax_formula(inputs_row, b_earnings)

    headline_a_cell = f"{HELPER_COL_A}{HELPER_HEADLINE_ROW}"
    headline_b_cell = f"{HELPER_COL_B}{HELPER_HEADLINE_ROW}"
    ws[headline_a_cell] = (
        f"=SUM({HELPER_COL_A}{HELPER_MEMBER_TAX_FIRST_ROW}:"
        f"{HELPER_COL_A}{HELPER_MEMBER_TAX_LAST_ROW})"
    )
    ws[headline_b_cell] = (
        f"=SUM({HELPER_COL_B}{HELPER_MEMBER_TAX_FIRST_ROW}:"
        f"{HELPER_COL_B}{HELPER_MEMBER_TAX_LAST_ROW})"
    )

    # Hide helper columns (L/M plus the per-register grid + matched-row col R).
    for col_letter in (HELPER_COL_A, HELPER_COL_B,
                       PER_REG_GAIN_A_COL, PER_REG_GAIN_B_COL,
                       PER_REG_DELTA_COL, MATCHED_ROW_COL):
        ws.column_dimensions[col_letter].hidden = True

    abs_a = f"${HELPER_COL_A}${HELPER_HEADLINE_ROW}"
    abs_b = f"${HELPER_COL_B}${HELPER_HEADLINE_ROW}"
    return abs_a, abs_b


def _build_metric_cards(ws: Worksheet, headline_a: str, headline_b: str) -> None:
    """Three boxed cards — Scenario A / Scenario B / Net effect."""
    _band(ws, BAND_HEADLINE_ROW, "Headline — total Div 296 tax")

    # v2.5 step 13 (Aiden polish): headline cards use a VERBOSE variant of the
    # standard labels to give the client full context on the big tile. Subtotal
    # and per-member tables keep the short form ("If no reset (default)" / "If
    # elected to reset" / "Difference") to stay scannable. CONTEXT.md notes the
    # headline-only override. Card formula is SIGNED (reset − default), so a
    # reset that reduces tax shows as a green-bracket negative.
    cards = [
        ("A:D", "If no Div 296 CostBase Reset (default)",   f"={headline_a}",                           FMT_CURRENCY,       CARD_FILL_A),
        ("E:H", "If elected to reset Div 296 CostBase Reset", f"={headline_b}",                         FMT_CURRENCY,       CARD_FILL_B),
        ("I:K", "Difference (Net Div 296 Tax)",             f"={headline_b}-{headline_a}",      FMT_CURRENCY_DELTA, CARD_FILL_DELTA),
    ]
    for merge_range, label, value_formula, value_fmt, fill in cards:
        start = merge_range.split(":")[0]
        end = merge_range.split(":")[1]

        # Label row
        label_cell = ws[f"{start}{CARD_LABEL_ROW}"]
        label_cell.value = label
        label_cell.font = CARD_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{start}{CARD_LABEL_ROW}:{end}{CARD_LABEL_ROW}")

        # Value row (big)
        value_cell = ws[f"{start}{CARD_VALUE_ROW}"]
        value_cell.value = value_formula
        value_cell.font = CARD_VALUE_FONT
        value_cell.number_format = value_fmt
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{start}{CARD_VALUE_ROW}:{end}{CARD_VALUE_ROW}")

        # Border + fill across both rows of the card
        for row in (CARD_LABEL_ROW, CARD_VALUE_ROW):
            for col_letter in _expand_cols(start, end):
                cell = ws[f"{col_letter}{row}"]
                cell.fill = fill
                cell.border = CARD_BORDER

    ws.row_dimensions[CARD_LABEL_ROW].height = 18
    ws.row_dimensions[CARD_VALUE_ROW].height = 36

    # v2.3: Combined headline explanation + Year-1 caveat.
    # Replaces the v2.2.0 standalone Year-1 caveat — explanation comes first so
    # the reader frames the headline figures before reading any subtotals.
    footnote = ws.cell(
        row=CARD_FOOTNOTE_ROW, column=1,
        value=("The headline figures summarise the estimated Division 296 "
               "impact under each scenario. They help compare the relative "
               "outcomes of each option rather than provide a final tax "
               "calculation.\nHeadline figures are Year 1 only — Div 296 is "
               "assessed annually, and the same earnings hit recurs each year "
               "a member's TSB remains above $3m (thresholds index over time)."),
    )
    footnote.font = Font(name="Arial", size=9, italic=True, color="666666")
    footnote.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True, indent=1)
    ws.merge_cells(f"A{CARD_FOOTNOTE_ROW}:{LAST_VISIBLE_COL}{CARD_FOOTNOTE_ROW}")
    ws.row_dimensions[CARD_FOOTNOTE_ROW].height = 48


def _expand_cols(start: str, end: str) -> list[str]:
    return [chr(c) for c in range(ord(start), ord(end) + 1)]


def _build_subtotals(ws: Worksheet, headline_a: str, headline_b: str,
                     gain_a_range: str, gain_b_range: str) -> None:
    """Four-row subtotal table: earnings, ord CGT, Div 296 tax, total burden."""
    _band(ws, BAND_SUBTOTALS_ROW, "Per-scenario subtotals")

    # Header row — v2.5 FB-5: labels standardised across all sections.
    header_cells = [
        ("A", "Subtotal"),
        ("B", "If no reset (default)"),
        ("C", "If elected to reset"),
        ("D", "Difference"),
    ]
    for col, text in header_cells:
        c = ws[f"{col}{SUBTOTAL_HEADER_ROW}"]
        c.value = text
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL
        c.alignment = CENTER

    # Reference to the Analyser's Fund Ordinary CGT total — same in both
    # scenarios (ordinary CGT doesn't depend on reset election; uses
    # original cost base via ordinary_raw_gain).
    # v3.2: address derived from analyser.FUND_ORD_CGT_CELL constant so any
    # future row shift on the Analyser tab propagates automatically.
    ord_cgt_ref = f"={ANALYSER_SHEET}!{analyser_tab.FUND_ORD_CGT_CELL}"

    # v2.3 C-3: per-row definitions surfaced as cell Comments on the col-A label
    # so the reader can hover for a plain-English explanation of each subtotal.
    # v3.1: "Div 296 earnings" formula nets gains and losses (was SUMIF(>0)).
    rows = [
        (SUBTOTAL_EARNINGS_ROW, "Div 296 earnings",
         f"=MAX(0, SUM({gain_a_range}))", f"=MAX(0, SUM({gain_b_range}))",
         "The total estimated Division 296 adjusted taxable capital gain "
         "across all assets included in the analysis, net of capital losses "
         "within the year (s102-5 method). Floored at zero — Div 296 earnings "
         "cannot be negative."),
        (SUBTOTAL_ORD_CGT_ROW,  "Ordinary CGT (unchanged by reset)",
         ord_cgt_ref, ord_cgt_ref,
         "Ordinary capital gains tax — fund CGT rate applied to net taxable "
         "capital gain after intra-year netting of capital losses (s102-5 ITAA "
         "1997 method). Same in both scenarios — the reset election affects "
         "Div 296, not ordinary CGT."),
        (SUBTOTAL_DIV296_ROW,   "Div 296 tax (headline)",
         f"={headline_a}", f"={headline_b}",
         "Division 296 additional tax payable. Each member's share is "
         "computed from their TSB proportion in the $3m–$10m band (taxed "
         "at 15%) plus the slice above $10m (taxed at 25%)."),
        (SUBTOTAL_BURDEN_ROW,   "TOTAL TAX BURDEN",
         f"=B{SUBTOTAL_ORD_CGT_ROW}+B{SUBTOTAL_DIV296_ROW}",
         f"=C{SUBTOTAL_ORD_CGT_ROW}+C{SUBTOTAL_DIV296_ROW}",
         "Ordinary CGT + Div 296 tax. Year 1 only — see the headline "
         "footnote about projecting these figures forward."),
    ]
    subtotal_border = Border(
        left=Side(style="thin", color="C5CECA"),
        right=Side(style="thin", color="C5CECA"),
        top=Side(style="thin", color="C5CECA"),
        bottom=Side(style="thin", color="C5CECA"),
    )
    for row, label, a_val, b_val, definition in rows:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = a_val
        ws[f"C{row}"] = b_val
        # v2.5 FB-5: Change is SIGNED reset − default (negative = saving).
        ws[f"D{row}"] = f"=C{row}-B{row}"
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{row}"].comment = Comment(definition, "v2.3")
        for col in ("B", "C"):
            ws[f"{col}{row}"].number_format = FMT_CURRENCY
            ws[f"{col}{row}"].alignment = Alignment(horizontal="right", vertical="center")
        # Change col uses signed-delta format (green brackets for negatives).
        ws[f"D{row}"].number_format = FMT_CURRENCY_DELTA
        ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center")

        # v2.3: light borders on all subtotal cells.
        for col in ("A", "B", "C", "D"):
            ws[f"{col}{row}"].border = subtotal_border

        # v2.4 FB-4: taller rows (was 20) so wrapped labels don't clip
        # and currency values breathe.
        ws.row_dimensions[row].height = 22

        # Emphasise the total-burden row
        if row == SUBTOTAL_BURDEN_ROW:
            top_border = Border(
                left=subtotal_border.left,
                right=subtotal_border.right,
                top=Side(style="medium", color="1D3B34"),
                bottom=subtotal_border.bottom,
            )
            for col in ("A", "B", "C", "D"):
                cell = ws[f"{col}{row}"]
                cell.font = TOTAL_BURDEN_FONT
                cell.fill = TOTAL_BURDEN_FILL
                cell.border = top_border
        else:
            ws[f"A{row}"].font = BODY_FONT

    # v2.5 FB-5: dropped the green-savings / red-cost conditional formatting.
    # Sign convention is now consistent across the tab (Change = reset − default,
    # negative = saving), and the [Red] brackets in FMT_CURRENCY_DELTA give the
    # eye the same signal without per-cell CF.

    # v2.3 C-3: short italic explanation immediately under the subtotals block.
    note_row = SUBTOTAL_BURDEN_ROW + 1
    note = ws.cell(
        row=note_row, column=1,
        value=("Difference = If elected to reset − If no reset. A negative "
               "value (shown in green brackets) means the reset reduces that "
               "line of tax. Hover over each subtotal label for a plain-English "
               "definition."),
    )
    note.font = Font(name="Arial", size=9, italic=True, color="666666")
    note.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True, indent=1)
    ws.merge_cells(f"A{note_row}:{LAST_VISIBLE_COL}{note_row}")
    ws.row_dimensions[note_row].height = 16


def _build_per_member_breakdown(ws: Worksheet) -> None:
    """v2.3: 4-row per-member breakdown showing each member's TSB and their
    share of the Div 296 tax under both scenarios. Empty members render blank
    so the block degrades gracefully for single-member or 2-member funds.

    Data already exists in hidden cols L/M (per-member tax helpers); this block
    just surfaces it visibly.
    """
    _band(ws, PER_MEMBER_BAND_ROW, "Per-member breakdown")

    # Header row — v2.5 FB-6: labels standardised across all sections.
    headers = [
        ("A", "Member"),
        ("B", "TSB"),
        ("C", "If no reset (default)"),
        ("D", "If elected to reset"),
        ("E", "Difference"),
    ]
    for col, text in headers:
        c = ws[f"{col}{PER_MEMBER_HEADER_ROW}"]
        c.value = text
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL
        c.alignment = CENTER

    # Data rows — one per member. v2.5 FB-6: member labels always show as
    # placeholders ("Member 1".."Member 4") regardless of TSB, so an empty
    # Inputs page still presents the 4-member structure on Comparison.
    # Numeric cells stay blank when TSB is empty/zero.
    for i in range(ASSUMPTIONS.member_count):
        row = PER_MEMBER_FIRST_ROW + i
        inputs_row = MEMBERS_FIRST_DATA_ROW + i
        helper_row = HELPER_MEMBER_TAX_FIRST_ROW + i
        tsb_ref = f"{INPUTS_SHEET}!B{inputs_row}"
        tax_a_ref = f"{HELPER_COL_A}{helper_row}"
        tax_b_ref = f"{HELPER_COL_B}{helper_row}"

        # Col A — Member label, ALWAYS shown (placeholder for empty members).
        label = ws.cell(row=row, column=1, value=f"Member {i+1}")
        label.font = BODY_FONT
        label.alignment = Alignment(horizontal="left", indent=1)

        # Col B — TSB
        tsb = ws.cell(row=row, column=2,
                      value=f'=IF({tsb_ref}>0,{tsb_ref},"")')
        tsb.number_format = FMT_CURRENCY
        tsb.alignment = Alignment(horizontal="right")

        # Col C — If no reset (default) Div 296 tax
        ca = ws.cell(row=row, column=3,
                     value=f'=IF({tsb_ref}>0,{tax_a_ref},"")')
        ca.number_format = FMT_CURRENCY
        ca.alignment = Alignment(horizontal="right")

        # Col D — If elected to reset Div 296 tax
        cb = ws.cell(row=row, column=4,
                     value=f'=IF({tsb_ref}>0,{tax_b_ref},"")')
        cb.number_format = FMT_CURRENCY
        cb.alignment = Alignment(horizontal="right")

        # Col E — Change SIGNED (reset − default; negative = saving).
        chg = ws.cell(row=row, column=5,
                      value=f'=IF({tsb_ref}>0,{tax_b_ref}-{tax_a_ref},"")')
        chg.number_format = FMT_CURRENCY_DELTA
        chg.alignment = Alignment(horizontal="right")

        ws.row_dimensions[row].height = 18

    # v2.5 step 13 (Aiden polish): NEW Total row aggregating the four members
    # so the per-member breakdown reconciles visibly to the headline.
    total_font = Font(name="Arial", size=11, bold=True, color="1D3B34")
    total_fill = PatternFill("solid", fgColor="EFF5F3")
    total_label = ws.cell(row=PER_MEMBER_TOTAL_ROW, column=1, value="Total")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = Alignment(horizontal="left", indent=1)

    sum_first = PER_MEMBER_FIRST_ROW
    sum_last = PER_MEMBER_LAST_ROW
    # Use SUMIF so blank cells (empty members) don't propagate #VALUE! errors.
    totals = [
        (2, f'=SUMIF(B{sum_first}:B{sum_last},">0")',   FMT_CURRENCY),
        (3, f'=SUMIF(C{sum_first}:C{sum_last},">0")',   FMT_CURRENCY),
        (4, f'=SUMIF(D{sum_first}:D{sum_last},">0")',   FMT_CURRENCY),
        (5, f"=D{PER_MEMBER_TOTAL_ROW}-C{PER_MEMBER_TOTAL_ROW}", FMT_CURRENCY_DELTA),
    ]
    for col_idx, formula, fmt in totals:
        cell = ws.cell(row=PER_MEMBER_TOTAL_ROW, column=col_idx, value=formula)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[PER_MEMBER_TOTAL_ROW].height = 18


def _build_per_asset_detail(
    ws: Worksheet, gain_a_range: str, gain_b_range: str, delta_range: str,
    headline_a: str, headline_b: str,
) -> None:
    # v2.5 FB-7: band header reworded — emphasises tax impact (the new sort
    # metric, displayed in the Change column).
    _band(ws, BAND_DETAIL_ROW,
          f"Top {DISPLAY_ROWS} assets — Div 296 tax impact if you elect to reset")

    # Panel titles row — v2.5 FB-7: layout is now Panel A | Panel B | Change.
    panel_a_first, panel_a_last = PANEL_A_COLS[0], PANEL_A_COLS[-1]
    panel_b_first, panel_b_last = PANEL_B_COLS[0], PANEL_B_COLS[-1]
    panel_a_title = ws[f"{panel_a_first}{PANEL_TITLE_ROW}"]
    panel_a_title.value = "If no reset (default)"
    panel_a_title.font = SECTION_BAND_FONT
    panel_a_title.fill = SECTION_BAND_FILL
    panel_a_title.alignment = CENTER
    ws.merge_cells(f"{panel_a_first}{PANEL_TITLE_ROW}:{panel_a_last}{PANEL_TITLE_ROW}")

    panel_b_title = ws[f"{panel_b_first}{PANEL_TITLE_ROW}"]
    panel_b_title.value = "If elected to reset"
    panel_b_title.font = SECTION_BAND_FONT
    panel_b_title.fill = SECTION_BAND_FILL
    panel_b_title.alignment = CENTER
    ws.merge_cells(f"{panel_b_first}{PANEL_TITLE_ROW}:{panel_b_last}{PANEL_TITLE_ROW}")

    delta_title = ws[f"{DELTA_COL}{PANEL_TITLE_ROW}"]
    delta_title.value = "Difference"
    delta_title.font = DELTA_FONT
    delta_title.fill = DELTA_HEADER_FILL
    delta_title.alignment = CENTER

    # Column sub-headers (29)
    sub_headers = ["Asset", "Proceeds", "Div 296 cost base",
                   "Div 296 adj gain (info only)", "Div 296 tax"]
    for col, header in zip(PANEL_A_COLS, sub_headers):
        c = ws[f"{col}{PANEL_HEADER_ROW}"]
        c.value = header
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL
        c.alignment = CENTER
    for col, header in zip(PANEL_B_COLS, sub_headers):
        c = ws[f"{col}{PANEL_HEADER_ROW}"]
        c.value = header
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL
        c.alignment = CENTER
    # v2.5 FB-7: sub-header reflects the new tax-delta metric (was "gain change").
    delta_sub = ws[f"{DELTA_COL}{PANEL_HEADER_ROW}"]
    delta_sub.value = "Div 296 tax"
    delta_sub.font = DELTA_FONT
    delta_sub.fill = DELTA_HEADER_FILL
    delta_sub.alignment = CENTER

    # Per-row data formulas — sort by |Δ| descending via LARGE / MATCH / INDEX.
    # Empty register rows score -1 on the delta range; LARGE pushes them past
    # rank DISPLAY_ROWS so matched=0 → blank cells in the visible panel.
    asset_a, proc_a, cb_a, gain_a, tax_a = PANEL_A_COLS
    asset_b, proc_b, cb_b, gain_b, tax_b = PANEL_B_COLS

    for offset in range(DISPLAY_ROWS):
        c_row = DATA_FIRST_ROW + offset
        k = offset + 1   # rank 1..DISPLAY_ROWS
        matched = f"${MATCHED_ROW_COL}{c_row}"   # absolute col, this row

        # Hidden: the register row that fills this visible slot. 0 = no asset
        # (fewer assets than DISPLAY_ROWS, or LARGE landed on an empty marker).
        ws[f"{MATCHED_ROW_COL}{c_row}"] = (
            f'=IFERROR('
            f'IF(LARGE({delta_range},{k})<=0,0,'
            f'MATCH(LARGE({delta_range},{k}),{delta_range},0)+{REGISTER_FIRST_DATA_ROW - 1}),'
            f'0)'
        )

        def _input(col: str) -> str:
            return f"INDEX({INPUTS_SHEET}!${col}:${col},{matched})"

        # v2.3: Inputs col letters shifted — D→C (orig), F→E (MV), H→G (proceeds)
        a_code  = _input("A")
        a_name  = _input("B")
        a_orig  = _input("C")   # was D
        a_mv    = _input("E")   # was F
        a_proc  = _input("G")   # was H

        # Pre-computed gains from the per-register grid (cols N/O).
        gain_a_lookup = f"INDEX(${PER_REG_GAIN_A_COL}:${PER_REG_GAIN_A_COL},{matched})"
        gain_b_lookup = f"INDEX(${PER_REG_GAIN_B_COL}:${PER_REG_GAIN_B_COL},{matched})"

        # Panel A (no reset → cost base = original)
        # v2.4 FB-2: "{code} - {name}" (was "{name} ({code})")
        ws[f"{asset_a}{c_row}"] = f'=IF({matched}=0,"",{a_code}&" - "&{a_name})'
        ws[f"{proc_a}{c_row}"]  = f'=IF({matched}=0,"",{a_proc})'
        ws[f"{cb_a}{c_row}"]    = f'=IF({matched}=0,"",{a_orig})'
        ws[f"{gain_a}{c_row}"]  = f'=IF({matched}=0,"",{gain_a_lookup})'
        ws[f"{tax_a}{c_row}"]   = (
            f'=IF({matched}=0,"",'
            f'IF(SUMIF({gain_a_range},">0")=0,0,'
            f'MAX(0,{gain_a}{c_row})/SUMIF({gain_a_range},">0")*{headline_a}))'
        )

        # Panel B (reset elected → cost base = MV)
        # v2.4 FB-2: "{code} - {name}" (was "{name} ({code})")
        ws[f"{asset_b}{c_row}"] = f'=IF({matched}=0,"",{a_code}&" - "&{a_name})'
        ws[f"{proc_b}{c_row}"]  = f'=IF({matched}=0,"",{a_proc})'
        ws[f"{cb_b}{c_row}"]    = f'=IF({matched}=0,"",{a_mv})'
        ws[f"{gain_b}{c_row}"]  = f'=IF({matched}=0,"",{gain_b_lookup})'
        ws[f"{tax_b}{c_row}"]   = (
            f'=IF({matched}=0,"",'
            f'IF(SUMIF({gain_b_range},">0")=0,0,'
            f'MAX(0,{gain_b}{c_row})/SUMIF({gain_b_range},">0")*{headline_b}))'
        )

        # v2.5 FB-7: Change column at the FAR RIGHT, signed TAX delta
        # (reset tax − default tax). Negative = reset reduces tax (green brackets).
        ws[f"{DELTA_COL}{c_row}"] = (
            f'=IF({matched}=0,"",{tax_b}{c_row}-{tax_a}{c_row})'
        )

        # Currency formatting on all numeric cells.
        for col in (proc_a, cb_a, gain_a, tax_a, proc_b, cb_b, gain_b, tax_b):
            ws[f"{col}{c_row}"].number_format = FMT_CURRENCY
        # Signed delta format for the Change col.
        ws[f"{DELTA_COL}{c_row}"].number_format = FMT_CURRENCY_DELTA

        # v2.3 C-4: per-column fills mirroring Analyser's per-asset table
        # (sand for proceeds, slate for ordinary/cost, sage for Div 296).
        # Applied to both Panel A and Panel B so the eye sweeps left-to-right
        # by data class, not by panel.
        ws[f"{proc_a}{c_row}"].fill = PROC_DATA_FILL
        ws[f"{cb_a}{c_row}"].fill = ORD_DATA_FILL
        ws[f"{gain_a}{c_row}"].fill = DIV_DATA_FILL
        ws[f"{tax_a}{c_row}"].fill = DIV_DATA_FILL
        ws[f"{proc_b}{c_row}"].fill = PROC_DATA_FILL
        ws[f"{cb_b}{c_row}"].fill = ORD_DATA_FILL
        ws[f"{gain_b}{c_row}"].fill = DIV_DATA_FILL
        ws[f"{tax_b}{c_row}"].fill = DIV_DATA_FILL

    # E3: highlight Div 296 cost base in panel B when it differs from panel A.
    # CF formula uses RELATIVE rows so it shifts per row in the range.
    cb_b_range = f"{PANEL_B_COLS[2]}{DATA_FIRST_ROW}:{PANEL_B_COLS[2]}{DATA_LAST_ROW}"
    diff_rule = FormulaRule(
        formula=[f'AND({PANEL_A_COLS[2]}{DATA_FIRST_ROW}<>"",'
                 f'{PANEL_B_COLS[2]}{DATA_FIRST_ROW}<>'
                 f'{PANEL_A_COLS[2]}{DATA_FIRST_ROW})'],
        fill=COST_BASE_ACCENT_FILL,
    )
    ws.conditional_formatting.add(cb_b_range, diff_rule)

    # v2.3 C-4: trap red-row CF — fires when the asset is in an unrealised
    # loss position (Panel A proceeds < cost base) AND the reset turns it
    # into a positive Div 296 gain (Panel B gain > 0). v2.5 FB-7: column
    # refs made symbolic (was hardcoded "$I" which referenced the wrong
    # column under the old layout; new symbolic refs auto-track the layout).
    panel_b_gain_col = PANEL_B_COLS[3]   # "I" under the v2.5 layout
    trap_range = (
        f"{PANEL_A_COLS[0]}{DATA_FIRST_ROW}:"
        f"{LAST_VISIBLE_COL}{DATA_LAST_ROW}"
    )
    trap_rule = FormulaRule(
        formula=[
            f"AND($A{DATA_FIRST_ROW}<>\"\","
            f"($B{DATA_FIRST_ROW}-$C{DATA_FIRST_ROW})<0,"
            f"${panel_b_gain_col}{DATA_FIRST_ROW}>0)"
        ],
        fill=TRAP_FILL,
    )
    ws.conditional_formatting.add(trap_range, trap_rule)

    # v2.5 FB-7 dropped the per-cell green/red CF here in favour of the
    # FMT_CURRENCY_DELTA format (negatives render as green brackets). v3.4
    # audit 3.1 re-adds a ONE-directional red highlight on POSITIVE (cost)
    # Change values via _build_difference_red_cf — the trap this model exists
    # to surface — while negatives keep the green-bracket format.

    # Overflow note — v2.5 FB-7: mentions tax delta (the new sort metric).
    overflow = ws.cell(
        row=DATA_OVERFLOW_NOTE_ROW, column=1,
        value=(f"Showing top {DISPLAY_ROWS} assets by absolute Div 296 tax "
               f"change when the reset is elected — see the Analyser tab for "
               f"the full register (up to {ASSUMPTIONS.asset_register_rows} rows)."),
    )
    overflow.font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.merge_cells(f"A{DATA_OVERFLOW_NOTE_ROW}:{LAST_VISIBLE_COL}{DATA_OVERFLOW_NOTE_ROW}")

    # v3.4 audit F4: error tripwire in the spacer row above the detail band.
    # Every register row scores numerically (incomplete/empty rows score -1
    # after the v3.4 blank guards), so COUNT < register depth means some row
    # ERRORED (e.g. text pasted into a numeric Inputs column) — in which case
    # LARGE() errors for every rank and the panel below renders silently empty.
    err_check = ws.cell(
        row=BAND_DETAIL_ROW - 1, column=1,
        value=(f'=IF(COUNT({delta_range})<{ASSUMPTIONS.asset_register_rows},'
               f'"⚠ Some register rows could not be evaluated — check for text in numeric '
               f'columns on Inputs. The asset ranking below may be incomplete.","")'),
    )
    err_check.font = Font(name="Arial", size=9, bold=True, italic=True, color="A61B1B")


def _build_footer_notes(ws: Worksheet) -> None:
    reminder = ws.cell(
        row=REMINDER_ROW, column=1,
        value=("Note: assets currently in an unrealised-loss position may contribute "
               "Div 296 tax IF you elect the reset that they do not contribute under "
               "the default outcome — see the Analyser tab for per-asset detail."),
    )
    reminder.font = Font(name="Arial", size=9, italic=True, color="666666")
    reminder.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{REMINDER_ROW}:{LAST_VISIBLE_COL}{REMINDER_ROW}")
    ws.row_dimensions[REMINDER_ROW].height = 24

    sort_note = ws.cell(
        row=SORT_NOTE_ROW, column=1,
        value=("Note: per-asset detail shows the top 10 assets where the reset "
               "election moves the Div 296 tax the most (by absolute change, "
               "either direction). See the Analyser tab for the full register."),
    )
    sort_note.font = Font(name="Arial", size=9, italic=True, color="666666")
    sort_note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{SORT_NOTE_ROW}:{LAST_VISIBLE_COL}{SORT_NOTE_ROW}")
    ws.row_dimensions[SORT_NOTE_ROW].height = 30


def _build_difference_red_cf(ws: Worksheet) -> None:
    """v3.4 audit 3.1: a POSITIVE Difference means the reset COSTS money (the
    trap this model exists to surface) — render it muted red, mirroring the
    Analyser fund-summary rule. Negative (a saving) keeps the green-bracket
    FMT_CURRENCY_DELTA styling. ISNUMBER guards skip blank/empty-member cells.
    """
    red = Font(name="Arial", size=10, bold=True, color="A61B1B")
    ranges = (
        f"I{CARD_VALUE_ROW}",                                       # Net-effect card
        f"D{SUBTOTAL_EARNINGS_ROW}:D{SUBTOTAL_BURDEN_ROW}",         # subtotal Difference
        f"E{PER_MEMBER_FIRST_ROW}:E{PER_MEMBER_LAST_ROW}",          # per-member Difference
        f"{DELTA_COL}{DATA_FIRST_ROW}:{DELTA_COL}{DATA_LAST_ROW}",  # per-asset Change
    )
    for rng in ranges:
        top = rng.split(":")[0]   # relative top-cell ref; rule auto-tracks down the range
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f"AND(ISNUMBER({top}),{top}>0)"], font=red),
        )


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False

    _build_header_block(ws)

    # Section band for the body — v2.5 polish: renamed to match what's actually
    # underneath (per-member TSB table) now that the right-side tiles are gone.
    _band(ws, BAND_BODY_ROW, "Members & TSB")

    _build_context_strip(ws)

    gain_a_range, gain_b_range, delta_range = _build_per_register_helpers(ws)
    headline_a, headline_b = _build_helpers(ws, gain_a_range, gain_b_range)

    _build_metric_cards(ws, headline_a, headline_b)
    _build_subtotals(ws, headline_a, headline_b, gain_a_range, gain_b_range)
    _build_per_member_breakdown(ws)
    _build_per_asset_detail(
        ws, gain_a_range, gain_b_range, delta_range, headline_a, headline_b,
    )
    _build_footer_notes(ws)
    _build_difference_red_cf(ws)

    # --- Print header watermark (large grey text on every printed page) ---
    ws.oddHeader.center.text = "ILLUSTRATIVE — NOT ADVICE"
    ws.oddHeader.center.size = 28
    ws.oddHeader.center.color = "CCCCCC"

    # --- Print setup: A4 landscape, narrow margins; let height spill if needed ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    # v2.5 FB-3: chart removed — single landscape A4 page is now enough.
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.4
    ws.print_area = f"A1:{LAST_VISIBLE_COL}{SORT_NOTE_ROW}"

    # --- Column widths (v2.5 FB-7 + FB-9 layout) ---
    # A-E  Panel A + subtotals: A holds long labels; B-D wider for the new
    #      "If no reset (default)" / "If elected to reset" headers (used by
    #      subtotals + per-member breakdown); E carries Change for subtotals
    #      and Tax for Panel A.
    # F-J  Panel B (asset, proceeds, cost base, gain, tax) — same shape as A.
    # K    Change column (signed tax delta) — rightmost; bracket-friendly width.
    # v2.5 step 13 (Aiden polish): D and I narrowed from 20/16 → 13 so the
    # subtotal and Panel B "cost base" cols match the eye-width of the Change
    # column rather than dominating the table.
    widths = {
        "A": 36, "B": 16, "C": 20, "D": 13, "E": 16,    # Panel A + subtotals
        "F": 26, "G": 14, "H": 16, "I": 13, "J": 14,    # Panel B
        "K": 16,                                         # Change
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Editable header-block cells stay editable when the sheet is protected.
    for row in (HEADER_FIRM_ROW, HEADER_PREPARED_FOR_ROW,
                HEADER_PREPARED_BY_ROW, HEADER_DATE_ROW):
        ws[f"B{row}"].protection = Protection(locked=False)

    # v2.5 step 13: re-lock the Comparison tab now that Aiden's formatting pass
    # has been ported into source. Match other tabs: sheet protected, but
    # column/row resizing allowed so the user can still tweak layout.
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws
