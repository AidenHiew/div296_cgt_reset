"""Inputs tab — the only data-entry sheet.

v3.0 layout — Control panel removed (toggles always-on internally per
Bill-correct calc; see CONTEXT.md v3.0 cut-over note). Sections renumbered.

    Row 1       Title
    Row 2       Sample-data badge (warns staff before client share)
    Row 3       TSB diagnostic banner (traffic light — green/amber/deep-amber)
    Row 5       Band: 1. Members
    Row 6       Members header row
    Rows 7-10   4 member rows
    Row 11      Members TSB total row
    Row 12      Split-% sum check (tautological since v2.4 split is auto-derived;
                kept as a visible 100% confirmation)
    Row 14      Band: 2. Asset register
    Row 15      Register header row
    Rows 16-65  50 register data rows
    Row 67      Band: 3. Advanced assumptions
    Rows 68-75  8 assumption rows (named ranges anchored here)

Section 2 (Members) columns (v3.0):
    A: Member label
    B: TSB ($)                                       — user input
    C: Split % of fund earnings (auto)               — TSB / sum(TSB)
    D: Proportion in $3m–$10m band (auto) = band1    — MAX(0, MIN(TSB,$10m)-$3m)/TSB
    E: Proportion above $10m (auto) = band2          — MAX(0, TSB-$10m)/TSB

band1 + band2 = total proportion of TSB above $3m. The two-band display
mirrors the calc-engine's actual structure 1:1 — a reviewer can manually
compute tax = earnings × split × (D × rate_tier1 + E × rate_tier2) and
reconcile against the workbook.

Downstream tabs read this layout via the constants below — do not
change row numbers without grepping for the constant names first.
"""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from div296 import named_ranges as nr
from div296.assumptions import ASSUMPTIONS
from div296.styles import (
    BODY_FONT, CENTER, FMT_CURRENCY, FMT_PERCENT, FMT_PERCENT_3,
    FMT_TEXT, INPUT_FILL, INPUT_FONT, SECTION_BAND_FILL,
    SECTION_BAND_FONT, THIN_BOX, TITLE_FONT,
)


# Cell-layout constants — single source of truth.
SHEET = "Inputs"

# --- Zone 1: Members (rows 5-12) ---
MEMBERS_HEADER_ROW = 6
MEMBERS_FIRST_DATA_ROW = 7
MEMBER_HEADERS = [
    "Member",
    "TSB ($)",
    "Split % of fund earnings (auto)",         # v2.4: TSB-derived formula
    "Proportion in $3m–$10m band (auto)",      # v3.0: band1
    "Proportion above $10m (auto)",            # v3.0: band2
]
MEMBERS_TOTAL_ROW = MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count       # row 11
SPLIT_CHECK_ROW = MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count + 1     # row 12

# --- Zone 2: Asset register (rows 14-65) ---
REGISTER_HEADER_ROW = 15
REGISTER_FIRST_DATA_ROW = 16
REGISTER_LAST_DATA_ROW = REGISTER_FIRST_DATA_ROW + ASSUMPTIONS.asset_register_rows - 1   # row 65

REGISTER_HEADERS = [
    ("Asset code", FMT_TEXT),
    ("Asset name", FMT_TEXT),
    ("Original cost base", FMT_CURRENCY),
    ("Current market value (as at today)", FMT_CURRENCY),
    ("Market value at 30 Jun 2026", FMT_CURRENCY),
    ("Valuation source / date", FMT_TEXT),
    ("Projected sale proceeds", FMT_CURRENCY),
    ("Projected gain/loss", FMT_CURRENCY),               # formula = proceeds − original CB
    ("Held > 12 months?", FMT_TEXT),
]

# Indices (1-based) for special-case col handling in the register loop.
REGISTER_COL_ORIG_CB = 3
REGISTER_COL_PROCEEDS = 7
REGISTER_COL_PROJ_GL = 8
REGISTER_COL_HELD = 9

SAMPLE_REGISTER_ROWS = [
    ("P1", "Commercial property",   800_000, 2_400_000, 2_400_000,
     "Independent val, 30/06/26",   2_600_000, "Yes"),
    ("S1", "Listed shares parcel",  300_000, 520_000,   520_000,
     "ASX close 30/06/26",          600_000,   "Yes"),
    ("L1", "Loss-making holding",   500_000, 100_000,   100_000,
     "Independent val, 30/06/26",   200_000,   "Yes"),
]

# --- Zone 3: Advanced assumptions (rows 67-75) ---
ADV_BAND_ROW = REGISTER_LAST_DATA_ROW + 2     # row 67
ADV_FIRST_ROW = ADV_BAND_ROW + 1              # row 68

ADV_ROWS = [
    (nr.RATE_TIER1,        "Div 296 additional rate — tier 1 ($3m–$10m)",   ASSUMPTIONS.rate_tier1,             FMT_PERCENT),
    (nr.RATE_TIER2,        "Div 296 additional rate — tier 2 (above $10m)", ASSUMPTIONS.rate_tier2,             FMT_PERCENT),
    (nr.THRESHOLD_1,       "Threshold 1",                                   ASSUMPTIONS.threshold_1,            FMT_CURRENCY),
    (nr.THRESHOLD_2,       "Threshold 2",                                   ASSUMPTIONS.threshold_2,            FMT_CURRENCY),
    (nr.DISCOUNT_RATE,     "CGT discount rate (1/3 = 33.333%)",             ASSUMPTIONS.discount_rate,          FMT_PERCENT_3),
    (nr.FUND_CGT_RATE,     "Fund CGT rate (accumulation phase)",            ASSUMPTIONS.fund_cgt_rate,          FMT_PERCENT),
    (nr.INDEXATION_INCR_1, "Indexation increment — threshold 1",            ASSUMPTIONS.indexation_increment_1, FMT_CURRENCY),
    (nr.INDEXATION_INCR_2, "Indexation increment — threshold 2",            ASSUMPTIONS.indexation_increment_2, FMT_CURRENCY),
]


def _band(ws: Worksheet, row: int, text: str, last_col_letter: str = "I") -> None:
    ws.cell(row=row, column=1, value=text).font = SECTION_BAND_FONT
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    for col_idx in range(1, ord(last_col_letter) - ord("A") + 2):
        ws.cell(row=row, column=col_idx).fill = SECTION_BAND_FILL


def _input_cell(ws: Worksheet, coord: str, value=None, number_format: str | None = None) -> None:
    cell = ws[coord]
    if value is not None:
        cell.value = value
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = THIN_BOX
    # Unlocked so it stays editable when the sheet is protected.
    cell.protection = Protection(locked=False)
    if number_format:
        cell.number_format = number_format


def _define_name(wb: Workbook, name: str, coord: str) -> None:
    """Create a workbook-scoped defined name pointing at Inputs!coord."""
    ref = f"'{SHEET}'!${coord[0]}${coord[1:]}" if coord[1:].isdigit() else f"'{SHEET}'!{coord}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)

    # --- Row 1: Title ---
    ws["A1"] = "Division 296 Cost Base Reset Model — Inputs"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    # --- Row 2: Sample-data badge ---
    badge = ws.cell(
        row=2, column=1,
        value=("⚠  Sample data preloaded — overwrite every cell with your fund's "
               "actual figures before sharing with a client."),
    )
    badge.font = Font(name="Arial", size=10, bold=True, italic=True, color="8A6D00")
    badge.fill = PatternFill("solid", fgColor="FFF4CE")
    badge.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 22

    # --- Row 3: TSB diagnostic — three-tier traffic light ---
    # v3.0: dropped the "consider enabling the tier toggle" nudge since the
    # toggle no longer exists. Deep-amber message now states the always-on
    # two-tier behaviour directly.
    last_member_row = MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count - 1
    max_tsb_expr = f"MAX(B{MEMBERS_FIRST_DATA_ROW}:B{last_member_row})"
    diag_formula = (
        f'=IF({max_tsb_expr}=0,'
        f'"➤  Enter member TSBs below to see whether Div 296 applies to this fund.",'
        f'IF({max_tsb_expr}<=threshold_1,'
        f'"✓  All members are below the $3m TSB threshold — Div 296 does not currently apply to this fund.",'
        f'IF({max_tsb_expr}<=threshold_2,'
        f'"⚠  At least one member has a TSB above $3m — Div 296 applies at 15% on the slice above $3m. See the Comparison tab for the modelled impact.",'
        f'"⚠⚠  At least one member has a TSB above $10m — Div 296 applies at 15% on the $3m–$10m band and 25% on the slice above $10m.")))'
    )
    diag = ws.cell(row=3, column=1, value=diag_formula)
    diag.font = Font(name="Arial", size=10, bold=True, color="666666")
    diag.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A3:I3")
    ws.row_dimensions[3].height = 22

    # Green when no member exceeds $3m.
    green_rule = FormulaRule(
        formula=[f"AND({max_tsb_expr}>0,{max_tsb_expr}<=threshold_1)"],
        fill=PatternFill("solid", fgColor="E1F5EE"),
    )
    # Light amber when at least one member is above $3m but below $10m.
    amber_rule = FormulaRule(
        formula=[f"AND({max_tsb_expr}>threshold_1,{max_tsb_expr}<=threshold_2)"],
        fill=PatternFill("solid", fgColor="FBE9E9"),
    )
    # Darker amber when at least one member is above $10m.
    deep_amber_rule = FormulaRule(
        formula=[f"{max_tsb_expr}>threshold_2"],
        fill=PatternFill("solid", fgColor="F4C28A"),
    )
    ws.conditional_formatting.add("A3:I3", green_rule)
    ws.conditional_formatting.add("A3:I3", amber_rule)
    ws.conditional_formatting.add("A3:I3", deep_amber_rule)

    # v3.0: The "tier10 nudge" CF on B6 that pulsed amber when MAX TSB > $10m
    # and the toggle was OFF is removed — B6 no longer exists.

    # --- Zone 1: Members ---
    _band(ws, MEMBERS_HEADER_ROW - 1, "1. Members")
    for col_idx, header in enumerate(MEMBER_HEADERS, start=1):
        c = ws.cell(row=MEMBERS_HEADER_ROW, column=col_idx, value=header)
        c.font = SECTION_BAND_FONT
        c.fill = SECTION_BAND_FILL
        c.alignment = CENTER

    last_mr = MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count - 1
    tsb_sum = f"SUM($B${MEMBERS_FIRST_DATA_ROW}:$B${last_mr})"
    # v2.5 FB-2: seed two members (Aiden's feedback — show split-balance behaviour).
    # Member 1 = $12m (§12 high-TSB case), Member 2 = $3.5m (just-above-threshold case).
    # Members 3–4 stay blank as illustrative placeholders.
    SAMPLE_TSB_BY_MEMBER = (12_000_000, 3_500_000, None, None)
    for i in range(ASSUMPTIONS.member_count):
        row = MEMBERS_FIRST_DATA_ROW + i
        ws.cell(row=row, column=1, value=f"Member {i+1}").font = BODY_FONT
        seed = SAMPLE_TSB_BY_MEMBER[i] if i < len(SAMPLE_TSB_BY_MEMBER) else None
        _input_cell(ws, f"B{row}", value=seed, number_format=FMT_CURRENCY)

        # Col C — auto-derived split %. Default-locked under sheet protection,
        # styled italic teal-grey to read as derived.
        split_formula = f"=IF({tsb_sum}>0,B{row}/{tsb_sum},0)"
        split_cell = ws.cell(row=row, column=3, value=split_formula)
        split_cell.number_format = FMT_PERCENT
        split_cell.font = Font(name="Arial", size=10, italic=True, color="1D3B34")
        split_cell.alignment = Alignment(horizontal="right", indent=1)

        # v3.0 Col D — band1 = proportion of TSB in the $3m–$10m slice.
        band1_formula = f"=IF(B{row}>0, MAX(0, MIN(B{row}, threshold_2) - threshold_1) / B{row}, 0)"
        band1_cell = ws.cell(row=row, column=4, value=band1_formula)
        band1_cell.number_format = FMT_PERCENT
        band1_cell.font = Font(name="Arial", size=10, italic=True, color="1D3B34")
        band1_cell.alignment = Alignment(horizontal="right", indent=1)

        # v3.0 Col E — band2 = proportion of TSB above $10m.
        band2_formula = f"=IF(B{row}>0, MAX(0, (B{row} - threshold_2)) / B{row}, 0)"
        band2_cell = ws.cell(row=row, column=5, value=band2_formula)
        band2_cell.number_format = FMT_PERCENT
        band2_cell.font = Font(name="Arial", size=10, italic=True, color="1D3B34")
        band2_cell.alignment = Alignment(horizontal="right", indent=1)

    # Member TSB total row — combined TSB across members for quick read.
    last_member_data_row = MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count - 1
    total_label = ws.cell(row=MEMBERS_TOTAL_ROW, column=1, value="Total")
    total_label.font = Font(name="Arial", size=10, bold=True, color="1D3B34")
    total_label.alignment = Alignment(horizontal="left", indent=1)
    total_tsb = ws.cell(
        row=MEMBERS_TOTAL_ROW, column=2,
        value=f"=SUM(B{MEMBERS_FIRST_DATA_ROW}:B{last_member_data_row})",
    )
    total_tsb.font = Font(name="Arial", size=10, bold=True, color="1D3B34")
    total_tsb.number_format = FMT_CURRENCY
    total_tsb.border = THIN_BOX
    total_tsb.fill = PatternFill("solid", fgColor="EFF5F3")  # light teal band
    # Apply the same light teal band to cols C-E for visual continuity.
    for col_idx in range(3, 6):
        c = ws.cell(row=MEMBERS_TOTAL_ROW, column=col_idx)
        c.fill = PatternFill("solid", fgColor="EFF5F3")

    # Member-split sanity row. Tautological since v2.4 (split% is auto-derived)
    # but kept as a visible 100% confirmation; gives the user a green light that
    # the workbook is in a valid TSB-populated state.
    ws.cell(row=SPLIT_CHECK_ROW, column=1, value="Split % sum (must equal 100%)").font = BODY_FONT
    check_formula = (
        f"=SUM(C{MEMBERS_FIRST_DATA_ROW}:C{MEMBERS_FIRST_DATA_ROW + ASSUMPTIONS.member_count - 1})"
    )
    chk = ws.cell(row=SPLIT_CHECK_ROW, column=3, value=check_formula)
    chk.number_format = FMT_PERCENT
    # Red when not 100%, green when 100%.
    chk_range = f"C{SPLIT_CHECK_ROW}"
    ws.conditional_formatting.add(
        chk_range,
        CellIsRule(operator="notEqual", formula=["1"], fill=PatternFill("solid", fgColor="FBE9E9")),
    )
    ws.conditional_formatting.add(
        chk_range,
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="E1F5EE")),
    )

    # --- Zone 2: Asset register (50 rows; sample data preloaded) ---
    _band(
        ws, REGISTER_HEADER_ROW - 1,
        f"2. Asset register (50 rows; sample data pre-loaded in rows "
        f"{REGISTER_FIRST_DATA_ROW}–{REGISTER_FIRST_DATA_ROW + len(SAMPLE_REGISTER_ROWS) - 1})",
    )

    # Header row coloured by column group for visual grouping
    # (Identity / Cost & value / Sale & CGT outcome).
    REG_HEADER_GROUPS = {
        # col_idx (1-based): group_fill
        1: SECTION_BAND_FILL,           # Identity
        2: SECTION_BAND_FILL,           # Identity
        3: PatternFill("solid", fgColor="3F4F4A"),  # Cost & value (slate)
        4: PatternFill("solid", fgColor="3F4F4A"),
        5: PatternFill("solid", fgColor="3F4F4A"),
        6: PatternFill("solid", fgColor="3F4F4A"),
        7: PatternFill("solid", fgColor="4A6B5E"),  # Sale & CGT outcome (sage-teal)
        8: PatternFill("solid", fgColor="4A6B5E"),
        9: PatternFill("solid", fgColor="4A6B5E"),
    }
    for col_idx, (header, _fmt) in enumerate(REGISTER_HEADERS, start=1):
        c = ws.cell(row=REGISTER_HEADER_ROW, column=col_idx, value=header)
        c.font = SECTION_BAND_FONT
        c.fill = REG_HEADER_GROUPS.get(col_idx, SECTION_BAND_FILL)
        c.alignment = CENTER

    # Held>12mo dropdown shared across all 50 rows.
    held_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(held_dv)

    for offset in range(ASSUMPTIONS.asset_register_rows):
        row = REGISTER_FIRST_DATA_ROW + offset
        sample = SAMPLE_REGISTER_ROWS[offset] if offset < len(SAMPLE_REGISTER_ROWS) else None
        # Sample data covers cols 1..7 + 9 (skip the formula-driven Projected G/L).
        # When walking REGISTER_HEADERS, the sample tuple index trails by 1 once
        # we pass col 8 (Projected G/L).
        for col_idx, (_header, fmt) in enumerate(REGISTER_HEADERS, start=1):
            coord = ws.cell(row=row, column=col_idx).coordinate
            if col_idx == REGISTER_COL_PROJ_GL:
                # Formula: proceeds (col G) − original cost base (col C).
                # Blank if either source cell is empty.
                value = (
                    f'=IF(AND(C{row}<>"",G{row}<>""),G{row}-C{row},"")'
                )
                _input_cell(ws, coord, value=value, number_format=fmt)
            else:
                # Sample data indexing: cols < PROJ_GL use col_idx-1 directly;
                # cols > PROJ_GL shift back by one to skip the formula col.
                if sample is None:
                    value = None
                elif col_idx < REGISTER_COL_PROJ_GL:
                    value = sample[col_idx - 1]
                else:
                    value = sample[col_idx - 2]
                _input_cell(ws, coord, value=value, number_format=fmt)
        held_dv.add(f"I{row}")

    # CF for Projected gain/loss column — green text for gain, red text for loss.
    pl_first = REGISTER_FIRST_DATA_ROW
    pl_last = REGISTER_LAST_DATA_ROW
    gl_range = f"H{pl_first}:H{pl_last}"
    ws.conditional_formatting.add(
        gl_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(H{pl_first}),H{pl_first}>0)'],
            font=Font(name="Arial", size=10, color="0B6E4F"),
        ),
    )
    ws.conditional_formatting.add(
        gl_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(H{pl_first}),H{pl_first}<0)'],
            font=Font(name="Arial", size=10, color="A61B1B"),
        ),
    )

    # Loss-position highlight — tint any register row whose current market value
    # (col D) is below its original cost base (col C). Cosmetic only.
    reg_first = REGISTER_FIRST_DATA_ROW
    reg_last = REGISTER_LAST_DATA_ROW
    loss_rule = FormulaRule(
        formula=[f'AND($C{reg_first}<>"",$D{reg_first}<>"",$D{reg_first}<$C{reg_first})'],
        fill=PatternFill("solid", fgColor="FBE9E9"),
    )
    ws.conditional_formatting.add(f"A{reg_first}:I{reg_last}", loss_rule)

    # Alternating row shading (zebra) for readability.
    zebra_rule = FormulaRule(
        formula=['MOD(ROW(),2)=0'],
        fill=PatternFill("solid", fgColor="F7F9F8"),
    )
    ws.conditional_formatting.add(f"A{reg_first}:I{reg_last}", zebra_rule)

    # --- Zone 3: Advanced assumptions (set-once constants — bottom) ---
    _band(ws, ADV_BAND_ROW, "3. Advanced assumptions (set once)")
    for i, (name, label, value, fmt) in enumerate(ADV_ROWS):
        row = ADV_FIRST_ROW + i
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        coord = f"B{row}"
        _input_cell(ws, coord, value=value, number_format=fmt)
        _define_name(wb, name, coord)

    # --- Column widths + freeze ---
    # 9 cols: A Code 32 / B Name 26 / C Orig CB 18 / D MV today 18 / E MV 30Jun 22
    # F Val source 24 / G Proceeds 18 / H Projected G/L 18 / I Held>12m 16
    widths = [32, 26, 18, 18, 22, 24, 18, 18, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    # Freeze below the title + sample badge + TSB diagnostic so they remain sticky.
    ws.freeze_panes = "A4"

    # --- Sheet protection (tamper-evident, passwordless) ---
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    return ws
