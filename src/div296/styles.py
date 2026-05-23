"""Colours, fills, fonts, number formats per spec §11."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Colour palette (spec §11) ---
COLOUR_BAND_DARK_TEAL = "1D3B34"
COLOUR_INPUT_TEXT_BLUE = "0000FF"
COLOUR_CROSS_SHEET_GREEN = "008000"
COLOUR_INPUT_FILL_GREEN = "E1F5EE"
COLOUR_TRAP_FILL = "FBE9E9"
COLOUR_TRAP_TEXT = "A32D2D"
COLOUR_WIN_TEXT = "0F6E56"
COLOUR_COST_BASE_HEADER = "185FA5"
COLOUR_DIV296_HEADER = "0F6E56"

# --- v2.0.0: per-asset column-group palette (Analyser tab) ---
COLOUR_PROC_HEADER = "D7CBB8"   # sand — Projected sale proceeds (raw input)
COLOUR_PROC_DATA   = "F7F4EE"   # 5% sand wash
COLOUR_ORD_HEADER  = "C9D5DA"   # slate — Ordinary CGT group
COLOUR_ORD_DATA    = "F4F6F7"   # 5% slate wash
COLOUR_DIV_HEADER  = "B0C9BD"   # sage-teal — Div 296 group
COLOUR_DIV_DATA    = "F2F6F4"   # 5% sage wash
COLOUR_RESET_HEADER = "C7A752"  # gold — Reset impact column header
COLOUR_RESET_DATA   = "FFF8E6"  # soft gold — Reset impact data cells
COLOUR_STATE_STRIP_FILL = "EFF5F3"   # light teal — Analyser state strip background
COLOUR_TOTALS_FILL = "E6ECEA"        # medium teal — distinct totals row fill
COLOUR_ROWNUM_FILL = "F2F2F2"        # very light grey — row-number column fill
COLOUR_ROWNUM_TEXT = "888888"        # muted — row-number column text

# --- Fonts ---
FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=COLOUR_BAND_DARK_TEAL)
SECTION_BAND_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10)
INPUT_FONT = Font(name=FONT_NAME, size=10, color=COLOUR_INPUT_TEXT_BLUE)
CROSS_SHEET_FONT = Font(name=FONT_NAME, size=10, color=COLOUR_CROSS_SHEET_GREEN)
TRAP_FONT = Font(name=FONT_NAME, size=10, color=COLOUR_TRAP_TEXT, bold=True)
WIN_FONT = Font(name=FONT_NAME, size=10, color=COLOUR_WIN_TEXT, bold=True)

# --- v2.0.0 fonts ---
RESET_HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
GROUP_HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color=COLOUR_BAND_DARK_TEAL)
ROWNUM_FONT       = Font(name=FONT_NAME, size=9, color=COLOUR_ROWNUM_TEXT)
TOTALS_FONT       = Font(name=FONT_NAME, size=11, bold=True, color=COLOUR_BAND_DARK_TEAL)

# --- Fills ---
SECTION_BAND_FILL = PatternFill("solid", fgColor=COLOUR_BAND_DARK_TEAL)
INPUT_FILL = PatternFill("solid", fgColor=COLOUR_INPUT_FILL_GREEN)
TRAP_FILL = PatternFill("solid", fgColor=COLOUR_TRAP_FILL)

# --- v2.0.0 fills ---
PROC_HEADER_FILL = PatternFill("solid", fgColor=COLOUR_PROC_HEADER)
PROC_DATA_FILL   = PatternFill("solid", fgColor=COLOUR_PROC_DATA)
ORD_HEADER_FILL  = PatternFill("solid", fgColor=COLOUR_ORD_HEADER)
ORD_DATA_FILL    = PatternFill("solid", fgColor=COLOUR_ORD_DATA)
DIV_HEADER_FILL  = PatternFill("solid", fgColor=COLOUR_DIV_HEADER)
DIV_DATA_FILL    = PatternFill("solid", fgColor=COLOUR_DIV_DATA)
RESET_HEADER_FILL = PatternFill("solid", fgColor=COLOUR_RESET_HEADER)
RESET_DATA_FILL   = PatternFill("solid", fgColor=COLOUR_RESET_DATA)
STATE_STRIP_FILL  = PatternFill("solid", fgColor=COLOUR_STATE_STRIP_FILL)
TOTALS_FILL       = PatternFill("solid", fgColor=COLOUR_TOTALS_FILL)
ROWNUM_FILL       = PatternFill("solid", fgColor=COLOUR_ROWNUM_FILL)

# --- Alignment ---
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# --- Borders ---
THIN = Side(style="thin", color="BFBFBF")
THIN_BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- Number formats ---
FMT_CURRENCY = '$#,##0;($#,##0);"-"'
FMT_PERCENT = "0.0%"
FMT_PERCENT_3 = "0.000%"     # for the CGT discount rate (spec calls for 33.333%)
FMT_INT = "#,##0"
FMT_TEXT = "@"               # explicit text format — used for non-numeric input columns
