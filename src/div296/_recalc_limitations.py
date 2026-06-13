"""Cells the pure-Python `formulas` recalc engine cannot evaluate.

Single source of truth shared by div296.build.validate_recalc (skip-list)
and tests/test_integration.py (assertion exclusions). All entries verified
correct in real Excel/LibreOffice by hand and pinned indirectly via
tests/test_calcs.py. Entries are SUFFIXES of the recalc solution keys
(e.g. "ANALYSER'!B71"), derived from layout constants so column/row shifts
can't strand them (audit 2026-06-10, finding 1.3).
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter

from div296.tabs import analyser as A_TAB
from div296.tabs import comparison as C_TAB

# Comparison per-asset detail panel — the IF(MATCH(LARGE(delta,k)...)) lookup
# chain that the pure-Python `formulas` engine returns #VALUE! for, while Excel
# and LibreOffice evaluate it correctly (README "Known validator false
# positives"; the rendered PDFs are proof, and tests/test_calcs.py asserts the
# §12 numbers via the Python formula mirror). The whole visible panel A..K ×
# the DISPLAY_ROWS data rows is excluded — derived from layout constants so a
# row/col shift can't strand it.
_PANEL_COLS = (*C_TAB.PANEL_A_COLS, *C_TAB.PANEL_B_COLS, C_TAB.DELTA_COL)
_PANEL_CELLS: tuple[str, ...] = tuple(
    f"COMPARISON'!{col}{row}"
    for row in range(C_TAB.DATA_FIRST_ROW, C_TAB.DATA_LAST_ROW + 1)
    for col in _PANEL_COLS
)

KNOWN_FORMULAS_LIMITATIONS: tuple[str, ...] = (
    f"ANALYSER'!{get_column_letter(A_TAB.HELPER_DISC_GAINS_COL)}{A_TAB.RECON_BAND_ROW}",    # disc-gains helper (col O)
    f"ANALYSER'!{get_column_letter(A_TAB.HELPER_NOND_GAINS_COL)}{A_TAB.RECON_BAND_ROW}",    # nond-gains helper (col P)
    f"ANALYSER'!{get_column_letter(A_TAB.HELPER_GROSS_LOSSES_COL)}{A_TAB.RECON_BAND_ROW}",  # gross-losses helper (col Q)
    f"ANALYSER'!B{A_TAB.RECON_ORD_CGT_ROW}",      # Fund Ord CGT (depends on helpers)
    f"ANALYSER'!B{A_TAB.RECON_LOSSES_ROW}",       # Carry-forward losses (depends on helpers)
    # Comparison: Ord CGT subtotal pulls Analyser!B71 → propagates #VALUE!.
    # Burden = Ord CGT + Div 296 tax, so burden propagates too.
    f"COMPARISON'!B{C_TAB.SUBTOTAL_ORD_CGT_ROW}",
    f"COMPARISON'!C{C_TAB.SUBTOTAL_ORD_CGT_ROW}",
    f"COMPARISON'!D{C_TAB.SUBTOTAL_ORD_CGT_ROW}",
    f"COMPARISON'!B{C_TAB.SUBTOTAL_BURDEN_ROW}",
    f"COMPARISON'!C{C_TAB.SUBTOTAL_BURDEN_ROW}",
    f"COMPARISON'!D{C_TAB.SUBTOTAL_BURDEN_ROW}",
    # Per-asset detail panel (LARGE/MATCH/INDEX false positives — see above).
    *_PANEL_CELLS,
)


def is_known_limitation(solution_key: str) -> bool:
    """True when a `formulas` solution key matches a known-limitation cell."""
    k = solution_key.upper()
    return any(k.endswith(suffix.upper()) for suffix in KNOWN_FORMULAS_LIMITATIONS)
