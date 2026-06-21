"""Calculator tab — the single data-entry + output sheet.

Layout (members as columns B–E; labels in col A):

  Row 1   Title
  Row 2   Sample-data badge
  Row 3   Income year (input B3) + threshold-loaded status (D3)
  Row 5   Band: 1. Fund pooled income
  Row 6   Dividends/distributions (grossed-up)      [input]
  Row 7   Interest                                  [input]
  Row 8   Rent                                      [input]
  Row 9   Other realised income                     [input]
  Row 10  Net realised capital gain (from helper)   [formula]
  Row 11  less: Deductible expenses                 [input]
  Row 12  POOLED TOTAL                              [formula]
  Row 14  Band: 2. CGT netting helper
  Row 15  Gross gains held >12m                     [input]
  Row 16  Gross gains held <12m                     [input]
  Row 17  Capital losses (current + brought-fwd)    [input]
  Row 18  (spare)
  Row 19  Unused capital loss carried forward       [formula output]
  Row 21  Band: 3. Members
  Row 22  Member name        (B–E)                  [input]
  Row 23  Opening TSB        (B–E)                   [input]
  Row 24  Closing TSB        (B–E)                   [input]
  Row 25  Share % of pooled earnings (B–E)          [input]
  Row 26  Prior-year Div 296 loss   (B–E)           [input]
  Row 27  Earnings override (optional) (B–E)        [input]
  Row 28  Earnings used      (B–E)                  [formula]
  Row 29  TSB used (ref)     (B–E)                  [formula]
  Row 30  Net Div 296 earnings (B–E)                [formula]
  Row 31  band1 proportion   (B–E)                  [formula]
  Row 32  band2 proportion   (B–E)                  [formula]
  Row 33  Tier-1 tax (15%)   (B–E)                  [formula]
  Row 34  Tier-2 tax (extra) (B–E)                  [formula]
  Row 35  TOTAL Division 296 tax (B–E)              [formula]
  Row 36  New carried-forward loss (B–E)            [formula]
  Row 37  Status             (B–E)                  [formula]
  Row 38  (hidden) pooled-share contribution (B–E)  [helper]
  Row 39  Pooled-share sum / count                  [helper]
  Row 40  FUND TOTAL Division 296 tax               [formula]
  Row 42-44  Liability block (3 plain lines)

  Hidden cols O–R: year table  (O=year, P=t1, Q=t2, R=greater_of)
  Resolved selectors (named): t1_sel, t2_sel, greater_of_sel, and the
  rate/discount constants live in col B of an assumptions strip (rows 46-48).
"""

from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from div296.styles import (
    BODY_FONT,
    FMT_CURRENCY,
    FMT_PERCENT,
    FMT_TEXT,
    INPUT_FILL,
    INPUT_FONT,
    SECTION_BAND_FILL,
    SECTION_BAND_FONT,
    THIN_BOX,
    TITLE_FONT,
)
from div296_calc import _formulas as F
from div296_calc import named_ranges as nr
from div296_calc.assumptions import (
    DISCOUNT_RATE,
    MEMBER_COUNT,
    RATE_TIER1,
    RATE_TIER2,
    YEAR_TABLE,
)

SHEET = "Calculator"

# --- row constants (single source of truth) ---
YEAR_ROW = 3
DIVIDENDS_ROW = 6
INTEREST_ROW = 7
RENT_ROW = 8
OTHER_ROW = 9
NET_CG_ROW = 10
EXPENSES_ROW = 11
POOLED_TOTAL_ROW = 12
CGT_OVER_ROW = 15
CGT_UNDER_ROW = 16
CGT_LOSS_ROW = 17
UNUSED_LOSS_ROW = 19
NAME_ROW = 22
OPENING_TSB_ROW = 23
CLOSING_TSB_ROW = 24
SHARE_ROW = 25
PRIOR_LOSS_ROW = 26
OVERRIDE_ROW = 27
EARNINGS_ROW = 28
TSB_REF_ROW = 29
NET_EARNINGS_ROW = 30
BAND1_ROW = 31
BAND2_ROW = 32
TIER1_TAX_ROW = 33
TIER2_TAX_ROW = 34
TOTAL_TAX_ROW = 35
NEW_LOSS_ROW = 36
STATUS_ROW = 37
POOLED_CONTRIB_ROW = 38
SHARE_GUARD_ROW = 39
FUND_TOTAL_ROW = 40
LIABILITY_ROW = 42

ASSUMP_FIRST_ROW = 46          # rate_tier1, rate_tier2, discount_rate strip
YEAR_TABLE_FIRST_ROW = 2       # hidden cols O–R start at row 2 (row 1 = headers)
YEAR_TABLE_LAST_ROW = 11       # capacity for ~10 future year rows

# Member columns B..E (1 + MEMBER_COUNT-1).
MEMBER_COLS = [get_column_letter(2 + i) for i in range(MEMBER_COUNT)]   # B,C,D,E

# Seeded coherent sample fund (a single pool; shares sum to 100%).
SAMPLE_POOLED = {
    DIVIDENDS_ROW: 120_000, INTEREST_ROW: 30_000, RENT_ROW: 80_000,
    OTHER_ROW: 0, EXPENSES_ROW: 40_000,
}
SAMPLE_CGT = {CGT_OVER_ROW: 300_000, CGT_UNDER_ROW: 0, CGT_LOSS_ROW: 60_000}
# Alice (B): pooled 60%, $4M.  Bob (C): pooled 40%, $12.9M (two-tier anchor).
SAMPLE_MEMBERS = {
    "B": {NAME_ROW: "Alice", OPENING_TSB_ROW: 3_800_000, CLOSING_TSB_ROW: 4_000_000,
          SHARE_ROW: 0.6, PRIOR_LOSS_ROW: 0},
    "C": {NAME_ROW: "Bob", OPENING_TSB_ROW: 12_500_000, CLOSING_TSB_ROW: 12_900_000,
          SHARE_ROW: 0.4, PRIOR_LOSS_ROW: 0},
}


def _band(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = SECTION_BAND_FONT
    ws.merge_cells(f"A{row}:E{row}")
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SECTION_BAND_FILL


def _label(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = BODY_FONT


def _input(ws, coord, value=None, number_format=FMT_CURRENCY):
    cell = ws[coord]
    if value is not None:
        cell.value = value
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = THIN_BOX
    cell.protection = Protection(locked=False)
    cell.number_format = number_format


def _formula(ws, coord, value, number_format=FMT_CURRENCY):
    cell = ws[coord]
    cell.value = value
    cell.number_format = number_format
    cell.font = Font(name="Arial", size=10, italic=True, color="1D3B34")


def _define_name(wb, name, coord):
    wb.defined_names[name] = DefinedName(
        name=name, attr_text=f"'{SHEET}'!${coord[0]}${coord[1:]}")


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)

    # --- Title ---
    ws["A1"] = "Ongoing Division 296 Calculator"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    # --- Hidden year table (cols O–R) ---
    ws["O1"] = "Year"
    ws["P1"] = "T1"
    ws["Q1"] = "T2"
    ws["R1"] = "GreaterOf"
    for i, (year, yt) in enumerate(sorted(YEAR_TABLE.items())):
        r = YEAR_TABLE_FIRST_ROW + i
        ws[f"O{r}"] = year
        ws[f"P{r}"] = yt.threshold_1
        ws[f"Q{r}"] = yt.threshold_2
        ws[f"R{r}"] = 1 if yt.use_greater_of else 0
    for col in ("O", "P", "Q", "R"):
        ws.column_dimensions[col].hidden = True

    year_rng_o = f"$O${YEAR_TABLE_FIRST_ROW}:$O${YEAR_TABLE_LAST_ROW}"
    t1_rng = f"$P${YEAR_TABLE_FIRST_ROW}:$P${YEAR_TABLE_LAST_ROW}"
    t2_rng = f"$Q${YEAR_TABLE_FIRST_ROW}:$Q${YEAR_TABLE_LAST_ROW}"
    gof_rng = f"$R${YEAR_TABLE_FIRST_ROW}:$R${YEAR_TABLE_LAST_ROW}"

    # --- Row 3: income year selector + resolved thresholds ---
    _label(ws, YEAR_ROW, "Income year")
    _input(ws, f"B{YEAR_ROW}", value="2026-27", number_format=FMT_TEXT)
    # constrain to table years
    # NOTE: a range-reference list DV takes the bare range (no leading '='),
    # mirroring openpyxl's idiom; the '="Yes,No"' form is only for literal lists.
    dv = DataValidation(type="list", formula1=year_rng_o, allow_blank=False,
                        showErrorMessage=True, errorTitle="Unknown income year",
                        error="Pick a year present in the threshold table (cols O–R).")
    ws.add_data_validation(dv)
    dv.add(f"B{YEAR_ROW}")
    ws[f"D{YEAR_ROW}"] = F.year_known_guard_formula(year_rng_o, f"$B${YEAR_ROW}")
    ws[f"D{YEAR_ROW}"].font = Font(name="Arial", size=10, bold=True, color="666666")

    # resolved selectors (named) — placed in hidden helper cells, col S
    ws["S2"] = F.threshold_lookup_formula(t1_rng, year_rng_o, f"$B${YEAR_ROW}")
    ws["S3"] = F.threshold_lookup_formula(t2_rng, year_rng_o, f"$B${YEAR_ROW}")
    ws["S4"] = F.threshold_lookup_formula(gof_rng, year_rng_o, f"$B${YEAR_ROW}")
    ws.column_dimensions["S"].hidden = True
    _define_name(wb, nr.T1_SEL, "S2")
    _define_name(wb, nr.T2_SEL, "S3")
    _define_name(wb, nr.GREATER_OF_SEL, "S4")

    # --- Assumptions strip (rows 46-48): rate constants as named cells ---
    for offset, (name, label, value, fmt) in enumerate([
        (nr.RATE_TIER1, "Div 296 rate — tier 1 ($3m–$10m)", RATE_TIER1, FMT_PERCENT),
        (nr.RATE_TIER2, "Div 296 rate — tier 2 (above $10m)", RATE_TIER2, FMT_PERCENT),
        (nr.DISCOUNT_RATE, "CGT discount (1/3)", DISCOUNT_RATE, "0.000%"),
    ]):
        r = ASSUMP_FIRST_ROW + offset
        _label(ws, r, label)
        _input(ws, f"B{r}", value=value, number_format=fmt)
        _define_name(wb, name, f"B{r}")

    # --- Band 1: pooled income ---
    _band(ws, 5,
          "1. Fund pooled income (realised, before fund 15% tax; exclude contributions)")
    for row, text in [
        (DIVIDENDS_ROW, "Dividends / distributions (grossed-up, incl. franking)"),
        (INTEREST_ROW, "Interest"), (RENT_ROW, "Rent"),
        (OTHER_ROW, "Other realised income"),
    ]:
        _label(ws, row, text)
        _input(ws, f"B{row}", value=SAMPLE_POOLED.get(row), number_format=FMT_CURRENCY)
    _label(ws, NET_CG_ROW, "Net realised capital gain (from CGT helper below)")
    _formula(ws, f"B{NET_CG_ROW}",
             F.cgt_net_formula(f"B{CGT_OVER_ROW}", f"B{CGT_UNDER_ROW}", f"B{CGT_LOSS_ROW}"))
    _label(ws, EXPENSES_ROW, "less: Deductible expenses")
    _input(ws, f"B{EXPENSES_ROW}", value=SAMPLE_POOLED.get(EXPENSES_ROW),
           number_format=FMT_CURRENCY)
    _label(ws, POOLED_TOTAL_ROW, "POOLED TOTAL")
    _formula(ws, f"B{POOLED_TOTAL_ROW}", F.pooled_total_formula(
        f"B{DIVIDENDS_ROW}", f"B{INTEREST_ROW}", f"B{RENT_ROW}",
        f"B{OTHER_ROW}", f"B{NET_CG_ROW}", f"B{EXPENSES_ROW}"))
    ws[f"B{POOLED_TOTAL_ROW}"].font = Font(name="Arial", size=11, bold=True, color="0F6E56")

    # --- Band 2: CGT helper ---
    _band(ws, 14, "2. CGT netting helper (raw figures off a CLASS/BGL report)")
    for row, text, val in [
        (CGT_OVER_ROW, "Gross capital gains — held > 12 months (discountable)",
         SAMPLE_CGT[CGT_OVER_ROW]),
        (CGT_UNDER_ROW, "Gross capital gains — held < 12 months",
         SAMPLE_CGT[CGT_UNDER_ROW]),
        (CGT_LOSS_ROW, "Capital losses (current year + brought-forward)",
         SAMPLE_CGT[CGT_LOSS_ROW]),
    ]:
        _label(ws, row, text)
        _input(ws, f"B{row}", value=val, number_format=FMT_CURRENCY)
    _label(ws, UNUSED_LOSS_ROW,
           "→ Unused capital loss carried forward (to NEXT year's losses)")
    _formula(ws, f"B{UNUSED_LOSS_ROW}", F.cgt_unused_loss_formula(
        f"B{CGT_OVER_ROW}", f"B{CGT_UNDER_ROW}", f"B{CGT_LOSS_ROW}"))

    # --- Band 3: members (cols B–E) ---
    _band(ws, 21, "3. Members (enter up to 4; segregated members use the override row)")
    row_labels = {
        NAME_ROW: "Member name", OPENING_TSB_ROW: "Opening TSB (1 Jul, ALL super)",
        CLOSING_TSB_ROW: "Closing TSB (30 Jun, ALL super)",
        SHARE_ROW: "Share % of pooled earnings", PRIOR_LOSS_ROW: "Prior-year Div 296 loss",
        OVERRIDE_ROW: "Earnings override (segregated; optional)",
        EARNINGS_ROW: "Earnings used", TSB_REF_ROW: "TSB used (ref)",
        NET_EARNINGS_ROW: "Net Div 296 earnings", BAND1_ROW: "Proportion $3m–$10m",
        BAND2_ROW: "Proportion above $10m", TIER1_TAX_ROW: "Tier-1 tax (15%)",
        TIER2_TAX_ROW: "Tier-2 tax (extra 10%)", TOTAL_TAX_ROW: "TOTAL Division 296 tax",
        NEW_LOSS_ROW: "New carried-forward Div 296 loss", STATUS_ROW: "Status",
    }
    for row, text in row_labels.items():
        _label(ws, row, text)

    for col in MEMBER_COLS:
        seed = SAMPLE_MEMBERS.get(col, {})
        _input(ws, f"{col}{NAME_ROW}", value=seed.get(NAME_ROW), number_format=FMT_TEXT)
        _input(ws, f"{col}{OPENING_TSB_ROW}", value=seed.get(OPENING_TSB_ROW),
               number_format=FMT_CURRENCY)
        _input(ws, f"{col}{CLOSING_TSB_ROW}", value=seed.get(CLOSING_TSB_ROW),
               number_format=FMT_CURRENCY)
        _input(ws, f"{col}{SHARE_ROW}", value=seed.get(SHARE_ROW),
               number_format=FMT_PERCENT)
        _input(ws, f"{col}{PRIOR_LOSS_ROW}", value=seed.get(PRIOR_LOSS_ROW),
               number_format=FMT_CURRENCY)
        _input(ws, f"{col}{OVERRIDE_ROW}", value=None, number_format=FMT_CURRENCY)

        _formula(ws, f"{col}{EARNINGS_ROW}", F.earnings_formula(
            f"{col}{NAME_ROW}", f"{col}{OVERRIDE_ROW}", f"{col}{SHARE_ROW}",
            f"$B${POOLED_TOTAL_ROW}"))
        _formula(ws, f"{col}{TSB_REF_ROW}", F.tsb_ref_formula(
            f"{col}{OPENING_TSB_ROW}", f"{col}{CLOSING_TSB_ROW}", nr.GREATER_OF_SEL))
        _formula(ws, f"{col}{NET_EARNINGS_ROW}", F.net_earnings_formula(
            f"{col}{EARNINGS_ROW}", f"{col}{PRIOR_LOSS_ROW}"))
        _formula(ws, f"{col}{BAND1_ROW}", F.band1_formula(
            f"{col}{TSB_REF_ROW}", nr.T1_SEL, nr.T2_SEL), number_format=FMT_PERCENT)
        _formula(ws, f"{col}{BAND2_ROW}", F.band2_formula(
            f"{col}{TSB_REF_ROW}", nr.T2_SEL), number_format=FMT_PERCENT)
        _formula(ws, f"{col}{TIER1_TAX_ROW}", F.tier1_tax_formula(
            f"{col}{NET_EARNINGS_ROW}", f"{col}{BAND1_ROW}", f"{col}{TSB_REF_ROW}",
            nr.T1_SEL))
        _formula(ws, f"{col}{TIER2_TAX_ROW}", F.tier2_tax_formula(
            f"{col}{NET_EARNINGS_ROW}", f"{col}{BAND2_ROW}", f"{col}{TSB_REF_ROW}",
            nr.T1_SEL))
        _formula(ws, f"{col}{TOTAL_TAX_ROW}",
                 f"=SUM({col}{TIER1_TAX_ROW},{col}{TIER2_TAX_ROW})")
        ws[f"{col}{TOTAL_TAX_ROW}"].font = Font(name="Arial", size=11, bold=True,
                                                color="0F6E56")
        _formula(ws, f"{col}{NEW_LOSS_ROW}", F.new_loss_formula(f"{col}{NET_EARNINGS_ROW}"))
        _formula(ws, f"{col}{STATUS_ROW}", F.status_formula(
            f"{col}{TSB_REF_ROW}", nr.T1_SEL), number_format=FMT_TEXT)
        # hidden pooled-share contribution helper
        _formula(ws, f"{col}{POOLED_CONTRIB_ROW}", F.pooled_share_contrib_formula(
            f"{col}{NAME_ROW}", f"{col}{OVERRIDE_ROW}", f"{col}{SHARE_ROW}"),
            number_format=FMT_PERCENT)
    ws.row_dimensions[POOLED_CONTRIB_ROW].hidden = True

    # --- Share guard ---
    last_col = MEMBER_COLS[-1]
    _label(ws, SHARE_GUARD_ROW, "Pooled-share check")
    ws[f"B{SHARE_GUARD_ROW}"] = F.pooled_count_formula(
        f"B{NAME_ROW}:{last_col}{NAME_ROW}", f"B{OVERRIDE_ROW}:{last_col}{OVERRIDE_ROW}")
    ws[f"C{SHARE_GUARD_ROW}"] = f"=SUM(B{POOLED_CONTRIB_ROW}:{last_col}{POOLED_CONTRIB_ROW})"
    ws[f"C{SHARE_GUARD_ROW}"].number_format = FMT_PERCENT
    ws[f"D{SHARE_GUARD_ROW}"] = F.share_guard_status_formula(
        f"B{SHARE_GUARD_ROW}", f"C{SHARE_GUARD_ROW}")
    ws.conditional_formatting.add(
        f"D{SHARE_GUARD_ROW}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("⚠",D{SHARE_GUARD_ROW}))'],
                    fill=PatternFill("solid", fgColor="FBE9E9")))
    ws.conditional_formatting.add(
        f"D{SHARE_GUARD_ROW}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("✓",D{SHARE_GUARD_ROW}))'],
                    fill=PatternFill("solid", fgColor="E1F5EE")))

    # --- Fund total ---
    _label(ws, FUND_TOTAL_ROW, "FUND TOTAL Division 296 tax")
    ws[f"B{FUND_TOTAL_ROW}"] = f"=SUM(B{TOTAL_TAX_ROW}:{last_col}{TOTAL_TAX_ROW})"
    ws[f"B{FUND_TOTAL_ROW}"].number_format = FMT_CURRENCY
    ws[f"B{FUND_TOTAL_ROW}"].font = Font(name="Arial", size=12, bold=True, color="0F6E56")

    # --- Liability block (3 plain lines) ---
    for i, line in enumerate([
        "Division 296 is assessed to the INDIVIDUAL, on top of the fund's own 15% tax.",
        "Payable personally OR by election to release from a super fund "
        "(release-authority deadline applies).",
        "Two carry-forwards: the member Div 296 loss (row 36) rolls into next year's "
        "prior-loss; the unused capital loss (row 19) rolls into next year's CGT-helper "
        "losses.",
    ]):
        c = ws.cell(row=LIABILITY_ROW + i, column=1, value=line)
        c.font = Font(name="Arial", size=9, italic=True, color="666666")
        ws.merge_cells(f"A{LIABILITY_ROW + i}:E{LIABILITY_ROW + i}")

    # --- Sample-data badge ---
    badge = ws.cell(row=2, column=1, value=(
        f'=IF(AND(B{DIVIDENDS_ROW}={SAMPLE_POOLED[DIVIDENDS_ROW]},'
        f'B{CLOSING_TSB_ROW}={SAMPLE_MEMBERS["B"][CLOSING_TSB_ROW]}),'
        f'"⚠ Sample data preloaded — overwrite with your fund\'s figures before '
        f'sharing.","")'))
    badge.font = Font(name="Arial", size=10, bold=True, italic=True, color="8A6D00")
    badge.fill = PatternFill("solid", fgColor="FFF4CE")
    ws.merge_cells("A2:E2")

    # --- Column widths, protection, print ---
    ws.column_dimensions["A"].width = 46
    for col in MEMBER_COLS:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "B4"

    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    ws.oddHeader.center.text = "ILLUSTRATIVE — NOT ADVICE"
    ws.oddHeader.center.size = 28
    ws.oddHeader.center.color = "CCCCCC"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:E{LIABILITY_ROW + 2}"
    return ws
