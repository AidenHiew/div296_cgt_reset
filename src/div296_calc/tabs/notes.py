"""Notes / disclaimer tab."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from div296.styles import SECTION_BAND_FILL, SECTION_BAND_FONT, TITLE_FONT
from div296_calc.assumptions import YEAR_TABLE

SHEET = "Notes"
THRESHOLDS_VERIFIED = "2026-06-20"   # update when YEAR_TABLE changes

LINES = [
    ("band", "Division 296 — ongoing calculator notes"),
    ("p", "Law basis: Treasury Laws Amendment (Building a Stronger and Fairer "
          "Super System) Act 2026 — Royal Assent 13 March 2026. First income "
          "year 2026-27 (first test date 30 June 2027)."),
    ("band", "Thresholds by year (user-maintained)"),
    ("p", f"thresholds verified {THRESHOLDS_VERIFIED}. At v0.1 only 2026-27 is "
          "confirmed ($3M / $10M, un-indexed). Later years are CPI-indexed "
          "($150k / $500k steps) — add each year's ATO-published figures to the "
          "hidden year table on the Calculator tab (cols O–R) as released."),
    ("p", "Years currently loaded: " + ", ".join(sorted(YEAR_TABLE))),
    ("band", "Earnings basis"),
    ("p", "Enacted method (approximated here): earnings ≈ taxable income − "
          "assessable contributions + net ECPI − NALI, measured BEFORE the "
          "fund's own 15% tax. v0.1 takes realised income components (less "
          "deductible expenses) directly. Contributions are NOT added back and "
          "there is NO $3M loss-floor (both were 2023-draft features)."),
    ("p", "Franking: dividends should be the grossed-up amount (cash + franking "
          "credits) per the fund return; exact treatment under final "
          "regulations to be confirmed."),
    ("p", "CGT discount caveat: the helper applies the 1/3 discount to ALL "
          "realised gains held >12 months regardless of pension/accumulation "
          "phase — an approximation pending final regs."),
    ("band", "Reset cost base"),
    ("p", "For 2026-27 onward, realised capital gains should be measured from "
          "the 30-Jun-2026 reset cost base IF the fund made that election — see "
          "the separate year-one CGT-reset tool."),
    ("band", "Two carry-forwards — do not conflate"),
    ("p", "Member Division 296 loss (Calculator row 36) → next year's per-member "
          "prior-loss input. Unused capital loss (Calculator row 19) → next "
          "year's CGT-helper capital-losses input."),
    ("band", "Scope & disclaimer"),
    ("p", "Accumulation and account-based pensions only; one calc per member "
          "across ALL their super (not per fund). Defined-benefit interests, "
          "NALI, and multi-fund aggregation are out of scope. Outputs are "
          "ESTIMATES subject to final regulations — not personal advice."),
]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET)
    r = 1
    for kind, text in LINES:
        if kind == "band":
            c = ws.cell(row=r, column=1, value=text)
            c.font = SECTION_BAND_FONT if r > 1 else TITLE_FONT
            if r > 1:
                c.fill = SECTION_BAND_FILL
            ws.merge_cells(f"A{r}:H{r}")
        else:
            c = ws.cell(row=r, column=1, value=text)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"A{r}:H{r}")
            ws.row_dimensions[r].height = 46
        r += 1

    ws.column_dimensions["A"].width = 20
    ws.oddHeader.center.text = "ILLUSTRATIVE — NOT ADVICE"
    ws.oddHeader.center.size = 28
    ws.oddHeader.center.color = "CCCCCC"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws
