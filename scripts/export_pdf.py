"""LibreOffice headless PDF export of the Comparison tab.

Renders the Comparison tab of a built workbook to a single-page A4
landscape PDF suitable for sharing with clients.

    python scripts/export_pdf.py dist/Division_296_Model_v0.1.0.xlsx
    python scripts/export_pdf.py dist/Division_296_Model_v0.1.0.xlsx --tab Comparison --out my_output.pdf
    python scripts/export_pdf.py dist/Division_296_Model_v0.1.0.xlsx --all-tabs

Requires LibreOffice installed and `soffice` on PATH. On Windows the
default install puts it at:
    C:\\Program Files\\LibreOffice\\program\\soffice.exe

Approach: copy the workbook to a temp dir, mark only the requested tab
as visible (others hidden), then run `soffice --convert-to pdf`. The
Comparison tab's print area + fit-to-page settings ensure it lands on
one A4 landscape page.
"""

from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


def _soffice_executable() -> str | None:
    for name in ("soffice", "soffice.exe", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    # Common Windows install path.
    win_default = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
    if win_default.exists():
        return str(win_default)
    return None


def _isolate_tab(src: Path, work_dir: Path, tab_name: str | None) -> Path:
    """Copy `src` to `work_dir` and (optionally) hide every tab except `tab_name`."""
    staged = work_dir / src.name
    shutil.copy2(src, staged)
    if tab_name is None:
        return staged

    wb = load_workbook(staged)
    if tab_name not in wb.sheetnames:
        raise SystemExit(
            f"Tab {tab_name!r} not found in {src.name}. Available: {wb.sheetnames}"
        )
    wb.active = wb.sheetnames.index(tab_name)
    for name in wb.sheetnames:
        wb[name].sheet_state = "visible" if name == tab_name else "hidden"
    wb.save(staged)
    return staged


def _soffice(soffice: str, work: Path, convert_to: str, target: Path) -> subprocess.CompletedProcess:
    """Invoke `soffice --headless --calc --convert-to <fmt>` against `target`, output in `work`."""
    cmd = [
        soffice, "--headless", "--calc",
        "--convert-to", convert_to,
        "--outdir", str(work),
        str(target),
    ]
    return subprocess.run(cmd, capture_output=True, text=True, check=False)


def export_pdf(src: Path, out: Path | None, tab: str | None) -> int:
    soffice = _soffice_executable()
    if soffice is None:
        print("ERROR: LibreOffice not found. Install it and put soffice on PATH.",
              file=sys.stderr)
        return 1

    if not src.exists():
        print(f"ERROR: input file does not exist: {src}", file=sys.stderr)
        return 1

    out_dir = (out.parent if out else src.parent).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="div296_pdf_") as td:
        work = Path(td)
        staged = _isolate_tab(src, work, tab)

        # Step 1: recalculate via a convert-to-xlsx round-trip so chart data
        # series (which reference formula cells) get cached values written.
        # Without this, the headless PDF renderer sees empty refs and the chart
        # renders with no bars and a default 0–12 axis.
        recalc_dir = work / "recalc"
        recalc_dir.mkdir()
        recalc_result = _soffice(soffice, recalc_dir, "xlsx", staged)
        if recalc_result.returncode != 0:
            print(recalc_result.stdout, file=sys.stdout)
            print(recalc_result.stderr, file=sys.stderr)
            return recalc_result.returncode
        recalc_xlsx = recalc_dir / staged.name
        if not recalc_xlsx.exists():
            print(f"ERROR: recalc step did not produce {recalc_xlsx}", file=sys.stderr)
            return 1

        # Step 2: convert the recalc'd workbook to PDF.
        pdf_result = _soffice(soffice, work, "pdf", recalc_xlsx)
        if pdf_result.returncode != 0:
            print(pdf_result.stdout, file=sys.stdout)
            print(pdf_result.stderr, file=sys.stderr)
            return pdf_result.returncode

        produced = work / (recalc_xlsx.stem + ".pdf")
        if not produced.exists():
            print(f"ERROR: LibreOffice did not produce {produced}", file=sys.stderr)
            print(pdf_result.stdout, file=sys.stdout)
            return 1

        final = out or (out_dir / f"{src.stem}{'_' + tab if tab else ''}.pdf")
        shutil.copy2(produced, final)
        print(f"Wrote {final}")
        return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Render a tab of the Div 296 workbook to PDF via LibreOffice headless.")
    parser.add_argument("xlsx", type=Path, help="Path to the built .xlsx")
    parser.add_argument(
        "--tab", default="Comparison",
        help="Tab to export (default: Comparison). Use --all-tabs to export every tab.",
    )
    parser.add_argument(
        "--all-tabs", action="store_true",
        help="Export the whole workbook (every visible tab) into one PDF.",
    )
    parser.add_argument(
        "--out", type=Path, default=None,
        help="Output PDF path (default: alongside the .xlsx).",
    )
    args = parser.parse_args(argv)

    tab = None if args.all_tabs else args.tab
    return export_pdf(args.xlsx.resolve(), args.out, tab)


if __name__ == "__main__":
    raise SystemExit(main())
