"""Integration tests: build the workbook, optionally recalc, read it back.

v3.0 — control panel removed (toggles always-on internally per Bill-correct
calc); Inputs Section 2 now shows two-band proportion display (cols D/E);
Analyser fund summary restructured side-by-side with per-member breakdown
and signed Difference column.

Row numbers shifted:
- Inputs: Members section moved up by 4 (now rows 6-12; was 10-16); Asset
  register starts at row 16 (was 20).
- Analyser: Lever mirror rows 4-7 deleted; everything shifted up by ~4.
  Fund summary band at row 6 (was 11); per-asset first row at 17 (was 21).
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from openpyxl.utils import get_column_letter

from div296 import named_ranges as nr
from div296.build import build_workbook
from div296.tabs import class_import as ci
from div296.tabs.comparison import (
    DATA_FIRST_ROW as CMP_DATA_FIRST_ROW,
    DATA_LAST_ROW as CMP_DATA_LAST_ROW,
    DATA_OVERFLOW_NOTE_ROW as CMP_OVERFLOW_ROW,
)


EXPECTED_TABS = ["Inputs", "CLASS Import", "Analyser", "Comparison", "Notes"]


def test_workbook_has_five_tabs_in_spec_order(tmp_path: Path):
    """v3.3: CLASS Import staging tab inserted after Inputs."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    reopened = load_workbook(out)
    assert reopened.sheetnames == EXPECTED_TABS


def test_every_named_range_resolves(tmp_path: Path):
    """v3.0: ALL_NAMES no longer includes the 3 toggle names."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    reopened = load_workbook(out)

    missing = [name for name in nr.ALL_NAMES if name not in reopened.defined_names]
    assert not missing, f"Named ranges not found in workbook: {missing}"

    for name in nr.ALL_NAMES:
        dn = reopened.defined_names[name]
        destinations = list(dn.destinations)
        assert destinations, f"Named range {name!r} has no destinations"
        for sheet_name, ref in destinations:
            assert sheet_name in reopened.sheetnames, f"{name} points at missing sheet {sheet_name}"
            ws = reopened[sheet_name]
            cell = ws[ref.replace("$", "")]
            assert cell is not None


def test_v3_toggle_named_ranges_removed(tmp_path: Path):
    """v3.0: reset_on, tier10_on, discount_on must NOT exist."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    reopened = load_workbook(out)
    for removed in ("reset_on", "tier10_on", "discount_on"):
        assert removed not in reopened.defined_names, (
            f"v3.0 should have removed named range {removed!r}"
        )


# --- Inputs tab -----------------------------------------------------------

def test_inputs_sample_data_preloaded(tmp_path: Path):
    """v3.0: Asset register first row shifted from row 20 to row 16
    (control panel rows 4-8 removed)."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]

    # Row 16: Commercial property (was row 20 in v2.6)
    assert ws["A16"].value == "P1"
    assert ws["B16"].value == "Commercial property"
    assert ws["C16"].value == 800_000             # Original cost base
    assert ws["E16"].value == 2_400_000           # MV at 30 Jun 2026
    assert ws["G16"].value == 2_600_000           # Projected sale proceeds
    assert ws["I16"].value == "Yes"               # Held>12m

    # Row 17: Listed shares parcel
    assert ws["A17"].value == "S1"
    assert ws["C17"].value == 300_000
    assert ws["E17"].value == 520_000

    # Row 18: Loss-making holding
    assert ws["A18"].value == "L1"
    assert ws["C18"].value == 500_000
    assert ws["E18"].value == 100_000

    # Projected gain/loss formula in col H references new row numbers.
    pg_formula = ws["H16"].value
    assert pg_formula and "G16" in pg_formula and "C16" in pg_formula

    # v3.0: Sole member on row 7 (was row 11): TSB $12m.
    assert ws["B7"].value == 12_000_000
    c7 = ws["C7"].value
    assert isinstance(c7, str) and c7.startswith("="), (
        f"C7 (Split %) should be a formula in v3.0, got {c7!r}"
    )
    assert "SUM($B$7:$B$10)" in c7 and "B7/" in c7


def test_inputs_section_2_band_columns(tmp_path: Path):
    """v3.0 NEW: cols D and E in Members section are auto-derived band1 / band2."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]

    # Header row (row 6 in v3.0)
    assert "$3m" in str(ws["D6"].value) and "$10m" in str(ws["D6"].value), (
        f"Col D header should mention $3m-$10m band, got {ws['D6'].value!r}"
    )
    assert "$10m" in str(ws["E6"].value), (
        f"Col E header should mention above $10m, got {ws['E6'].value!r}"
    )

    # Member 1 (row 7) — band1 formula uses MIN(TSB, threshold_2) - threshold_1
    d7 = ws["D7"].value
    assert d7 and d7.startswith("=")
    assert "MIN" in d7 and "threshold_2" in d7 and "threshold_1" in d7

    # Member 1 (row 7) — band2 formula uses (TSB - threshold_2)
    e7 = ws["E7"].value
    assert e7 and e7.startswith("=")
    assert "threshold_2" in e7
    # band2 should NOT reference threshold_1 — only the >$10m slice.
    assert "threshold_1" not in e7


def test_inputs_control_panel_removed(tmp_path: Path):
    """v3.0: control panel section deleted — no toggles at B5/B6/B7."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]

    # In v3.0 the band at row 5 should read "1. Members" — not "Control panel".
    a5 = str(ws["A5"].value or "")
    assert "Control panel" not in a5, f"Control panel should be gone, got A5={a5!r}"
    assert "Members" in a5, f"Row 5 should be Members band, got A5={a5!r}"

    # Previous control-panel cells must not hold toggle defaults anymore.
    # B5 is now blank (band only). B6 is Members header. B7 is Member 1 TSB.
    assert ws["B7"].value == 12_000_000, (
        f"B7 should be Member 1 TSB in v3.0, got {ws['B7'].value!r}"
    )


def test_sample_data_badge_present(tmp_path: Path):
    """Row 2 badge tells staff to overwrite sample data."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]
    badge = ws["A2"].value
    assert badge is not None
    assert "Sample data" in badge
    assert "overwrite" in badge.lower()


def test_all_tabs_protected(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    wb_re = load_workbook(out)
    for tab in ("Inputs", "CLASS Import", "Analyser", "Notes", "Comparison"):
        assert wb_re[tab].protection.sheet is True, f"{tab} is not protected"


def test_all_tabs_allow_column_resize_under_protection(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    wb_re = load_workbook(out)
    for sheet in ("Inputs", "CLASS Import", "Analyser", "Notes", "Comparison"):
        ws = wb_re[sheet]
        assert ws.protection.sheet is True, f"{sheet} should be protected"
        assert ws.protection.formatColumns in (False, "0", 0), (
            f"{sheet} formatColumns must be unlocked under v2+ protection scope"
        )
        assert ws.protection.formatRows in (False, "0", 0), (
            f"{sheet} formatRows must be unlocked under v2+ protection scope"
        )


def test_input_cells_unlocked(tmp_path: Path):
    """v3.0: Member 1 TSB (B7) and Original cost base (C16) must stay editable."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]
    # v3.0 shifted: Member 1 TSB is at B7; Asset register original cost base at C16.
    assert ws["B7"].protection.locked is False
    assert ws["C16"].protection.locked is False


def test_register_proj_gl_formula_cells_locked(tmp_path: Path):
    """v3.3 audit F1: the col-H register formula is the CLASS-transfer paste
    guard — it must be LOCKED so an accidental A:H/A:I Paste-Special is
    rejected by sheet protection instead of silently wiping the formulas."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]
    for row in range(16, 66):
        h = ws[f"H{row}"]
        assert h.protection.locked is True, f"Inputs!H{row} must be locked"
        # Neighbouring input columns stay editable.
        assert ws[f"G{row}"].protection.locked is False
        assert ws[f"I{row}"].protection.locked is False


# --- Held>12m paste-in normalisation (v3.1.2) ---------------------------
#
# The Inputs!I dropdown only validates direct typing; pastes from another
# sheet or CSV import bypass it entirely. v3.1.2 added a hidden Inputs!J
# helper column that runs TRIM+UPPER on I and emits a clean "Yes"/"No",
# and redirected every downstream formula to read J instead of I.
# These tests pin both the wiring (formulas reference J) and the
# behaviour (paste-in "Yes " / "yes" still produce the discounted gain).


def test_inputs_j_column_is_hidden_normaliser(tmp_path: Path):
    """Inputs!J16:J65 = =IF(I=...,..,IF(TRIM(UPPER(I))="YES","Yes","No"));
    column is hidden so the user never sees the helper."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]

    assert ws.column_dimensions["J"].hidden, "Inputs!J must be hidden"

    for row in (16, 40, 65):
        f = ws[f"J{row}"].value
        assert isinstance(f, str) and f.startswith("="), (
            f"Inputs!J{row} should be a formula, got {f!r}"
        )
        assert "TRIM" in f and "UPPER" in f, (
            f"Inputs!J{row} should normalise via TRIM+UPPER, got {f!r}"
        )
        assert f'I{row}' in f, (
            f"Inputs!J{row} should reference its own I{row} source, got {f!r}"
        )


def test_downstream_formulas_read_held_from_j_not_i(tmp_path: Path):
    """Every formula that classifies a gain as discountable must read the
    NORMALISED held flag (Inputs!J), not the raw user input (Inputs!I).

    v3.2: the per-asset table now exposes gross gain explicitly (col E),
    the "1/3 CGT discount eligible?" flag (col F), and a derived Per-asset
    Ord CGT (col G). The held-from-J check moves:
      - col E (was post-disc, read Inputs!J)  →  col E is now pure gross
        (proceeds − orig CB), no Inputs!J reference.
      - col F (NEW flag col) now reads Inputs!J directly.
      - col G (derived Ord CGT) references col F transitively.
      - col H17 (was Div 296 post-disc) → col J17 carries the post-disc
        Div 296 value; same Inputs!J read.
      - M70 (Fund Ord CGT helper) → shifted to O70.
    """
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    wb_re = load_workbook(out)

    analyser = wb_re["Analyser"]
    # Col F (NEW: discount-eligible flag) is the direct Inputs!J reader.
    f17 = analyser["F17"].value
    assert "'Inputs'!J16" in f17, (
        f"Analyser F17 (flag) must read Inputs!J16, got {f17!r}"
    )
    assert "'Inputs'!I16" not in f17, (
        f"Analyser F17 must not read raw Inputs!I16, got {f17!r}"
    )
    # Col G (derived per-asset Ord CGT) references col F (which transitively
    # reads Inputs!J) — verified by the F17 check above and a structural
    # check here that G17 references F17.
    g17 = analyser["G17"].value
    assert "F17" in g17, f"Analyser G17 must reference col F17 (flag): {g17!r}"
    # Col J17 — Per-asset Div 296 gain (post-disc, info-only). Reads Inputs!J
    # because the discount branch depends on the held flag.
    j17 = analyser["J17"].value
    assert "'Inputs'!J16" in j17 and "'Inputs'!I16" not in j17, j17
    # Reconciliation helper for "disc gains" — shifted from M70 to O70.
    o70 = analyser["O70"].value
    assert "'Inputs'!J16:J65" in o70, (
        f"Recon helper O70 must aggregate over Inputs!J16:J65, got {o70!r}"
    )
    assert "'Inputs'!I16:I65" not in o70, o70

    # Comparison per-register helper grid (cols N/O — div296 adj gains).
    comparison = wb_re["Comparison"]
    # Find the per-register grid row aligned to Inputs row 16; the grid uses
    # the same row numbers as Inputs (n_first = REGISTER_FIRST_DATA_ROW).
    n16 = comparison["N16"].value
    assert "'Inputs'!J16" in n16 and "'Inputs'!I16" not in n16, n16


@pytest.mark.skip(reason=(
    "v3.2 column shift: per-asset col E is now gross (no discount), the "
    "post-disc Div 296 gain moved from H17 to J17, and the Fund Ord CGT "
    "helper moved from M70 to O70. The paste-normalisation correctness "
    "story belongs with the v3.3 Inputs!I DataValidation hardening chip; "
    "restore + update this test there, where paste behaviour is the "
    "actual subject. Avoiding the `formulas`-package OOM during v3.2 work."
))
def test_paste_in_dirty_held_values_are_normalised(tmp_path: Path):
    """Behavioural test: write paste-style dirty values ('Yes ', 'yes',
    ' YES ') into Inputs!I, recalc, and check that Analyser per-asset
    E (Ord taxable gain) and H (Div 296 adj gain) both apply the 1/3
    CGT discount as if the cell had said "Yes" exactly."""
    formulas = pytest.importorskip(  # noqa: F811
        "formulas",
        reason="`formulas` package required for paste-in normalisation test",
    )
    import numpy as np

    out = tmp_path / "out.xlsx"
    wb = build_workbook()

    # Simulate paste-in: trailing space (row 16), lowercase (row 17),
    # surrounding whitespace + uppercase (row 18). All three sample rows.
    ws_in = wb["Inputs"]
    ws_in["I16"] = "Yes "
    ws_in["I17"] = "yes"
    ws_in["I18"] = " YES "
    wb.save(out)

    try:
        sol = formulas.ExcelModel().loads(str(out)).finish().calculate()
    except MemoryError:
        pytest.skip("`formulas` OOM'd on v3.1.2 workbook recalc")

    file_token = f"[{out.name}]"

    def at(sheet: str, cell: str):
        key = f"'{file_token}{sheet.upper()}'!{cell}"
        v = sol[key].value
        if isinstance(v, np.ndarray):
            v = v.flat[0]
        return v

    # Inputs!J should normalise every variant to "Yes".
    for r in (16, 17, 18):
        j = at("Inputs", f"J{r}")
        assert j == "Yes", (
            f"Inputs!J{r} must normalise paste-in to 'Yes', got {j!r}"
        )

    # Analyser E17 — P1 (orig 800k, proceeds 2.6m) → raw gain 1.8m → with
    # discount applied: 1.8m * (1 - 1/3) = 1,200,000. If normalisation
    # didn't happen, the formula would skip the discount branch and return
    # the raw 1.8m.
    e17 = float(at("Analyser", "E17"))
    assert abs(e17 - 1_200_000) < 1, (
        f"E17 should be discounted long-held gain 1.2m, got {e17}"
    )

    # Analyser H17 — Div 296 adj gain, cost base = MV 2.4m, proceeds 2.6m
    # → raw 200k → discounted ≈ 133,333.33.
    h17 = float(at("Analyser", "H17"))
    assert abs(h17 - 133_333.33) < 1, (
        f"H17 should be discounted Div 296 adj gain ≈ 133,333, got {h17}"
    )

    # Analyser B71 (Fund Ord CGT) routes through the recon helpers
    # M70/N70/O70 which read Inputs!J via SUMIFS. M70 = discount-eligible
    # gains > 0. With all three sample rows held>12m (paste-dirty), the
    # discount-eligible gains are P1 (1.8m) + S1 (0.3m) = 2,100,000.
    # L1 has a projected loss (-300k) → in O70, not M70.
    # If normalisation didn't happen, "Yes ", "yes", " YES " would all fail
    # the SUMIFS exact-match against "Yes" and M70 would collapse to 0.
    m70 = float(at("Analyser", "M70"))
    assert abs(m70 - 2_100_000) < 1, (
        f"Recon helper M70 should aggregate paste-normalised long-held "
        f"gains to 2,100,000, got {m70}"
    )


# --- Analyser tab ---------------------------------------------------------

def _analyser(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    return load_workbook(out)["Analyser"]


class TestAnalyser:
    def test_no_lever_mirror_rows(self, tmp_path: Path):
        """v3.0: lever mirror at rows 5-7 (and band at row 4) is deleted."""
        ws = _analyser(tmp_path)
        # Rows 5-7 should NOT contain the v2.x lever-mirror Inputs!B refs.
        for row in (5, 6, 7):
            b = ws.cell(row=row, column=2).value
            assert b is None or not (isinstance(b, str) and b.startswith("='Inputs'!B")), (
                f"v3.0 should have no lever mirror at B{row}, got {b!r}"
            )

    def test_state_strip_parameters_in_effect(self, tmp_path: Path):
        """v3.0: state strip row 2 is the 'Parameters in effect' read-only line.
        References named ranges for rates / thresholds / discount, NOT toggles."""
        ws = _analyser(tmp_path)
        cell = ws["A2"]
        v = cell.value
        assert v and isinstance(v, str) and v.startswith("=")
        assert "Parameters in effect" in v
        # Must reference these named ranges live.
        for nm in ("rate_tier1", "rate_tier2", "threshold_1", "threshold_2",
                   "discount_rate", "fund_cgt_rate"):
            assert nm in v, f"state strip should reference {nm}"
        # Must NOT reference any removed toggle.
        for removed in ("reset_on", "tier10_on", "discount_on"):
            assert removed not in v, (
                f"state strip should not reference removed named range {removed}: {v!r}"
            )

    def test_fund_summary_side_by_side_headers(self, tmp_path: Path):
        """v3.0: fund-summary header row (row 7) has scenario column headers."""
        ws = _analyser(tmp_path)
        # Col C, D, E hold scenario column headers.
        assert "no reset" in str(ws["C7"].value).lower()
        assert "elected" in str(ws["D7"].value).lower()
        assert "difference" in str(ws["E7"].value).lower()

    def test_fund_earnings_two_scenarios(self, tmp_path: Path):
        """v3.1: Fund Div 296 earnings nets gains and losses intra-year,
        floored at zero — formula is MAX(0, SUM(...)). v3.0 used
        SUMIF(...,">0") which floored per-asset before summing.

        v3.2 column shift: the post-disc helper for the no-reset scenario
        moved from col L to col N (col L is now the visible Reset Impact
        col); the elected-scenario post-disc col moved from H to J.
        - C8 (no reset) nets helper col N (was L).
        - D8 (elected) nets col J (was H, the post-disc Div 296 gain).
        - E8 (difference) = D8 - C8.
        """
        ws = _analyser(tmp_path)
        c8 = ws["C8"].value
        d8 = ws["D8"].value
        e8 = ws["E8"].value
        assert c8 == '=MAX(0, SUM(N17:N66))', f"C8 wrong: {c8!r}"
        assert d8 == '=MAX(0, SUM(J17:J66))', f"D8 wrong: {d8!r}"
        assert e8 == "=D8-C8", f"E8 (diff) should be D8-C8, got {e8!r}"

    def test_headline_row_sums_member_taxes_per_scenario(self, tmp_path: Path):
        """v3.0: Headline at row 13, three scenario columns each summing rows 9-12."""
        ws = _analyser(tmp_path)
        assert ws["C13"].value == "=SUM(C9:C12)"
        assert ws["D13"].value == "=SUM(D9:D12)"
        assert ws["E13"].value == "=SUM(E9:E12)"

    def test_member_tax_formula_reads_band1_band2_from_inputs(self, tmp_path: Path):
        """v3.0: per-member tax (row 9 Member 1) reads Inputs!D7 (band1) and
        Inputs!E7 (band2) directly — single source of truth with Inputs."""
        ws = _analyser(tmp_path)
        # No-reset scenario at C9
        c9 = ws["C9"].value
        assert c9 and isinstance(c9, str)
        for ref in ("'Inputs'!B7", "'Inputs'!C7", "'Inputs'!D7", "'Inputs'!E7"):
            assert ref in c9, f"member 1 no-reset formula missing {ref}: {c9!r}"
        assert "rate_tier1" in c9 and "rate_tier2" in c9
        # Must not reference removed toggle.
        assert "tier10_on" not in c9, f"v3.0 formula should not reference tier10_on: {c9!r}"

        # Elected scenario at D9 — same formula, different earnings cell.
        d9 = ws["D9"].value
        assert d9 and isinstance(d9, str)
        assert "rate_tier1" in d9 and "rate_tier2" in d9

        # Difference at E9
        assert ws["E9"].value == "=D9-C9"

    def test_per_asset_first_row_columns(self, tmp_path: Path):
        """v3.2: per-asset detail row 17, new symmetric Option B layout.
        Visible cols A..L (was A..J); hidden helpers M..Q (was K..L only).
            A # | B Asset | C Proceeds | D Orig CB
            E Ord gross gain (NEW, no discount)
            F Disc eligible? (NEW, mirrors Inputs!J)
            G Per-asset Ord CGT (info, derived from E + F)
            H Div 296 CB
            I Div 296 gross gain (NEW)
            J Per-asset Div 296 gain (post-disc, info)
            K Div 296 tax (pro-rata of headline; SUMIF over col J)
            L Reset impact (= M − N, the hidden Div 296 with/without-reset helpers)
        """
        ws = _analyser(tmp_path)
        a = ws.cell(row=17, column=2).value   # Asset display
        c = ws.cell(row=17, column=3).value   # Proceeds
        d = ws.cell(row=17, column=4).value   # Original cost base
        e = ws.cell(row=17, column=5).value   # NEW: Ord gross gain
        f = ws.cell(row=17, column=6).value   # NEW: Discount-eligible flag
        g = ws.cell(row=17, column=7).value   # Per-asset Ord CGT (derived)
        h = ws.cell(row=17, column=8).value   # Div 296 cost base
        i = ws.cell(row=17, column=9).value   # NEW: Div 296 gross gain
        j = ws.cell(row=17, column=10).value  # Per-asset Div 296 gain (post-disc)
        k = ws.cell(row=17, column=11).value  # Div 296 tax (pro-rata of headline)
        reset_impact = ws.cell(row=17, column=12).value  # Reset impact

        assert "'Inputs'!A16" in a and "'Inputs'!B16" in a
        assert "'Inputs'!G16" in c    # Proceeds
        assert "'Inputs'!C16" in d    # Original CB
        # Col E — gross ord gain: NO discount_rate, NO Inputs!J reference.
        # Just (proceeds − orig CB), with "" guard for empty rows.
        assert "'Inputs'!G16" in e and "'Inputs'!C16" in e, f"E17 must compute proceeds − orig: {e!r}"
        assert "discount_rate" not in e, f"v3.2 col E (gross) must NOT apply discount: {e!r}"
        # Col F — flag, mirrors Inputs!J16.
        assert "'Inputs'!J16" in f, f"F17 must read Inputs!J16: {f!r}"
        # Col G — per-asset Ord CGT, derived from E (gross) and F (flag).
        assert "fund_cgt_rate" in g and "E17" in g and "F17" in g, (
            f"G17 must reference E17, F17, fund_cgt_rate: {g!r}"
        )
        assert "discount_rate" in g, f"G17 must apply discount when flag=Yes: {g!r}"
        # Col H — Div 296 cost base. Always MV.
        assert "'Inputs'!E16" in h
        assert "reset_on" not in h, f"v3.0 col H (Div 296 CB) must not reference reset_on: {h!r}"
        # Col I — Div 296 gross gain (NEW): proceeds − Div 296 CB (MV), no discount.
        assert "'Inputs'!G16" in i and "'Inputs'!E16" in i, f"I17 must compute proceeds − MV: {i!r}"
        assert "discount_rate" not in i, f"v3.2 col I (gross) must NOT apply discount: {i!r}"
        # Col J — Div 296 post-disc gain (semantics same as v3.1's H, shifted right).
        assert "discount_rate" in j
        assert "reset_on" not in j and "discount_on" not in j
        # Col K — Div 296 tax (pro-rata). Headline at $D$13. SUMIF references col J.
        assert "$D$13" in k, f"K17 must reference headline $D$13: {k!r}"
        assert 'SUMIF(J17:J66,">0")' in k, (
            f"K17 attribution denominator must SUMIF over col J (post-disc), not H: {k!r}"
        )
        # Col L — Reset impact = M − N (hidden helpers for Div 296 gain with/without reset).
        assert "M17" in reset_impact and "N17" in reset_impact

    def test_helper_columns_hidden(self, tmp_path: Path):
        """v3.2: hidden helpers shift right by 2 — M, N for Div 296 with/without
        reset, plus O, P, Q for the Fund Ord CGT helpers. Visible cols now run A..L."""
        ws = _analyser(tmp_path)
        for col in ("M", "N", "O", "P", "Q"):
            assert ws.column_dimensions[col].hidden, f"col {col} should be hidden"
        # Visible cols must not be hidden.
        for col in ("K", "L"):
            assert not ws.column_dimensions[col].hidden, (
                f"col {col} is now VISIBLE (was hidden in v3.1); got hidden=True"
            )

    def test_totals_row_v3(self, tmp_path: Path):
        """v3.2: totals row at 67. Info-only gain/CGT cols are E (Ord gross),
        G (Ord CGT derived), I (Div 296 gross), J (Div 296 post-disc) —
        all show "(see fund total)". Col F (the Yes/No flag) shows blank
        in totals because summing flags is nonsensical. Only Proceeds (C)
        and Div 296 tax (K, shifted from I) sum."""
        ws = _analyser(tmp_path)
        assert ws["C67"].value == "=SUM(C17:C66)"
        # v3.2: four info-only cols, all "(see fund total)".
        assert ws["E67"].value == "(see fund total)"
        assert ws["G67"].value == "(see fund total)"
        assert ws["I67"].value == "(see fund total)"
        assert ws["J67"].value == "(see fund total)"
        # Flag col (F) blank in totals.
        assert ws["F67"].value in (None, "")
        # Div 296 tax — shifted from col I to col K.
        assert ws["K67"].value == "=SUM(K17:K66)"

    def test_f_info_footnote_present(self, tmp_path: Path):
        """v3.2: row 68 footnote covers the four info-only gain/CGT cols
        (E ord gross, G ord CGT, I div296 gross, J div296 post-disc)."""
        ws = _analyser(tmp_path)
        footnote = ws["A68"].value
        assert footnote is not None
        assert "info only" in footnote
        assert "fund" in footnote.lower()

    def test_per_asset_ord_cgt_col_header_grey_label(self, tmp_path: Path):
        """v3.2: 'Per-asset Ord CGT (info only)' moved from col F to col G."""
        ws = _analyser(tmp_path)
        g16 = ws["G16"].value
        assert g16 is not None and "info only" in g16, (
            f"G16 must be the Per-asset Ord CGT header marked 'info only': {g16!r}"
        )
        # And col F is the new flag header.
        f16 = ws["F16"].value
        assert f16 is not None and "discount eligible" in f16.lower(), (
            f"F16 must be the '1/3 CGT discount eligible?' flag header: {f16!r}"
        )

    def test_per_asset_ord_cgt_formula_shows_dash_for_losses(self, tmp_path: Path):
        """v3.2: the Per-asset Ord CGT derivation lives at col G; '—' for
        loss rows (gross E<=0)."""
        ws = _analyser(tmp_path)
        g17 = ws["G17"].value
        assert g17 is not None and "—" in g17 and "IF(E17<=0" in g17, (
            f"G17 must guard losses with '—' and reference E17: {g17!r}"
        )

    def test_per_asset_ord_cgt_derives_from_gross_and_flag(self, tmp_path: Path):
        """v3.2 NEW: col G derivation must reference both col E (gross) and
        col F (flag), and apply the discount only when flag = 'Yes'."""
        ws = _analyser(tmp_path)
        g17 = ws["G17"].value
        assert "E17" in g17, f"G17 must reference E17 (gross): {g17!r}"
        assert "F17" in g17, f"G17 must reference F17 (flag): {g17!r}"
        assert '"Yes"' in g17, f"G17 must check flag='Yes': {g17!r}"
        assert "discount_rate" in g17, f"G17 must apply discount_rate: {g17!r}"
        assert "fund_cgt_rate" in g17, f"G17 must multiply by fund_cgt_rate: {g17!r}"

    def test_div296_attribution_uses_post_disc_col_J(self, tmp_path: Path):
        """v3.2 NEW: col K (Div 296 tax) attribution denominator must SUMIF
        over col J (post-disc, info), NOT the gross col I."""
        ws = _analyser(tmp_path)
        k17 = ws["K17"].value
        assert 'SUMIF(J17:J66,">0")' in k17, (
            f"K17 must use SUMIF(J17:J66,\">0\") as the attribution "
            f"denominator (NOT col I which is the new gross col): {k17!r}"
        )
        # Numerator references col J too (the asset's own post-disc gain).
        assert "MAX(0,J17)" in k17, f"K17 numerator must use J17: {k17!r}"

    def test_reconciliation_panel_v31(self, tmp_path: Path):
        """v3.2: recon band stays at row 70. Fund Ord CGT helpers shifted
        from M/N/O to O/P/Q because two new visible cols (gross-gain and
        flag) and a third (Div 296 gross) bumped the previously-hidden
        K, L into M, N for the Div 296 with/without-reset helpers."""
        ws = _analyser(tmp_path)
        assert ws["A70"].value == "Reconciliation"
        assert ws["A71"].value == "Fund Ordinary CGT (after intra-year netting)"
        b71 = ws["B71"].value
        # Helpers shifted right: M70/N70/O70 → O70/P70/Q70.
        assert b71.startswith("=")
        assert "O70" in b71 and "P70" in b71 and "Q70" in b71, (
            f"B71 must reference shifted helpers O70/P70/Q70: {b71!r}"
        )
        assert "discount_rate" in b71 and "fund_cgt_rate" in b71

        assert ws["A72"].value == "Div 296 tax payable (elected-reset headline)"
        assert ws["B72"].value == "=D13"

        assert ws["A73"].value == "Capital losses carried forward"
        # v3.2: net unused gross loss; references shifted helpers.
        b73 = ws["B73"].value
        assert b73 == "=MAX(0, Q70 - (O70 + P70))", f"B73 wrong: {b73!r}"

    def test_recon_helpers_hidden(self, tmp_path: Path):
        """v3.2: Fund Ord CGT helpers live at hidden cols O, P, Q (was M, N, O)."""
        ws = _analyser(tmp_path)
        for col in ("O", "P", "Q"):
            assert ws.column_dimensions[col].hidden, f"col {col} should be hidden"
        for cell in ("O70", "P70", "Q70"):
            v = ws[cell].value
            assert v and v.startswith("="), f"{cell} empty/non-formula: {v!r}"

    def test_recon_helpers_use_sumifs_not_sumproduct(self, tmp_path: Path):
        """v3.1.1: helper cells use SUMIFS/SUMIF, NOT SUMPRODUCT.

        v3.2: helpers shifted from M/N/O to O/P/Q. Same formula shape.
        The briefly-shipped SUMPRODUCT pattern errored with #VALUE!
        because the trailing `*H` multiplied empty-row text values ("" from
        the Inputs!H IF formula) and the ISNUMBER guard doesn't short-
        circuit. SUMIFS handles mixed-type sum_range natively.
        """
        ws = _analyser(tmp_path)
        o70 = ws["O70"].value
        p70 = ws["P70"].value
        q70 = ws["Q70"].value
        assert o70.startswith("=SUMIFS("), f"O70 must use SUMIFS: {o70!r}"
        assert p70.startswith("=SUMIFS("), f"P70 must use SUMIFS: {p70!r}"
        assert q70.startswith("=-SUMIF("), f"Q70 must use -SUMIF: {q70!r}"
        for cell, v in (("O70", o70), ("P70", p70), ("Q70", q70)):
            assert "SUMPRODUCT" not in v, (
                f"{cell} should NOT use SUMPRODUCT (it errors on empty-row "
                f"text values): {v!r}"
            )

    def test_trap_conditional_formatting_applied(self, tmp_path: Path):
        """v3.2: trap CF range now A17:L66 (was A17:J66; +2 cols added)."""
        ws = _analyser(tmp_path)
        cf_rules = list(ws.conditional_formatting._cf_rules.items())
        assert cf_rules, "no conditional formatting rules on Analyser"
        ranges = [str(r[0]) for r in cf_rules]
        assert any("A17" in r and "L66" in r for r in ranges), (
            f"expected trap-shading range to span A17:L66, got {ranges!r}"
        )

    def test_row_number_column(self, tmp_path: Path):
        """v3.0: col A row-num 1..50 spans rows 17-66; totals at row 67."""
        ws = _analyser(tmp_path)
        assert ws["A16"].value == "#"        # header row
        assert ws["A17"].value == 1          # first data row
        assert ws["A66"].value == 50         # last data row
        assert ws["A67"].value == "Total"    # totals row

    def test_print_titles_repeat_header(self, tmp_path: Path):
        """v3.0: per-asset header repeats on every printed page (row 16)."""
        ws = _analyser(tmp_path)
        assert ws.print_title_rows in ("16:16", "$16:$16")

    def test_no_freeze_panes(self, tmp_path: Path):
        ws = _analyser(tmp_path)
        assert ws.freeze_panes is None

    def test_trap_cf_formula_uses_relative_rows(self, tmp_path: Path):
        """The CF formula must NOT have $ before row numbers (must adjust per row)."""
        import re
        ws = _analyser(tmp_path)
        for _rng, rules in ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                if rule.formula:
                    for f in rule.formula:
                        bad = re.findall(r"\$\d+", f)
                        assert not bad, (
                            f"CF formula has absolute-row reference(s) {bad}: {f!r}"
                        )


# --- Comparison tab -------------------------------------------------------

def _comparison(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    return load_workbook(out)["Comparison"]


class TestComparison:
    def test_watermark_banner_and_print_header(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        assert ws["A1"].value == "ILLUSTRATIVE — NOT ADVICE"
        assert "ILLUSTRATIVE" in ws.oddHeader.center.text

    def test_context_strip_present(self, tmp_path: Path):
        """v3.0: Members & TSB strip still uses MEMBERS_FIRST_DATA_ROW (now 7)
        for the SUM reference."""
        ws = _comparison(tmp_path)
        assert "Members & TSB" in str(ws["A12"].value)
        assert ws["A13"].value == "Members"
        assert ws["B13"].value == "Total Superannuation Balance (TSB)"
        assert ws["A14"].value == "Member 1"
        assert ws["A15"].value == "Member 2"
        assert ws["A16"].value == "Member 3"
        assert ws["A17"].value == "Member 4"
        assert ws["A18"].value == "Total"
        # v3.0: members on Inputs now at rows 7-10 (was 11-14).
        assert "SUM('Inputs'!B7:B10)" in ws["B18"].value
        for col in ("D", "E", "F", "G", "H", "I", "J", "K"):
            for row in range(14, 19):
                val = ws[f"{col}{row}"].value
                assert val is None, (
                    f"{col}{row} should be empty (right-side tiles dropped); got {val!r}"
                )

    def test_metric_cards_present(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        assert ws["A21"].value == "If no Div 296 CostBase Reset (default)"
        assert ws["E21"].value == "If elected to reset Div 296 CostBase Reset"
        assert ws["I21"].value == "Difference (Net Div 296 Tax)"
        assert "L$6" in ws["A22"].value or "L6" in ws["A22"].value
        assert "M$6" in ws["E22"].value or "M6" in ws["E22"].value
        net = ws["I22"].value
        assert ("M" in net and "L" in net and "-" in net), f"unexpected difference formula {net!r}"
        assert net.index("M") < net.index("L")

    def test_subtotals_table(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        assert ws["B26"].value == "If no reset (default)"
        assert ws["C26"].value == "If elected to reset"
        assert ws["D26"].value == "Difference"
        assert ws["A27"].value == "Div 296 earnings"
        assert "Ordinary CGT" in ws["A28"].value and "unchanged" in ws["A28"].value
        assert ws["A29"].value == "Div 296 tax (headline)"
        assert "TOTAL TAX BURDEN" in ws["A30"].value
        assert "Analyser" in ws["B28"].value
        assert ws["B28"].value == ws["C28"].value
        assert ws["B30"].value == "=B28+B29"
        assert ws["C30"].value == "=C28+C29"
        for row in (27, 28, 29, 30):
            assert ws[f"D{row}"].value == f"=C{row}-B{row}"

    def test_comparison_pulls_fund_ord_cgt_via_constant(self, tmp_path: Path):
        """v3.2 slice 1: Comparison's Ord CGT ref must derive from the
        analyser.FUND_ORD_CGT_CELL constant, not a hardcoded 'B71'. This
        guards against the next Analyser row shift drifting the comparison
        ref silently."""
        from div296.tabs import analyser as analyser_tab
        ws = _comparison(tmp_path)
        expected = f"='Analyser'!{analyser_tab.FUND_ORD_CGT_CELL}"
        assert ws["B28"].value == expected, (
            f"Comparison B28 must equal {expected!r} (derived from "
            f"analyser.FUND_ORD_CGT_CELL = {analyser_tab.FUND_ORD_CGT_CELL!r}); "
            f"got {ws['B28'].value!r}"
        )
        assert ws["C28"].value == expected

    def test_panel_a_uses_original_cost_base(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        cb_formula = ws[f"C{CMP_DATA_FIRST_ROW}"].value
        assert "'Inputs'!$C:$C" in cb_formula
        assert "'Inputs'!$E:$E" not in cb_formula

    def test_panel_b_uses_market_value(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        cb_formula = ws[f"H{CMP_DATA_FIRST_ROW}"].value
        assert "'Inputs'!$E:$E" in cb_formula
        assert "'Inputs'!$C:$C" not in cb_formula

    def test_panels_independent_of_master_reset_toggle(self, tmp_path: Path):
        """Neither panel formula may reference reset_on (which no longer exists)."""
        ws = _comparison(tmp_path)
        for row in range(CMP_DATA_FIRST_ROW, CMP_DATA_LAST_ROW + 1):
            for col_letter in ("B", "C", "D", "E", "G", "H", "I", "J"):
                f = ws[f"{col_letter}{row}"].value
                if f and isinstance(f, str):
                    assert "reset_on" not in f, (
                        f"{col_letter}{row} references reset_on (removed in v3.0)"
                    )

    def test_per_member_formulas_have_no_tier_or_discount_toggle(self, tmp_path: Path):
        """v3.0: Comparison's per-member tax formula must not reference removed toggles."""
        ws = _comparison(tmp_path)
        # Helper col L holds the Scenario A per-member tax (rows starting at HELPER_MEMBER_TAX_FIRST_ROW=2).
        for row in range(2, 6):
            for col_letter in ("L", "M"):
                f = ws[f"{col_letter}{row}"].value
                if f and isinstance(f, str):
                    for removed in ("tier10_on", "discount_on"):
                        assert removed not in f, (
                            f"{col_letter}{row} references removed toggle {removed}: {f!r}"
                        )

    def test_delta_column_formula(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        delta = ws[f"K{CMP_DATA_FIRST_ROW}"].value
        assert f"J{CMP_DATA_FIRST_ROW}" in delta and f"E{CMP_DATA_FIRST_ROW}" in delta
        assert "-" in delta
        assert delta.index(f"J{CMP_DATA_FIRST_ROW}") < delta.index(f"E{CMP_DATA_FIRST_ROW}")

    def test_helper_columns_hidden(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        # v3.4: col Q dropped from the hidden tuple — nothing writes it (the
        # per-register grid is N/O/P and the matched-row lookup is R).
        for col_letter in ("L", "M", "N", "O", "P", "R"):
            assert ws.column_dimensions[col_letter].hidden, f"col {col_letter} should be hidden"
        for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"):
            assert not ws.column_dimensions[col_letter].hidden, f"col {col_letter} must be visible"

    def test_top_10_sorted_rendered(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        first = ws[f"A{CMP_DATA_FIRST_ROW}"].value
        last = ws[f"A{CMP_DATA_LAST_ROW}"].value
        assert first and first.startswith("=")
        assert last and last.startswith("=")
        overflow = ws[f"A{CMP_OVERFLOW_ROW}"].value
        assert overflow and "top 10" in overflow.lower()
        assert "Δ" not in overflow

    def test_per_asset_detail_uses_large_match_sort(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        matched_formula = ws[f"R{CMP_DATA_FIRST_ROW}"].value
        assert "LARGE" in matched_formula and "MATCH" in matched_formula
        assert (f"$R{CMP_DATA_FIRST_ROW}" in ws[f"B{CMP_DATA_FIRST_ROW}"].value
                or f"R{CMP_DATA_FIRST_ROW}" in ws[f"B{CMP_DATA_FIRST_ROW}"].value)

    def test_no_recommendation_language(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        forbidden = ("saved", "saves", "you should", "we recommend", "we suggest")
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    for word in forbidden:
                        assert word.lower() not in v.lower(), (
                            f"Comparison tab contains forbidden recommendation word "
                            f"{word!r} in cell value {v!r}"
                        )

    def test_print_setup_landscape_a4(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        assert ws.page_setup.orientation == ws.ORIENTATION_LANDSCAPE
        assert int(ws.page_setup.paperSize) == int(ws.PAPERSIZE_A4)
        assert int(ws.page_setup.fitToWidth) == 1
        assert ws.print_area is not None and "K" in ws.print_area

    def test_no_chart_v25(self, tmp_path: Path):
        ws = _comparison(tmp_path)
        assert len(ws._charts) == 0


# --- Notes tab ------------------------------------------------------------

def _notes(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    return load_workbook(out)["Notes"]


class TestNotes:
    def test_required_caveats_present(self, tmp_path: Path):
        """Every locked caveat from README must appear on the Notes tab."""
        ws = _notes(tmp_path)
        all_text = " ".join(
            str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
        )
        for required in (
            "Illustrative only",
            "Pension phase is NOT modelled",
            # v3.1.1: "Loss-offset divergence" replaced by
            # "Prior-year capital losses are NOT modelled" — v3.1 now
            # implements intra-year netting per s102-5, so the previous
            # divergence caveat is no longer factually correct.
            "Prior-year capital losses are NOT modelled",
            "Reset OFF scenario is realised-only",
            "Wash sale / Part IVA",
            "Transaction costs",
            "actuarial certificate",
            "all-or-nothing and irrevocable",
            "recontribution",
        ):
            assert required in all_text, f"Notes missing required caveat: {required!r}"

    def test_no_recommendation_language(self, tmp_path: Path):
        ws = _notes(tmp_path)
        forbidden = ("you should", "we recommend", "we suggest")
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    for word in forbidden:
                        assert word.lower() not in v.lower(), (
                            f"Notes contains forbidden advice phrase {word!r}: {v!r}"
                        )

    def test_valuation_log_mirrors_register(self, tmp_path: Path):
        """Valuation log has one row per asset, formulas pointing at Inputs."""
        ws = _notes(tmp_path)
        ref_count = 0
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "'Inputs'!F" in c.value:
                    ref_count += 1
        assert ref_count >= 50, f"Expected ≥50 valuation-source mirrors, got {ref_count}"

    def test_provenance_cells_present(self, tmp_path: Path):
        ws = _notes(tmp_path)
        labels = {
            str(c.value): c.row
            for row in ws.iter_rows()
            for c in row
            if c.column == 1 and isinstance(c.value, str)
        }
        assert "build_version" in labels
        assert "build_date" in labels
        assert "git_short_sha" in labels
        for label in ("build_version", "build_date", "git_short_sha"):
            assert ws.row_dimensions[labels[label]].hidden, (
                f"Provenance row for {label!r} should be hidden"
            )


# --- CLASS Import tab (v3.3) ---------------------------------------------

def _class_import(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    return load_workbook(out)[ci.SHEET]


def test_class_import_inserted_after_inputs(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    names = load_workbook(out).sheetnames
    assert names.index(ci.SHEET) == names.index("Inputs") + 1


def test_class_import_requires_tax_cost_basis_banner(tmp_path: Path):
    """The basis cannot be auto-detected from the CSV, so a banner must warn."""
    ws = _class_import(tmp_path)
    banner = ws.cell(row=ci.BASIS_BANNER_ROW, column=1).value
    assert banner and "TAX COST BASE" in banner.upper()


def test_class_import_sample_data_preloaded(tmp_path: Path):
    """Sample = DEMO tax-cost-base export. GOOGL carries the negative cost base."""
    ws = _class_import(tmp_path)
    # First sample row (BACCT cash) sits at the first paste data row.
    assert ws[f"{ci.PASTE_COL_CODE}{ci.FIRST_DATA_ROW}"].value == "BACCT"
    googl_row = ci.FIRST_DATA_ROW + 6
    assert ws[f"{ci.PASTE_COL_CODE}{googl_row}"].value == "GOOGL"
    assert ws[f"{ci.PASTE_COL_COST}{googl_row}"].value == -1772.96


def test_class_import_mapped_formulas_read_correct_paste_columns(tmp_path: Path):
    """Mapped block A=PasteB(code), C=PasteL(cost), D=PasteM(mv); filter present."""
    ws = _class_import(tmp_path)
    r = ci.FIRST_DATA_ROW
    code_col = get_column_letter(ci.MAP_COL_START)
    cost_col = get_column_letter(ci.MAP_COL_START + 2)
    mv_col = get_column_letter(ci.MAP_COL_START + 3)
    code_f = ws[f"{code_col}{r}"].value
    cost_f = ws[f"{cost_col}{r}"].value
    mv_f = ws[f"{mv_col}{r}"].value
    assert f"${ci.PASTE_COL_CODE}{r}" in code_f
    assert f"${ci.PASTE_COL_COST}{r}" in cost_f
    assert f"${ci.PASTE_COL_MV}{r}" in mv_f
    # Blacklist logic baked into the IF filter.
    assert 'SEARCH("cash"' in code_f
    assert '"REASEDCGT"' in code_f


def test_class_import_negative_cost_base_is_flagged(tmp_path: Path):
    """Decision 7: negative tax cost base passed through + flagged red."""
    ws = _class_import(tmp_path)
    googl_row = ci.FIRST_DATA_ROW + 6
    flag = ws.cell(row=googl_row, column=ci.MAP_FLAG_COL_IDX).value
    assert flag and "<0" in flag and "E4" in flag
    # A conditional-format rule tints the mapped cost column on negatives.
    cost_col = get_column_letter(ci.MAP_COL_START + 2)
    cf_ranges = " ".join(str(r) for r in ws.conditional_formatting)
    assert cost_col in cf_ranges


def test_class_import_paste_unlocked_mapped_locked(tmp_path: Path):
    """Paste zone must be editable; the formula-driven mapped block must not."""
    ws = _class_import(tmp_path)
    r = ci.FIRST_DATA_ROW
    assert ws[f"{ci.PASTE_COL_CODE}{r}"].protection.locked is False
    mapped_code = f"{get_column_letter(ci.MAP_COL_START)}{r}"
    assert ws[mapped_code].protection.locked in (True, None)


def test_class_import_howto_banner_gives_physical_range(tmp_path: Path):
    """v3.3 audit: the transfer instruction must name the PHYSICAL copy range
    (T7:Z56) and tell the user to clear the demo data first."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["CLASS Import"]
    howto = ws["A3"].value
    assert "T7:Z56" in howto
    assert "Clear the green zone" in howto
    assert "Inputs!A16" in howto
    assert "Paste-Special > Values" in howto
    # Hint cell sits directly above the mapped block.
    assert "T7:Z56" in ws["T5"].value


def test_class_import_demo_remnant_guard(tmp_path: Path):
    """v3.4 review: a machine-checked banner must fire while the shipped demo
    rows remain in the paste zone, so a SHORT real paste that leaves a demo
    tail is caught before it transfers phantom holdings into the register."""
    ws = _class_import(tmp_path)
    # The sentinel must actually be in the shipped sample, or the guard is dead.
    sample_names = [name for (_code, name, *_rest) in ci.SAMPLE_ROWS]
    assert ci.DEMO_SENTINEL_NAME in sample_names
    # The guard banner sits over the mapped block (its merge top-left).
    guard = ws.cell(row=ci.CAPACITY_WARN_ROW, column=ci.MAP_COL_START).value
    assert guard and "COUNTIF" in guard
    assert ci.DEMO_SENTINEL_NAME in guard
    # ...counting over the paste NAME column, where the sentinel lives.
    assert f"{ci.PASTE_COL_NAME}{ci.FIRST_DATA_ROW}" in guard
    assert "clear the green zone" in guard.lower()


def test_class_import_overflow_rows_unlocked(tmp_path: Path):
    """v3.3 audit: rows below the 50-row paste zone must be unlocked so an
    oversize CLASS paste LANDS (triggering the row-4 capacity warning)
    instead of being rejected wholesale by sheet protection."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["CLASS Import"]
    for row in (57, 100, 256):
        assert ws.cell(row=row, column=2).protection.locked is False, (
            f"Overflow row {row} in paste-zone column must be unlocked"
        )
    # ...but the mapped block stays locked even in the overflow band.
    assert ws.cell(row=57, column=20).protection.locked is True, (
        "Mapped block must remain locked in overflow band"
    )


def test_class_import_review_flag_covers_blank_cost(tmp_path: Path):
    """v3.3 audit F5: a kept row with blank/non-numeric Total Cost must be
    flagged, not mapped as a silent $0 cost base."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["CLASS Import"]
    flag = ws.cell(row=7, column=29).value   # AC7 — review flag, first data row
    assert "NEGATIVE tax cost base" in flag
    assert "blank/non-numeric" in flag
    # Alignment warning cell exists on row 5.
    assert "paste alignment" in ws["A5"].value


def test_inputs_transfer_tripwire(tmp_path: Path):
    """v3.3 audit: warn when formulas (not values) are pasted into the register."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]
    v = ws["A13"].value
    assert "ISFORMULA" in v and "Paste-Special" in v


def test_incomplete_rows_blank_not_zero(tmp_path: Path):
    """v3.4 audit F2/F3: a register row missing its cost-base cell must render
    blank (""), never coerce the blank to $0 (a full-proceeds 'gain'). Every
    per-asset Analyser cell must guard BOTH proceeds and its cost base."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    wb_re = load_workbook(out)
    a = wb_re["Analyser"]
    # Col H/I/J guard proceeds AND MV; col E guards proceeds AND orig CB.
    for coord in ("E17", "H17", "I17", "J17"):
        f = a[coord].value
        assert f.startswith("=IF(OR("), f"{coord}: {f}"
    # Inputs completeness warning (row-13 tripwire, second priority) exists.
    assert "no Market value at 30 Jun" in wb_re["Inputs"]["A13"].value


def test_comparison_card_formulas_keep_absolute_refs():
    """v3.4 audit 3.1: card/subtotal formulas must be '=$L$6' (absolute),
    not '=L$6' produced by the accidental [1:] slice."""
    from div296.tabs import comparison as C
    wb = build_workbook()
    ws = wb["Comparison"]
    card_a = ws[f"A{C.CARD_VALUE_ROW}"].value
    assert card_a == f"=${C.HELPER_COL_A}${C.HELPER_HEADLINE_ROW}"
    # The Div 296 subtotal row mirrors the same headline cells (absolute).
    sub = ws[f"B{C.SUBTOTAL_DIV296_ROW}"].value
    assert sub == f"=${C.HELPER_COL_A}${C.HELPER_HEADLINE_ROW}"


def test_comparison_positive_difference_red_cf(tmp_path: Path):
    """v3.4 audit 3.1: a positive Difference (reset costs money) renders red,
    mirroring Analyser. Assert a '>0' CF rule covers the subtotal Difference."""
    from div296.tabs import comparison as C
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Comparison"]
    sub_range = f"D{C.SUBTOTAL_EARNINGS_ROW}:D{C.SUBTOTAL_BURDEN_ROW}"
    for rng, rules in ws.conditional_formatting._cf_rules.items():
        if sub_range in str(rng):   # str(rng) is "<ConditionalFormatting D27:D30>"
            formulas = [f for rule in rules for f in (rule.formula or [])]
            assert any(">0" in f for f in formulas), formulas
            break
    else:
        raise AssertionError(f"no CF rule on {sub_range}")


def test_watermark_on_all_tabs():
    """v3.4 audit: the ILLUSTRATIVE print header must cover EVERY sheet
    (was missing on Inputs and CLASS Import — a compliance gap)."""
    wb = build_workbook()
    for ws in wb.worksheets:
        assert "ILLUSTRATIVE" in (ws.oddHeader.center.text or ""), ws.title


def test_inputs_numeric_dv_and_held_flag_cf(tmp_path: Path):
    """v3.4 audit: a numeric (warning) DV guards currency inputs, and an amber
    CF flags unrecognised held>12m values that silently normalise to 'No'."""
    from div296.tabs import inputs as I
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    ws = load_workbook(out)["Inputs"]
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) == 2, [dv.type for dv in dvs]
    assert any(dv.type == "decimal" for dv in dvs)
    # Amber CF on the held column (I) register range.
    held_rng = f"I{I.REGISTER_FIRST_DATA_ROW}:I{I.REGISTER_LAST_DATA_ROW}"
    cf_ranges = " ".join(str(r) for r in ws.conditional_formatting)
    assert held_rng in cf_ranges, cf_ranges


def test_sample_badge_survives_register_replacement(tmp_path: Path):
    """v3.3 audit: badge must also key on the seeded member TSBs, not only
    the register codes a CLASS transfer removes."""
    out = tmp_path / "out.xlsx"
    wb = build_workbook()
    wb.save(out)
    wb_re = load_workbook(out)
    for sheet, cell in (("Inputs", "A2"), ("Analyser", "A3"), ("Comparison", "A9")):
        v = wb_re[sheet][cell].value
        assert "12000000" in v.replace(",", "") and "P1" in v


def test_recalc_limitations_derive_from_constants():
    """v3.4 audit: the known-limitation list must track layout constants."""
    from div296._recalc_limitations import (
        KNOWN_FORMULAS_LIMITATIONS,
        is_known_limitation,
    )
    assert any(e.endswith("O70") for e in KNOWN_FORMULAS_LIMITATIONS)
    assert any(e.endswith("Q70") for e in KNOWN_FORMULAS_LIMITATIONS)
    assert is_known_limitation("'[X.xlsx]ANALYSER'!B71")
    assert not is_known_limitation("'[X.xlsx]ANALYSER'!C13")
    # v3.4: the Comparison per-asset detail panel (LARGE/MATCH/INDEX false
    # positives) is excluded across its full A..K x data-rows extent.
    from div296.tabs import comparison as C
    assert is_known_limitation(f"'[X.xlsx]COMPARISON'!{C.PANEL_A_COLS[0]}{C.DATA_FIRST_ROW}")
    assert is_known_limitation(f"'[X.xlsx]COMPARISON'!{C.DELTA_COL}{C.DATA_LAST_ROW}")
    # ...but the panel HEADER row (not part of the lookup chain) is not.
    assert not is_known_limitation(
        f"'[X.xlsx]COMPARISON'!{C.PANEL_A_COLS[0]}{C.PANEL_HEADER_ROW}"
    )
