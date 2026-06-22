"""Adviser edition (--edition adviser) — omits CLASS Import staging tab.

For advisers who don't use CLASS Super as their administration software.
The full edition must remain unchanged; the adviser edition drops only the
CLASS Import sheet and rewrites the two user-facing strings that name CLASS
(Inputs trip-banner text and the two CLASS-specific Notes entries).
"""

from pathlib import Path

from openpyxl import load_workbook

from div296 import build as build_mod
from div296.build import build_workbook
from div296.tabs.inputs import TRANSFER_CHECK_ROW


FULL_TABS = ["Inputs", "CLASS Import", "Analyser", "Comparison", "Notes"]
ADVISER_TABS = ["Inputs", "Analyser", "Comparison", "Notes"]


def _save(wb, tmp_path: Path, name: str) -> Path:
    out = tmp_path / name
    wb.save(out)
    return out


def test_full_edition_default_keeps_class_import(tmp_path: Path):
    wb = build_workbook()
    assert wb.sheetnames == FULL_TABS
    reopened = load_workbook(_save(wb, tmp_path, "full.xlsx"))
    assert reopened.sheetnames == FULL_TABS


def test_adviser_edition_omits_class_import(tmp_path: Path):
    wb = build_workbook(include_class_import=False)
    assert wb.sheetnames == ADVISER_TABS
    reopened = load_workbook(_save(wb, tmp_path, "adv.xlsx"))
    assert reopened.sheetnames == ADVISER_TABS
    assert "CLASS Import" not in reopened.sheetnames


def test_adviser_inputs_trip_banner_drops_class_mention(tmp_path: Path):
    wb = build_workbook(include_class_import=False)
    ws = wb["Inputs"]
    cell_value = ws.cell(row=TRANSFER_CHECK_ROW, column=1).value or ""
    assert "CLASS Import" not in cell_value, (
        "Adviser-edition Inputs trip-banner must not name CLASS Import"
    )
    assert "Paste-Special" in cell_value, "underlying paste-safety hint must remain"


def test_full_inputs_trip_banner_still_names_class(tmp_path: Path):
    wb = build_workbook(include_class_import=True)
    ws = wb["Inputs"]
    cell_value = ws.cell(row=TRANSFER_CHECK_ROW, column=1).value or ""
    assert "CLASS Import" in cell_value, (
        "Full-edition trip-banner must still reference CLASS Import (regression guard)"
    )


def test_adviser_notes_drops_class_specific_entries(tmp_path: Path):
    wb = build_workbook(include_class_import=False)
    ws = wb["Notes"]
    seen = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert not any("CLASS Import" in (s or "") for s in seen), (
        f"Adviser-edition Notes still mentions CLASS Import in: "
        f"{[s for s in seen if 'CLASS Import' in (s or '')]}"
    )


def test_full_notes_still_has_class_entries(tmp_path: Path):
    wb = build_workbook(include_class_import=True)
    ws = wb["Notes"]
    seen = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert any("CLASS Import" in (s or "") for s in seen), (
        "Full-edition Notes must keep the CLASS Import entries (regression guard)"
    )


def test_cli_adviser_writes_suffixed_filename(tmp_path: Path):
    out = tmp_path / "x.xlsx"
    rc = build_mod.main(["--edition", "adviser", "--no-validate", "--output", str(out)])
    assert rc == 0
    reopened = load_workbook(out)
    assert reopened.sheetnames == ADVISER_TABS


def test_cli_default_filename_suffix_only_for_adviser(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(build_mod, "_dist_dir", lambda: tmp_path)
    assert build_mod.main(["--no-validate"]) == 0
    assert build_mod.main(["--edition", "adviser", "--no-validate"]) == 0
    names = sorted(p.name for p in tmp_path.glob("*.xlsx"))
    assert any("_Adviser_Edition" in n for n in names), f"missing adviser suffix in {names}"
    assert any("_Adviser_Edition" not in n for n in names), f"missing plain full filename in {names}"
