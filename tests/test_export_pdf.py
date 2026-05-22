"""Unit tests for scripts/export_pdf.py.

LibreOffice isn't required: we only test the pure-Python helpers
(`_isolate_tab`, `_soffice_executable`, CLI surface). The actual PDF
conversion is exercised manually when LibreOffice is installed.
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from div296.build import build_workbook


@pytest.fixture(scope="module")
def export_pdf_module():
    """Load scripts/export_pdf.py as a module (it lives outside `src/`)."""
    repo_root = Path(__file__).resolve().parents[1]
    script_path = repo_root / "scripts" / "export_pdf.py"
    spec = importlib.util.spec_from_file_location("export_pdf", script_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules["export_pdf"] = module
    spec.loader.exec_module(module)
    return module


@pytest.fixture(scope="module")
def built_xlsx(tmp_path_factory) -> Path:
    out = tmp_path_factory.mktemp("export_pdf") / "model.xlsx"
    wb = build_workbook()
    wb.save(out)
    return out


def test_isolate_tab_hides_others(export_pdf_module, built_xlsx, tmp_path):
    staged = export_pdf_module._isolate_tab(built_xlsx, tmp_path, "Comparison")
    wb = load_workbook(staged)
    assert wb["Comparison"].sheet_state == "visible"
    for hidden in ("Inputs", "Analyser", "Notes"):
        assert wb[hidden].sheet_state == "hidden", f"{hidden} should be hidden"
    assert wb.active.title == "Comparison"


def test_isolate_tab_none_keeps_all_visible(export_pdf_module, built_xlsx, tmp_path):
    staged = export_pdf_module._isolate_tab(built_xlsx, tmp_path, None)
    wb = load_workbook(staged)
    for name in ("Inputs", "Analyser", "Comparison", "Notes"):
        assert wb[name].sheet_state == "visible"


def test_isolate_tab_rejects_unknown(export_pdf_module, built_xlsx, tmp_path):
    with pytest.raises(SystemExit, match="not found"):
        export_pdf_module._isolate_tab(built_xlsx, tmp_path, "NoSuchTab")


def test_main_errors_clearly_when_input_missing(export_pdf_module, tmp_path, capsys):
    # If soffice isn't installed this returns 1 with a "LibreOffice not found"
    # error; if it IS installed we'll get 1 with "input file does not exist".
    rc = export_pdf_module.main([str(tmp_path / "does_not_exist.xlsx")])
    assert rc == 1
    captured = capsys.readouterr()
    combined = captured.out + captured.err
    assert ("does not exist" in combined) or ("LibreOffice not found" in combined)
